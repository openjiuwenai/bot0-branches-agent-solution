"""FastAPI Web 服务 —— 化工炼化物料平衡数据导入、累积与导出。

每个装置模块（2#加氢裂化、DCC催化裂解）支持天级/小时级两种粒度，
导入时自动检测实际粒度并路由到对应存储；支持按时间段清除数据。
"""
from __future__ import annotations

import io
import json
import os
import tempfile
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List
from urllib.parse import quote

import pandas as pd
from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request

from . import data_store
from . import analyzer
from .validator import (_prepare_long_df, _prepare_long_df_full, _pivot_to_wide, field_structure_from_prep,
                          _classify_columns, _val_to_jsonable)

BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
TEMPLATE_DIR = BASE_DIR / "templates"

app = FastAPI(title="化工炼化物料平衡数据", version="0.3.0")


class NoCacheMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        if request.url.path.startswith("/static") or request.url.path in ("/", "/analysis", "/tank-analysis"):
            response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        return response


app.add_middleware(NoCacheMiddleware)
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
templates = Jinja2Templates(directory=str(TEMPLATE_DIR))


@app.get("/")
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/api/sources")
async def api_sources():
    """列出所有装置模块及其天级/小时级存储摘要。"""
    return {"modules": [data_store.get_module_summary(m["id"]) for m in data_store.MODULES]}


@app.post("/api/modules/add")
async def api_add_module(name: str = Form(...)):
    """添加新的数据类型（装置模块）。"""
    try:
        m = data_store.add_module(name)
        return {"ok": True, "module": m, "modules": [data_store.get_module_summary(mm["id"]) for mm in data_store.MODULES]}
    except ValueError as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/modules/rename")
async def api_rename_module(module_id: str = Form(...), name: str = Form(...)):
    """重命名已有数据类型（仅改显示名，不影响已导入数据）。"""
    try:
        m = data_store.rename_module(module_id, name)
        return {"ok": True, "module": m, "modules": [data_store.get_module_summary(mm["id"]) for mm in data_store.MODULES]}
    except ValueError as e:
        return {"ok": False, "error": str(e)}


def _granularity_warning(granularity, actual_gran):
    if granularity == actual_gran:
        return None
    g_sel = data_store.granularity_name(granularity)
    g_act = data_store.granularity_name(actual_gran)
    return (f"您选择的是「{g_sel}」，但检测到数据实际为「{g_act}」，"
            f"已自动导入到{g_act}数据库。")

@app.post("/api/import/sheets")
async def api_import_sheets(file: UploadFile = File(...)):
    """Read Excel file sheet name list."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    suffix = Path(file.filename).suffix.lower()
    if suffix not in (".xlsx", ".xls"):
        raise HTTPException(status_code=400, detail="Only .xlsx / .xls supported")
    tmp_path = None
    try:
        content = await file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        import pandas as _pd
        xl = _pd.ExcelFile(tmp_path)
        sheets = xl.sheet_names
        return {"sheets": sheets}
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=400, detail=f"Failed to read sheet list: {e}")
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass




@app.post("/api/import")
async def api_import(file: UploadFile = File(...),
                     module_id: str = Form(...),
                     granularity: str = Form(...),
                     sheet_name: str = Form(default="")):
    """导入 Excel（第一步）：解析并比对字段结构模板。

    - 无模板（首次导入）：返回 needs_confirmation=True 及字段清单，等待确认。
    - 有模板且结构一致：按模板确认的字段直接导入并返回宽表预览。
    - 有模板但结构不一致：返回 needs_confirmation=True 及差异，等待确认。
    """
    if module_id not in data_store.module_ids():
        raise HTTPException(status_code=400, detail="未知装置模块")
    if granularity not in ("daily", "hourly"):
        raise HTTPException(status_code=400, detail="粒度须为 daily 或 hourly")
    if not file.filename:
        raise HTTPException(status_code=400, detail="未提供文件")
    suffix = Path(file.filename).suffix.lower()
    if suffix not in (".xlsx", ".xls"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 格式")

    tmp_path = None
    try:
        content = await file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        prep = _prepare_long_df(tmp_path, sheet_name=sheet_name if sheet_name else None)
        if prep.errors:
            raise HTTPException(status_code=400, detail=";".join(prep.errors))

        # Even if time/flow columns can't be auto-detected, return field structure
        # so the user can manually configure them in the field adjustment modal.
        auto_detect_warning = None
        if prep.time_col is None or prep.flow_col is None:
            auto_detect_warning = "Unable to auto-detect time/value column. Please set manually in field settings."

        if prep.time_col is not None:
            actual_gran = data_store.detect_granularity(prep.df, prep.time_col)
        else:
            actual_gran = granularity
        actual_sid = data_store.granularity_source_id(module_id, actual_gran)
        mismatch = (granularity != actual_gran)
        gran_warn = _granularity_warning(granularity, actual_gran)

        fs = field_structure_from_prep(prep)
        incoming_all = fs["field_order"]
        template = data_store.load_template(actual_sid)

        if template is None:
            # 首次导入：等待用户确认字段结构
            return JSONResponse({
                "needs_confirmation": True, "is_first": True, "has_template": False,
                "module_id": module_id, "module_name": data_store.module_name(module_id),
                "actual_granularity": actual_gran, "actual_source_id": actual_sid,
                "granularity_mismatch": mismatch, "granularity_warning": gran_warn,
                "fields": fs["fields"], "field_order": incoming_all,
                "time_col": fs["time_col"], "flow_col": fs["flow_col"],
                "auto_detect_warning": auto_detect_warning,
                "value_cols": fs["value_cols"], "adjustments": prep.adjustments,
                "diff": None,
                "selected_sheet": prep.sheet_name or sheet_name or "",
            })

        # 比对基于模板确认的导入字段(included_fields)，而非全部字段。
        # 额外列(如序号)不影响匹配，只要模板需要的字段都存在即可直接导入。
        tpl_included = template.get("included_fields", []) or template.get("all_fields", [])
        diff = data_store.compare_field_structures(tpl_included, [c for c in tpl_included if c in incoming_all])

        def _norm_name(s):
            """规范化字段名：去空格，统一 _extracted_date 为日期，用于模糊匹配。"""
            s = str(s).replace(" ", "").replace("　", "").strip()
            if s == "_extracted_date":
                return "日期"
            return s

        # 构建规范化名 -> 原始名的映射（incoming）
        norm_to_incoming = {}
        for c in incoming_all:
            norm_to_incoming.setdefault(_norm_name(c), c)
        # 构建规范化名 -> 模板名的映射
        norm_to_tpl = {}
        for c in tpl_included:
            norm_to_tpl.setdefault(_norm_name(c), c)

        # 仅当模板字段缺失时才认为不匹配；额外字段不触发确认
        missing = [c for c in tpl_included if _norm_name(c) not in norm_to_incoming]
        extra = [c for c in incoming_all if _norm_name(c) not in norm_to_tpl]
        diff["missing"] = missing
        diff["extra"] = extra
        diff["match"] = (not missing)
        if diff["match"]:
            # 结构一致：按模板确认的字段直接导入（套用模板排除的计算列等）
            # 构建 模板字段名 -> 实际列名 的映射（基于规范化匹配）
            tpl_to_actual = {}
            for tc in template.get("included_fields", []):
                nc = _norm_name(tc)
                if nc in norm_to_incoming:
                    tpl_to_actual[tc] = norm_to_incoming[nc]
            included_tpl = [c for c in template.get("included_fields", []) if c in tpl_to_actual]
            included_actual = [tpl_to_actual[c] for c in included_tpl]
            if prep.time_col and prep.time_col not in included_actual:
                included_actual = [prep.time_col] + included_actual
            t_value_cols_tpl = [c for c in template.get("value_cols", []) if c in tpl_to_actual]
            t_value_cols_actual = [tpl_to_actual[c] for c in t_value_cols_tpl]
            # 重命名为模板字段名，保持与首次导入一致
            rename_map = {tpl_to_actual[tc]: tc for tc in included_tpl if tpl_to_actual[tc] != tc}
            # 重新读取完整数据（_prepare_long_df 仅读前 200 行用于快速检测）
            prep_full = _prepare_long_df_full(tmp_path, sheet_name=sheet_name if sheet_name else None)
            filtered = prep_full.df[included_actual].copy()
            if rename_map:
                filtered = filtered.rename(columns=rename_map)
            final_included = [rename_map.get(c, c) for c in included_actual]
            # 更新 time_col 为重命名后的名称
            final_time_col = rename_map.get(prep.time_col, prep.time_col)
            data_store.save_import(actual_sid, filtered, final_time_col, prep.flow_col,
                                   prep.mapping, final_included, t_value_cols_tpl)
            preview = data_store.build_wide_preview(actual_sid)
            preview["adjustments"] = prep.adjustments
            preview["action"] = "import"
            preview["granularity_mismatch"] = mismatch
            preview["actual_granularity"] = actual_gran
            preview["granularity_warning"] = gran_warn
            preview["needs_confirmation"] = False
            preview["template_matched"] = True
            preview["selected_sheet"] = prep.sheet_name or sheet_name or ""
            return JSONResponse(preview)

        # 结构不一致：返回差异，等待用户确认
        return JSONResponse({
            "needs_confirmation": True, "is_first": False, "has_template": True,
            "module_id": module_id, "module_name": data_store.module_name(module_id),
            "actual_granularity": actual_gran, "actual_source_id": actual_sid,
            "granularity_mismatch": mismatch, "granularity_warning": gran_warn,
            "fields": fs["fields"], "field_order": incoming_all,
            "time_col": fs["time_col"], "flow_col": fs["flow_col"],
            "value_cols": fs["value_cols"], "adjustments": prep.adjustments,
            "diff": diff,
            "selected_sheet": prep.sheet_name or sheet_name or "",
        })
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"处理失败: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass


@app.post("/api/import/confirm")
async def api_import_confirm(file: UploadFile = File(...),
                             module_id: str = Form(...),
                             granularity: str = Form(...),
                             sheet_name: str = Form(default=""),
                             included_fields: str = Form(default="[]"),
                             value_cols: str = Form(default="[]"),
                             time_col: str = Form(default=""),
                             field_types: str = Form(default="{}"),
                             dtypes: str = Form(default="{}"),
                             field_renames: str = Form(default="{}"),
                             included_orig: str = Form(default="[]")):
    """导入 Excel（第二步）：用户确认字段结构后，保存模板并导入数据。

    included_fields / value_cols 为 JSON 数组字符串。
    time_col 为用户指定的时间列名。
    field_types / dtypes 为 JSON 对象字符串（字段名→类型）。
    首次导入建立模板；后续导入结构变化经用户确认后更新模板。
    """
    if module_id not in data_store.module_ids():
        raise HTTPException(status_code=400, detail="未知装置模块")
    if granularity not in ("daily", "hourly"):
        raise HTTPException(status_code=400, detail="粒度须为 daily 或 hourly")
    if not file.filename:
        raise HTTPException(status_code=400, detail="未提供文件")
    suffix = Path(file.filename).suffix.lower()
    if suffix not in (".xlsx", ".xls"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 格式")

    tmp_path = None
    try:
        content = await file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        prep = _prepare_long_df_full(tmp_path, sheet_name=sheet_name if sheet_name else None)
        if prep.errors:
            raise HTTPException(status_code=400, detail=";".join(prep.errors))

        # Use user-specified time_col if auto-detection failed
        user_specified_time = time_col.strip() if time_col else ""
        if prep.time_col is None and user_specified_time:
            # User manually set the time column - find it in the data
            if user_specified_time in prep.df.columns:
                prep.time_col = user_specified_time
                prep.mapping[user_specified_time] = "time"
            else:
                # Try case-insensitive match
                for col in prep.df.columns:
                    if str(col).strip() == user_specified_time:
                        prep.time_col = col
                        prep.mapping[col] = "time"
                        break
        # If still no time_col but we have _extracted_date from header, use it
        if prep.time_col is None and "_extracted_date" in prep.df.columns:
            prep.time_col = "_extracted_date"
            prep.mapping["_extracted_date"] = "time"

        if prep.time_col is None:
            raise HTTPException(status_code=400, detail="无法识别时间列，请在字段设置中手动指定时间列")

        try:
            included = json.loads(included_fields) if included_fields else []
            vcols = json.loads(value_cols) if value_cols else []
            ftypes = json.loads(field_types) if field_types else {}
            dtypes_map = json.loads(dtypes) if dtypes else {}
            renames = json.loads(field_renames) if field_renames else {}
            included_orig_list = json.loads(included_orig) if included_orig else []
        except Exception:
            raise HTTPException(status_code=400, detail="字段参数格式错误")
        if not isinstance(included, list) or not isinstance(vcols, list):
            raise HTTPException(status_code=400, detail="字段参数格式错误")

        # 仅在勾选（包含）的字段中检查重名：两个勾选了相同字段名才报错
        dup_in_included = [n for n in set(included) if included.count(n) > 1]
        if dup_in_included:
            raise HTTPException(status_code=400,
                detail="勾选的字段名重复，请确保每个勾选字段名唯一：" + "、".join(dup_in_included))

        # 构建完整重命名映射：显式重命名 + 隐式（空表头字段的显示名）
        rename_map = {}
        if renames and isinstance(renames, dict):
            rename_map = {k: v for k, v in renames.items() if k in prep.df.columns and v}
        full_rename = dict(rename_map)
        if included_orig_list and len(included_orig_list) == len(included):
            for orig, final in zip(included_orig_list, included):
                if orig != final and orig not in full_rename:
                    full_rename[orig] = final

        # 保存原始列名（重命名前），用于按原始列名精确选取列
        original_cols = list(prep.df.columns)
        if full_rename:
            valid = {k: v for k, v in full_rename.items() if k in prep.df.columns and v}
            if valid:
                prep.df = prep.df.rename(columns=valid)

        # 用户指定的时间列（优先于自动检测）
        user_time_col = time_col.strip() if time_col else prep.time_col

        # 按原始列名选取勾选的列（避免重名列导致选取多列）
        final_to_orig = {}
        for orig in original_cols:
            fn = full_rename.get(orig, orig)
            if fn not in final_to_orig:
                final_to_orig[fn] = orig
        if included_orig_list:
            present_orig = [c for c in included_orig_list if c in original_cols]
        else:
            present_orig = []
            for nm in included:
                orig = final_to_orig.get(nm, nm)
                if orig in original_cols and orig not in present_orig:
                    present_orig.append(orig)
        time_orig = final_to_orig.get(user_time_col, user_time_col) if user_time_col else None
        if time_orig and time_orig in original_cols and time_orig not in present_orig:
            present_orig = [time_orig] + present_orig
            if user_time_col and user_time_col not in included:
                included = [user_time_col] + included
        if not present_orig:
            raise HTTPException(status_code=400, detail="未选择任何有效字段")
        # 按位置选取列，列名为重命名后的名字
        col_positions = [original_cols.index(c) for c in present_orig]
        filtered = prep.df.iloc[:, col_positions].copy()
        present = list(filtered.columns)
        vcols = [c for c in vcols if c in filtered.columns]

        # 按用户指定的数据类型转换列
        dtype_map = {"文本": "object", "整数": "Int64", "小数": "float64",
                     "日期时间": "datetime64[ns]", "布尔": "bool"}
        for col_name, dlabel in dtypes_map.items():
            if col_name not in filtered.columns:
                continue
            target = dtype_map.get(dlabel)
            if not target:
                continue
            try:
                if target == "datetime64[ns]":
                    filtered[col_name] = pd.to_datetime(filtered[col_name], errors="coerce")
                elif target == "Int64":
                    filtered[col_name] = pd.to_numeric(filtered[col_name], errors="coerce").astype("Int64")
                elif target == "float64":
                    filtered[col_name] = pd.to_numeric(filtered[col_name], errors="coerce")
                elif target == "bool":
                    filtered[col_name] = filtered[col_name].astype(bool)
                else:
                    filtered[col_name] = filtered[col_name].astype(target)
            except Exception:
                pass  # 转换失败时保留原始类型

        actual_gran = data_store.detect_granularity(filtered, user_time_col)
        actual_sid = data_store.granularity_source_id(module_id, actual_gran)
        mismatch = (granularity != actual_gran)
        gran_warn = _granularity_warning(granularity, actual_gran)

        # 建立/更新字段结构模板
        fs = field_structure_from_prep(prep)
        template = {
            "all_fields": fs["field_order"],
            "included_fields": present,
            "value_cols": vcols,
            "time_col": user_time_col,
            "flow_col": prep.flow_col,
            "field_types": ftypes,
            "dtypes": dtypes_map,
        }
        data_store.save_template(actual_sid, template)

        data_store.save_import(actual_sid, filtered, user_time_col, prep.flow_col,
                               prep.mapping, present, vcols)
        preview = data_store.build_wide_preview(actual_sid)
        preview["adjustments"] = prep.adjustments
        preview["action"] = "import"
        preview["granularity_mismatch"] = mismatch
        preview["actual_granularity"] = actual_gran
        preview["granularity_warning"] = gran_warn
        preview["needs_confirmation"] = False
        preview["template_matched"] = False
        preview["template_updated"] = True
        return JSONResponse(preview)
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"处理失败: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass


@app.post("/api/clear")
async def api_clear(module_id: str = Form(...),
                    granularity: str = Form(...),
                    start_date: str = Form(default=""),
                    end_date: str = Form(default="")):
    """清除指定模块、粒度、时间段的数据。granularity=all 同时清除天级和小时级。"""
    if module_id not in data_store.module_ids():
        raise HTTPException(status_code=400, detail="未知装置模块")
    if granularity not in ("daily", "hourly", "all"):
        raise HTTPException(status_code=400, detail="粒度须为 daily、hourly 或 all")

    gran_list = ["daily", "hourly"] if granularity == "all" else [granularity]
    results = {}
    total_cleared = 0
    for g in gran_list:
        sid = data_store.granularity_source_id(module_id, g)
        r = data_store.clear_data(sid, start_date, end_date)
        results[g] = r
        total_cleared += r["cleared"]

    return {
        "module_id": module_id,
        "module_name": data_store.module_name(module_id),
        "granularity": granularity,
        "results": results,
        "total_cleared": total_cleared,
    }


@app.get("/api/source/{source_id}/preview")
async def api_source_preview(source_id: str):
    """查看指定数据源已存储数据的宽表预览。"""
    if source_id not in data_store.source_ids():
        raise HTTPException(status_code=400, detail="未知数据源")
    preview = data_store.build_wide_preview(source_id)
    preview["action"] = "view"
    return JSONResponse(preview)


def _pivot_from_store(source_id: str):
    store = data_store.load_store(source_id)
    if not store or store.get("long_df") is None or store["long_df"].empty:
        return None, [], [], "", ["该数据源暂无数据，请先导入"]
    df = store["long_df"]
    time_col = store.get("time_col") or ""
    flow_col = store.get("flow_col")
    value_cols = store.get("value_cols")
    try:
        wide_df, flat_cols, header_rows = _pivot_to_wide(df, time_col, flow_col, value_cols)
        return wide_df, flat_cols, header_rows, time_col, []
    except Exception as e:
        return None, [], [], time_col, [f"宽表转换失败: {e}"]


def _field_structure_from_store(source_id: str) -> Dict[str, Any]:
    """从已导入存储的数据提取字段结构，供"调整字段结构"使用。

    复用 field_structure_from_prep 的输出格式，但从存储的长表 + 模板构建，
    无需重新解析 Excel。
    """
    store = data_store.load_store(source_id)
    if not store or store.get("long_df") is None or store["long_df"].empty:
        return {"fields": [], "field_order": [], "time_col": None, "flow_col": None,
                "value_cols": [], "errors": ["该数据源暂无数据"]}
    df = store["long_df"]
    time_col = store.get("time_col") or ""
    flow_col = store.get("flow_col")
    value_cols = store.get("value_cols") or []
    template = data_store.load_template(source_id)
    field_types = (template or {}).get("field_types") or {}
    dtypes_map = (template or {}).get("dtypes") or {}
    included_fields = (template or {}).get("included_fields") or []

    # 自动分类描述列/数值列（用于补全无模板时的字段类型）
    if time_col and time_col in df.columns:
        _desc, auto_values, _dropped, _dn = _classify_columns(df, time_col)
    else:
        auto_values = []

    fields: List[Dict[str, Any]] = []
    field_order: List[str] = []
    _empty_counter = 0
    _seen_cols = set()
    for c in df.columns:
        cs = str(c).strip()
        if cs in _seen_cols:
            continue
        _seen_cols.add(cs)
        is_empty_header = cs.startswith("Unnamed:") or cs == "" or cs == "nan"
        if is_empty_header:
            _empty_counter += 1
            display_name = f"(空字段名{_empty_counter})"
        else:
            display_name = str(c)
        if is_empty_header and df[c].notna().sum() == 0:
            _empty_counter -= 1
            continue
        # 字段类型：优先模板 → 自动识别
        if c in field_types:
            ftype = field_types[c]
        elif c == time_col:
            ftype = "时间列"
        elif c in value_cols:
            ftype = "数值列"
        else:
            ftype = "描述列"
        # 数据类型：优先模板 → 实际 dtype
        if c in dtypes_map:
            dtype_label = dtypes_map[c]
        else:
            dtype_label = _dtype_label_from_pandas(str(df[c].dtype))
        try:
            raw = df[c].dropna().unique()[:4]
            samples = [_val_to_jsonable(v) for v in raw]
        except Exception:
            samples = []
        is_included = (c in included_fields) if included_fields else (not is_empty_header)
        fields.append({
            "name": display_name, "orig_name": str(c),
            "field_type": ftype, "dtype_label": dtype_label,
            "samples": samples, "is_time": (c == time_col),
            "is_value": (c in value_cols), "included": is_included,
            "is_empty_header": is_empty_header,
        })
        field_order.append(str(c))
    return {
        "fields": fields, "field_order": field_order,
        "time_col": time_col, "flow_col": flow_col, "value_cols": value_cols,
        "errors": [], "source_id": source_id,
    }


def _dtype_label_from_pandas(dtype_str: str) -> str:
    """pandas dtype → 中文标签，与前端 _dtypeToLabel 一致。"""
    if dtype_str.startswith("int"):
        return "整数"
    if dtype_str.startswith("float"):
        return "小数"
    if dtype_str.startswith("datetime"):
        return "日期时间"
    if dtype_str.startswith("bool"):
        return "布尔"
    return "文本"


@app.get("/api/source/{source_id}/fields")
async def api_source_fields(source_id: str):
    """获取已导入数据源的字段结构，用于手动调整。"""
    if source_id not in data_store.source_ids():
        raise HTTPException(status_code=400, detail="未知数据源")
    return JSONResponse(_field_structure_from_store(source_id))


@app.post("/api/source/{source_id}/adjust")
async def api_source_adjust(source_id: str,
                            included_fields: str = Form(default="[]"),
                            value_cols: str = Form(default="[]"),
                            time_col: str = Form(default=""),
                            field_types: str = Form(default="{}"),
                            dtypes: str = Form(default="{}"),
                            field_renames: str = Form(default="{}"),
                            included_orig: str = Form(default="[]")):
    """手动调整已导入数据的字段结构，按新配置重新透视并更新存储与模板。

    参数与 /api/import/confirm 一致，但数据源来自已有存储而非重新解析文件。
    """
    if source_id not in data_store.source_ids():
        raise HTTPException(status_code=400, detail="未知数据源")
    store = data_store.load_store(source_id)
    if not store or store.get("long_df") is None or store["long_df"].empty:
        raise HTTPException(status_code=400, detail="该数据源暂无数据，无法调整")

    try:
        included = json.loads(included_fields) if included_fields else []
        vcols = json.loads(value_cols) if value_cols else []
        ftypes = json.loads(field_types) if field_types else {}
        dtypes_map = json.loads(dtypes) if dtypes else {}
        renames = json.loads(field_renames) if field_renames else {}
        included_orig_list = json.loads(included_orig) if included_orig else []
    except Exception:
        raise HTTPException(status_code=400, detail="字段参数格式错误")
    if not isinstance(included, list) or not isinstance(vcols, list):
        raise HTTPException(status_code=400, detail="字段参数格式错误")

    # 仅在勾选（包含）的字段中检查重名
    dup_in_included = [n for n in set(included) if included.count(n) > 1]
    if dup_in_included:
        raise HTTPException(status_code=400,
            detail="勾选的字段名重复，请确保每个勾选字段名唯一：" + "、".join(dup_in_included))

    df = store["long_df"]
    # 构建完整重命名映射
    rename_map = {}
    if renames and isinstance(renames, dict):
        rename_map = {k: v for k, v in renames.items() if k in df.columns and v}
    full_rename = dict(rename_map)
    if included_orig_list and len(included_orig_list) == len(included):
        for orig, final in zip(included_orig_list, included):
            if orig != final and orig not in full_rename:
                full_rename[orig] = final

    # 保存原始列名，按位置选取列
    original_cols = list(df.columns)
    if full_rename:
        valid = {k: v for k, v in full_rename.items() if k in df.columns and v}
        if valid:
            df = df.rename(columns=valid)
    old_time_col = store.get("time_col") or ""
    user_time_col = time_col.strip() if time_col else old_time_col

    # 按原始列名选取勾选的列
    final_to_orig = {}
    for orig in original_cols:
        fn = full_rename.get(orig, orig)
        if fn not in final_to_orig:
            final_to_orig[fn] = orig
    if included_orig_list:
        present_orig = [c for c in included_orig_list if c in original_cols]
    else:
        present_orig = []
        for nm in included:
            orig = final_to_orig.get(nm, nm)
            if orig in original_cols and orig not in present_orig:
                present_orig.append(orig)
    time_orig = final_to_orig.get(user_time_col, user_time_col) if user_time_col else None
    if time_orig and time_orig in original_cols and time_orig not in present_orig:
        present_orig = [time_orig] + [c for c in present_orig if c != time_orig]
    if not present_orig:
        raise HTTPException(status_code=400, detail="未选择任何有效字段")
    col_positions = [original_cols.index(c) for c in present_orig]
    filtered = df.iloc[:, col_positions].copy()
    present = list(filtered.columns)
    vcols = [c for c in vcols if c in filtered.columns]

    # 按用户指定的数据类型转换列
    dtype_map = {"文本": "object", "整数": "Int64", "小数": "float64",
                 "日期时间": "datetime64[ns]", "布尔": "bool"}
    for col_name, dlabel in dtypes_map.items():
        if col_name not in filtered.columns:
            continue
        target = dtype_map.get(dlabel)
        if not target:
            continue
        try:
            if target == "datetime64[ns]":
                filtered[col_name] = pd.to_datetime(filtered[col_name], errors="coerce")
            elif target == "Int64":
                filtered[col_name] = pd.to_numeric(filtered[col_name], errors="coerce").astype("Int64")
            elif target == "float64":
                filtered[col_name] = pd.to_numeric(filtered[col_name], errors="coerce")
            elif target == "bool":
                filtered[col_name] = filtered[col_name].astype(bool)
            else:
                filtered[col_name] = filtered[col_name].astype(target)
        except Exception:
            pass

    flow_col = vcols[0] if vcols else (store.get("flow_col") or "")

    # 更新存储（不累积，直接替换为调整后的长表）
    store["long_df"] = filtered
    store["time_col"] = user_time_col
    store["flow_col"] = flow_col
    store["value_cols"] = vcols
    store["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    data_store._save_store(source_id, store)

    # 更新模板
    template = {
        "all_fields": [str(c) for c in df.columns],
        "included_fields": present,
        "value_cols": vcols,
        "time_col": user_time_col,
        "flow_col": flow_col,
        "field_types": ftypes,
        "dtypes": dtypes_map,
    }
    data_store.save_template(source_id, template)

    preview = data_store.build_wide_preview(source_id)
    preview["action"] = "adjust"
    preview["needs_confirmation"] = False
    return JSONResponse(preview)


def _write_wide_excel(wide_df, flat_cols, header_rows, time_col: str) -> bytes:
    """把宽表 DataFrame + 多级表头写成 Excel 字节流。"""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "物料平衡"

    for ri, hrow in enumerate(header_rows, start=1):
        ci = 1
        for cell in hrow:
            cs = int(cell.get("colspan", 1) or 1)
            rs = int(cell.get("rowspan", 1) or 1)
            ws.cell(row=ri, column=ci, value=cell.get("text", ""))
            if cs > 1 or rs > 1:
                ws.merge_cells(start_row=ri, start_column=ci,
                               end_row=ri + rs - 1, end_column=ci + cs - 1)
            ci += cs

    data_start_row = len(header_rows) + 1
    for ridx, (_, row) in enumerate(wide_df.iterrows()):
        for cidx, col in enumerate(flat_cols, start=1):
            v = row.get(col)
            if pd.isna(v):
                v = None
            elif hasattr(v, "strftime"):
                v = v.strftime("%Y-%m-%d %H:%M:%S") if hasattr(v, "hour") else v.strftime("%Y-%m-%d")
            ws.cell(row=data_start_row + ridx, column=cidx, value=v)

    for ci in range(1, len(flat_cols) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.post("/api/export")
async def api_export(source_id: str = Form(...),
                     start_date: str = Form(default=""),
                     end_date: str = Form(default=""),
                     filename: str = Form(default="")):
    """按数据源 + 自定义时间段导出累积宽表为 Excel。"""
    if source_id not in data_store.source_ids():
        raise HTTPException(status_code=400, detail="未知数据源")
    wide_df, flat_cols, header_rows, time_col, errors = _pivot_from_store(source_id)
    if errors:
        raise HTTPException(status_code=400, detail="；".join(errors))
    if wide_df is None or wide_df.empty:
        raise HTTPException(status_code=400, detail="该数据源无数据，无法导出")

    if time_col and time_col in wide_df.columns:
        tser = pd.to_datetime(wide_df[time_col], errors="coerce")
        mask = pd.Series(True, index=wide_df.index)
        if start_date:
            t0 = pd.to_datetime(start_date, errors="coerce")
            if pd.notna(t0):
                mask &= (tser >= t0)
        if end_date:
            t1 = pd.to_datetime(end_date, errors="coerce")
            if pd.notna(t1):
                t1_end = t1 + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
                mask &= (tser <= t1_end)
        wide_df = wide_df[mask]
        if wide_df.empty:
            raise HTTPException(status_code=400, detail="所选时间段内无数据，请调整时间范围")

    xlsx_bytes = _write_wide_excel(wide_df, flat_cols, header_rows, time_col)

    name = data_store.source_name(source_id)
    stem = filename.strip() or (name + "_导出")
    for ch in '\\/:*?"<>|':
        stem = stem.replace(ch, "_")
    if not stem.lower().endswith(".xlsx"):
        stem = stem + ".xlsx"
    disp = "attachment; filename*=UTF-8''" + quote(stem)

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": disp},
    )


# ==================== 总收率分析 ====================

@app.get("/analysis")
async def analysis_page(request: Request):
    """总收率趋势分析页面。"""
    return templates.TemplateResponse("analysis.html", {"request": request})


@app.get("/api/analysis/sources")
async def api_analysis_sources():
    """列出分析数据源。分析直接基于导入数据计算，有导入数据即可分析。"""
    result = []
    for s in analyzer.ANALYSIS_SOURCES:
        sid = s["id"]
        store = data_store.load_store(sid)
        has_import = bool(store and store.get("long_df") is not None and not store["long_df"].empty)
        result.append({
            "source_id": sid, "name": s["name"], "source_name": s["name"],
            "has_data": has_import,
            "has_analysis": has_import, "has_import": has_import,
        })
    return {"sources": result}


@app.post("/api/analysis/import")
async def api_analysis_import(file: UploadFile = File(...), source_id: str = Form(...)):
    """导入宽表 Excel，计算总收率并存储，返回趋势数据。"""
    if source_id not in analyzer.analysis_source_ids():
        raise HTTPException(status_code=400, detail="未知数据源")
    if not file.filename:
        raise HTTPException(status_code=400, detail="未提供文件")
    suffix = Path(file.filename).suffix.lower()
    if suffix not in (".xlsx", ".xls"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 格式")

    tmp_path = None
    try:
        content = await file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        parsed = analyzer.parse_wide_balance_excel(tmp_path)
        if parsed["errors"]:
            raise HTTPException(status_code=400, detail="；".join(parsed["errors"]))
        data = analyzer.save_analysis(source_id, parsed, file.filename)
        header = {k: v for k, v in data["header_info"].items() if k != "colnames"}
        return JSONResponse({
            "has_data": True, "source_id": source_id,
            "source_name": data["source_name"],
            "points": data["points"], "stats": data["stats"],
            "adjustments": data["adjustments"], "header_info": header,
            "filename": data["filename"], "imported_at": data["imported_at"],
            "yield_min": data["yield_min"], "yield_max": data["yield_max"],
            "errors": [],
        })
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"处理失败: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass


@app.get("/api/analysis/yield")
async def api_analysis_yield(source_id: str, neg_filter: str = "filter", unit: str = ""):
    """获取指定数据源的总收率趋势，直接基于导入数据计算，保证与导入数据一致。"""
    if source_id not in analyzer.analysis_source_ids():
        raise HTTPException(status_code=400, detail="未知数据源")
    store = data_store.load_store(source_id)
    has_import = bool(store and store.get("long_df") is not None and not store["long_df"].empty)
    if not has_import:
        return JSONResponse({
            "source_id": source_id,
            "source_name": analyzer.analysis_source_name(source_id),
            "has_data": False, "points": [], "stats": {},
            "errors": ["该数据源暂无导入数据，请先在导入页面导入数据"],
            "yield_min": analyzer.YIELD_MIN, "yield_max": analyzer.YIELD_MAX,
        })
    result = analyzer.compute_yield_from_store(store, neg_filter=neg_filter, unit=unit)
    points = result.get("points", [])
    return JSONResponse({
        "source_id": source_id,
        "source_name": analyzer.analysis_source_name(source_id),
        "has_data": len(points) > 0,
        "points": points,
        "stats": result.get("stats", {}),
        "adjustments": result.get("adjustments", []),
        "header_info": result.get("header_info", {}),
        "errors": result.get("errors", []),
        "yield_min": analyzer.YIELD_MIN, "yield_max": analyzer.YIELD_MAX,
        "filename": store.get("name", ""),
        "imported_at": store.get("last_updated", ""),
    })

@app.get("/api/analysis/distribution")
async def api_analysis_distribution(source_id: str, neg_filter: str = "filter", start_date: str = "", end_date: str = "", exclude_dates: str = "", unit: str = ""):
    """获取进/出方向各物料占比（饼图）。"""
    if source_id not in analyzer.analysis_source_ids():
        raise HTTPException(status_code=400, detail="未知数据源")
    store = data_store.load_store(source_id)
    has_import = bool(store and store.get("long_df") is not None and not store["long_df"].empty)
    if not has_import:
        return JSONResponse({
            "source_id": source_id,
            "has_data": False,
            "in_materials": [], "out_materials": [],
            "errors": ["该数据源暂无导入数据"],
        })
    result = analyzer.compute_material_distribution_from_store(store, neg_filter=neg_filter, start_date=start_date, end_date=end_date, exclude_dates=exclude_dates, unit=unit)
    return JSONResponse({
        "source_id": source_id,
        "has_data": result.get("has_data", False),
        "in_materials": result.get("in_materials", []),
        "out_materials": result.get("out_materials", []),
        "errors": result.get("errors", []),
    })


@app.get("/api/analysis/material-trend")
async def api_analysis_material_trend(source_id: str, neg_filter: str = "filter"):
    """获取指定数据源的各物料流量时间序列趋势，直接基于导入数据计算。"""
    if source_id not in analyzer.analysis_source_ids():
        raise HTTPException(status_code=400, detail="未知数据源")
    store = data_store.load_store(source_id)
    has_import = bool(store and store.get("long_df") is not None and not store["long_df"].empty)
    if not has_import:
        return JSONResponse({
            "source_id": source_id,
            "source_name": analyzer.analysis_source_name(source_id),
            "has_data": False, "materials": [], "dates": [], "series": {},
            "errors": ["该数据源暂无导入数据，请先在导入页面导入数据"],
        })
    result = analyzer.compute_material_trend_from_store(store, neg_filter=neg_filter)
    return JSONResponse({
        "source_id": source_id,
        "source_name": analyzer.analysis_source_name(source_id),
        "has_data": result.get("has_data", False),
        "materials": result.get("materials", []),
        "dates": result.get("dates", []),
        "series": result.get("series", {}),
        "name_col": result.get("name_col", ""),
        "value_col": result.get("value_col", ""),
        "errors": result.get("errors", []),
    })


@app.get("/api/analysis/io-flow")
async def api_analysis_io_flow(source_id: str, neg_filter: str = "filter"):
    """获取指定数据源的进/出物料总量时间序列，支持按装置筛选。"""
    if source_id not in analyzer.analysis_source_ids():
        raise HTTPException(status_code=400, detail="未知数据源")
    store = data_store.load_store(source_id)
    has_import = bool(store and store.get("long_df") is not None and not store["long_df"].empty)
    if not has_import:
        return JSONResponse({
            "source_id": source_id,
            "source_name": analyzer.analysis_source_name(source_id),
            "has_data": False, "dates": [], "series": {}, "units": [],
            "errors": ["该数据源暂无导入数据"],
        })
    result = analyzer.compute_io_flow_from_store(store, neg_filter=neg_filter)
    return JSONResponse({
        "source_id": source_id,
        "source_name": analyzer.analysis_source_name(source_id),
        "has_data": result.get("has_data", False),
        "dates": result.get("dates", []),
        "series": result.get("series", {}),
        "units": result.get("units", []),
        "errors": result.get("errors", []),
    })





# ---------------------------------------------------------------------------
# 罐表分析页面
# ---------------------------------------------------------------------------

@app.get("/tank-analysis")
async def tank_analysis_page(request: Request):
    """罐表分析页面。"""
    return templates.TemplateResponse("tank_analysis.html", {"request": request})


@app.get("/api/tank-analysis/sources")
async def api_tank_sources():
    """列出所有含罐表数据的数据源。"""
    return {"sources": analyzer.tank_data_sources()}


@app.get("/api/tank-analysis/overview")
async def api_tank_overview(source_id: str):
    """获取罐表数据概览：罐号列表、油品列表、数值列、日期范围。"""
    result = analyzer.compute_tank_overview(source_id)
    return JSONResponse(result)


@app.get("/api/tank-analysis/trend")
async def api_tank_trend(source_id: str,
                         tanks: str = "", materials: str = "",
                         time_point: str = "前尺"):
    """计算罐/原料库存变化趋势。

    tanks: 逗号分隔的罐号列表
    materials: 逗号分隔的油品名称列表
    time_point: 时间点 前尺 / 12点 / 0点
    """
    tank_list = [s for s in tanks.split(",") if s.strip()] if tanks else []
    mat_list = [s for s in materials.split(",") if s.strip()] if materials else []
    result = analyzer.compute_tank_trend(source_id, tanks=tank_list,
                                         materials=mat_list, time_point=time_point)
    return JSONResponse(result)
