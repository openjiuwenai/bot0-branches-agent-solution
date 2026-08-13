# -*- coding: utf-8 -*-
"""
原油加工预排产
读「参数_原油/参数_罐/参数_掺炼/参数_全局 + 原油到港计划/原油加工计划/月初状态」，
排 CDU 连续加工序列 + 罐物流仿真(到港入G罐->G罐进料，或到港入T罐→T→G整批倒罐→G罐进料)，回填日网格。

v3 变更：
  - 参数_掺炼 sheet：改用绝对负荷(吨/h)表示掺炼配方，不再用比例
    格式：配方编号 | 主力油种 | 主力油负荷限额(吨/h) | 掺炼油种 | 掺炼油负荷 | 下限 | 上限
  - T罐→G罐传输：一次性整批转移（下一批次所需量），不再每天分批传输
  - 月加工计划预排产sheet：按到港计划靠泊日期/数量填写到库量行（T罐对应行）
"""
import calendar, math, time
from collections import defaultdict
from dataclasses import dataclass, field
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from ortools.sat.python import cp_model

# 油轮卸油到库量单元格高亮样式：红底 + 白字粗体
UNLOAD_FILL = PatternFill(fill_type="solid", fgColor="FF0000")
UNLOAD_FONT = Font(color="FFFFFF", bold=True)
NO_FEED_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")  # 不供料灰块

SRC = "原油加工月计划模板.xlsx"
WT2T = 10000.0  # 万吨→吨

# 回退默认
DEF_RATE_HR = 680   # 吨/小时
DEF_GCAP = 85000
DEF_TCAP = 85000
DEF_UNLOAD_TPH = 1800   # 卸油速度（吨/小时）
MIN_BATCH_H = 48        # 单批次最短加工时长（小时）
MAX_BATCH_H = 156       # 单批次最长加工时长（小时）
EDGE_MIN_BATCH_H = 24   # 月初月末批次最短加工时长（小时）：首批(st==0)/末批(en==H)可短到此值
SWITCH_RESID_TON = 1400 # G罐换油残余阈值（吨）：旧油可用残余≤此值即视为可换油（接收新油）
CLEAR_MARGIN_TON = 1500 # 腾容安全余量（吨）：腾容约束在最小需求上多清此量，避免零余量下靠泊船差几百吨装不下
EXTRA_MIN_TON = 0.0     # 非计划库存油纳入炼制的可用库存下限（吨）；> 此值且 >1e-6 才纳入
UNLOAD_PREP_H_DEF = 6    # 卸油准备时间默认（小时）：G罐停止供料后X小时开始卸油
FEED_PREP_H_DEF = 24     # 原油供料准备时间默认（小时）：卸油入罐后静置/化验等，之后才可供料
STOP_FEED_HOD_DEF = 24   # 卸油当日停止供料时间默认（24小时制）：24=用满靠泊当日
GG_TPH_DEF = 1000        # G→G输油速度默认（吨/小时）：本次占位不使用

def sched_sheet(wb):
    for n in wb.sheetnames:
        if "预排产" in n: return n
    return wb.sheetnames[0]

def norm(s): return str(s).strip().upper().replace("\n", "")
def tank_id(raw): return str(raw).replace("\n", " ").split()[0].strip().upper()
def yes(v): return str(v).strip() in ("是", "Y", "y", "TRUE", "True", "1")

def detect_extra_domestic(P):
    """识别「月初在G罐有可用库存、不在加工计划、且参数_原油已登记为国内」的原油，
    按CDU剩余产能部分纳入炼制：
      - 仅统计G罐中的非计划库存（T罐不腾容，不纳入）
      - 按CDU剩余产能（H×max_rate − ΣPROC）限额加入，超限不加入
      - 写入 P["PROC"][c]=可用库存（软目标），并记入 P["EXTRA_DOMESTIC"]。
    未登记 → 告警跳过（保持死库存）；进口 → 跳过。原地修改 P。"""
    from collections import defaultdict
    G = P.get("GTANKS", [])
    HEEL = P.get("HEEL", {})
    RATE_HR = P.get("RATE_HR", {})
    H = P["DAYS"] * 24

    # 只统计 G 罐中的非计划库存油
    extra_target = defaultdict(float)
    for t in G:
        d = P.get("INIT", {}).get(t, {})
        c = d.get("crude")
        if not c or c in P["PROC"]:
            continue
        avail_net = d.get("ton", 0.0) - HEEL.get(t, 0.0)
        if avail_net <= max(EXTRA_MIN_TON, 1e-6):
            continue
        extra_target[c] += avail_net

    # 计算 CDU 剩余产能
    plan_total = sum(P["PROC"].values())
    max_rate = max(RATE_HR.values()) if RATE_HR else 715
    cdu_capacity = H * max_rate
    spare_capacity = max(0, cdu_capacity - plan_total)
    print(f"  CDU产能={cdu_capacity:.0f}t  原计划={plan_total:.0f}t  剩余额度={spare_capacity:.0f}t")

    # 按剩余产能限额加入（超限的不加入）
    extra_domestic = set()
    used = 0.0
    for c, tgt in sorted(extra_target.items(), key=lambda x: x[1]):  # 小量优先
        if c not in RATE_HR:
            print(f"  月初G罐库存油 {c}({tgt:.0f}吨)未在『参数_原油』登记，暂不纳入炼制")
            continue
        if P.get("ORIGIN", {}).get(c) != "dom":
            print(f"  非计划库存油 {c}（产地{P.get('ORIGIN',{}).get(c)}）非国内，跳过纳入")
            continue
        if used + tgt > spare_capacity:
            print(f"  非计划库存油 {c}({tgt:.0f}吨)超出剩余产能额度({spare_capacity-used:.0f}t)，不纳入")
            continue
        P["PROC"][c] = tgt
        extra_domestic.add(c)
        used += tgt
        print(f"  纳入非计划国内库存油 {c}: 软目标={tgt:.0f}吨（G罐腾罐+补总量）")
    P["EXTRA_DOMESTIC"] = extra_domestic


# ============================================================
# 1. 读参数 + 业务输入
# ============================================================
def read_all(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    P = {}

    # 全局参数
    g = {}
    if "参数_全局" in wb.sheetnames:
        ws = wb["参数_全局"]
        for r in range(2, ws.max_row + 1):
            k = ws.cell(r, 1).value
            if k: g[str(k).strip()] = ws.cell(r, 2).value
    def gnum(key, default):
        """读全局参数为数值，缺失/非法时回退默认（默认值的类型决定 int/float）。"""
        v = g.get(key)
        if v is None: return default
        try: return type(default)(v)
        except (ValueError, TypeError): return default

    ym = str(g.get("排产年月", "2026-02")).split("-")
    P["YEAR"], P["MONTH"] = int(ym[0]), int(ym[1])
    P["DAYS"] = calendar.monthrange(P["YEAR"], P["MONTH"])[1]
    P["UNLOAD_TPH"] = float(g.get("卸油速度(吨/小时)", DEF_UNLOAD_TPH))
    P["TRANSFER_TPH"] = float(g.get("T→G输油速度(吨/小时)", 700))
    # 管道时序/能力 CP-SAT 建模：选配，默认关。
    # 开启时模型规模 ~14× 膨胀(6712变量)，当前 overlap 线性化使 CP-SAT 在时限内
    # 难以求解(UNKNOWN 超时)，故默认关；精确建模待后续优化求解时间，详见
    # docs/superpowers/specs/2026-06-20-T2G管道CP-SAT建模-design.md
    P["PIPELINE_MODEL"] = yes(g.get("管道建模(是/否)", "否"))
    # 批次时长 / 求解策略：从 参数_全局 读，缺失回退模块常量
    P["MIN_BATCH_H"] = gnum("单批次最短小时数", MIN_BATCH_H)
    P["MAX_BATCH_H"] = gnum("单批次最长小时数", MAX_BATCH_H)
    # 月初月末批次最短时长：首批(st==0)/末批(en==H)可短到此值（其余批次仍守 MIN_BATCH_H）
    P["EDGE_MIN_BATCH_H"] = gnum("月初月末批次最短小时数", EDGE_MIN_BATCH_H)
    # 注：单批次优选最短/最长加工小时数（PREF）已废弃、不再读取（2026-07-05 去优选时长评分）
    # G罐换油残余阈值（吨）：旧油可用残余≤此值即视为可换油，缺失回退模块常量
    P["SWITCH_RESID_TON"] = gnum("换油残余阈值(吨)", SWITCH_RESID_TON)
    # 腾容安全余量（吨）：腾容约束在最小需求上多清此量，避免零余量下靠泊船差几百吨装不下
    P["CLEAR_MARGIN"] = gnum("腾容安全余量(吨)", CLEAR_MARGIN_TON)
    P["TIME_LIMIT"] = gnum("单轮求解时限(秒)", 240)
    P["N_ROUNDS"] = gnum("多轮调度轮数", 8)
    P["N_PREFERRED"] = max(1, gnum("优选解数量", 3))   # 优选解输出数量；<1 钳为 1
    P["NUM_WORKERS"] = gnum("并行搜索线程数", 8)
    # 卸油/供料时序参数：卸油准备时间、原油供料准备时间、卸油当日停止供料时间、G→G输油速度
    P["UNLOAD_PREP_H"] = gnum("卸油准备时间(小时)", UNLOAD_PREP_H_DEF)
    P["FEED_PREP_H"]   = gnum("原油供料准备时间(小时)", FEED_PREP_H_DEF)
    P["STOP_FEED_HOD"] = gnum("卸油当日停止供料时间", STOP_FEED_HOD_DEF)
    P["GG_TPH"]        = float(g.get("G→G输油速度(吨/小时)", GG_TPH_DEF))
    P["MAX_UNLOAD_TANKS"] = max(1, gnum("最大一次卸油罐数", 2))   # 一船卸油最多占用几个罐(G+T合计)
    # 月度滞期天数上限：技术经济指标 sheet「月度最大滞期天数」行（值在 B 列，按行标签匹配读，
    # 不依赖固定单元格）。注意这是「限制」；靠泊计划已含的滞期天数从 原油到港计划 小计读取
    # (DEMURRAGE_INCLUDED)，tank_fill 可追加的滞期 = 限制 − 已含。
    P["DEMURRAGE_BUDGET"] = 0.0
    for n in wb.sheetnames:
        if "技术经济指标" in n:
            ws2 = wb[n]
            for r in range(1, ws2.max_row + 1):
                k = ws2.cell(r, 1).value
                if k and "月度最大滞期天数" in str(k):
                    try: P["DEMURRAGE_BUDGET"] = float(ws2.cell(r, 2).value)
                    except (ValueError, TypeError): pass
            break

    # 油种参数（以标准小时加工量为单位）
    rate_hr = {}; origin = {}; no_consec = set(); can_single = {}; allow_mix = {}
    if "参数_原油" in wb.sheetnames:
        ws = wb["参数_原油"]
        for r in range(2, ws.max_row + 1):
            nm = ws.cell(r, 1).value
            if not nm: continue
            c = norm(nm)
            hv = ws.cell(r, 2).value
            rate_hr[c] = int(round(float(hv))) if hv else DEF_RATE_HR  # 保留原值(不再整十)，取整保 int 域
            origin[c] = "imp" if "进口" in str(ws.cell(r, 3).value) else "dom"
            can_single[c] = yes(ws.cell(r, 4).value)
            if yes(ws.cell(r, 5).value): no_consec.add(c)
            mixv = ws.cell(r, 6).value
            allow_mix[c] = ({norm(x) for x in str(mixv).split(",") if x.strip()}
                            if mixv else set())
    P["RATE_HR"] = rate_hr   # 吨/小时，模板原值
    P["RATE"] = {c: v * 24 for c, v in rate_hr.items()}  # 吨/天，内部用
    P["ORIGIN"] = origin
    P["NO_CONSEC"] = no_consec
    P["CAN_SINGLE"] = can_single
    P["ALLOW_MIX"] = allow_mix   # 卸油允许混合油种：油种→白名单集合（空集=不允许与任何油种混）

    # 罐参数
    gtanks = []; ttanks = []; cap = {}; allow = {}; farness = {}; enabled = {}
    avail_cap = {}; heel = {}
    if "参数_罐" in wb.sheetnames:
        ws = wb["参数_罐"]
        for r in range(2, ws.max_row + 1):
            t = ws.cell(r, 1).value
            if not t: continue
            t = tank_id(t)
            role = str(ws.cell(r, 2).value or "")
            en = yes(ws.cell(r, 6).value); enabled[t] = en          # 启用 col6
            capv = ws.cell(r, 3).value                              # 罐容上限 col3
            heelv = ws.cell(r, 4).value                             # 底油 col4
            acapv = ws.cell(r, 5).value                             # 可用容量 col5
            cap[t] = float(capv) if capv else (DEF_GCAP if "进料" in role else DEF_TCAP)
            if acapv:
                avail_cap[t] = float(acapv)
            else:
                avail_cap[t] = cap[t] - (float(heelv) if heelv else 0.0)
            heel[t] = float(heelv) if heelv else 0.0
            allow[t] = str(ws.cell(r, 7).value or "任意").strip()   # 允许油种 col7
            farness[t] = ws.cell(r, 8).value or 1                   # 卸船远近 col8
            if not en: continue
            if "进料" in role or ("G" in t and "储" not in role):
                gtanks.append(t)
            else:
                ttanks.append(t)
    P["GTANKS"] = gtanks; P["TTANKS"] = ttanks
    P["CAP"] = cap; P["ALLOW"] = allow; P["FARNESS"] = farness
    P["AVAIL_CAP"] = avail_cap
    P["HEEL"] = heel

    # MAIN/BLEND 罐组自动判定：进料G罐可用容量 ≥ G1101可用容量 → MAIN，否则 BLEND
    big_thr = avail_cap.get("G1101", DEF_GCAP)
    P["MAIN_TANKS"]  = {t for t in gtanks if avail_cap.get(t, 0.0) >= big_thr}
    P["BLEND_TANKS"] = {t for t in gtanks if avail_cap.get(t, 0.0) <  big_thr}

    # 掺炼配方（v3：绝对负荷表示）
    # 格式：配方编号 | 主力油种 | 主力油负荷限额(吨/h) | 掺炼油种 | 掺炼油负荷 | 下限 | 上限
    # 5列格式: 配方编号|主力油种|主力油负荷限额|掺炼油种|掺炼负荷可取值(逗号枚举)
    # 同一配方编号下，主力油行1行 + 掺炼油行1~2行
    def parse_cands(s):
        out = []
        for tok in str(s).replace("，", ",").split(","):
            tok = tok.strip()
            if tok:
                try: v = round(float(tok) / 10) * 10  # 10步长校验
                except ValueError: continue
                if v > 0 and v not in out: out.append(v)
        return out
    recipes = {}
    if "参数_掺炼" in wb.sheetnames:
        ws = wb["参数_掺炼"]
        cur = None; cur_data = None; skip = False
        for r in range(2, ws.max_row + 1):
            rid = ws.cell(r, 1).value
            blend_crude = norm(ws.cell(r, 4).value or "")
            cands_cell = ws.cell(r, 5).value
            if rid:
                rid_str = str(rid).strip()
                if "示例" in rid_str or "说明" in rid_str:
                    skip = True; continue
                skip = False; cur = rid_str
                main_crude = norm(ws.cell(r, 2).value or "")
                main_cap_hr = ws.cell(r, 3).value
                if not main_cap_hr and main_crude in rate_hr:
                    main_cap_hr = rate_hr[main_crude]
                main_cap_hr = int(round(float(main_cap_hr))) if main_cap_hr else DEF_RATE_HR
                cur_data = {"main_crude": main_crude, "total_cap_hr": main_cap_hr, "blends": []}
                if blend_crude and cands_cell is not None:
                    cur_data["blends"].append({"crude": blend_crude, "cands": parse_cands(cands_cell)})
                recipes[cur] = cur_data
            elif not skip and cur_data is not None:
                if blend_crude and cands_cell is not None:   # 续行:第2个掺炼油
                    cur_data["blends"].append({"crude": blend_crude, "cands": parse_cands(cands_cell)})
    P["RECIPES_RAW"] = {k: v for k, v in recipes.items() if v["main_crude"]}

    # 加工计划: 油种→加工量(吨)
    ws = wb["原油加工计划"]; proc = {}; proc_total = None
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, 1).value
        if not nm: continue
        amt = ws.cell(r, 4).value
        if norm(nm) == "小计":   # 总加工量小计(D列)，达标判定的国内油总量目标
            if amt: proc_total = abs(float(amt)) * WT2T
            continue
        if amt: proc[norm(nm)] = abs(float(amt)) * WT2T
    P["PROC"] = {k: v for k, v in proc.items() if v > 0}
    P["PROC_TOTAL"] = proc_total if proc_total else sum(P["PROC"].values())

    # 月初罐状态（从预排产sheet读）——须在 detect_extra_domestic 之前，PROC 已就绪
    ws = wb[sched_sheet(wb)]; init = {}; r = 2
    while r <= ws.max_row:
        tag = ws.cell(r, 1).value
        if tag:
            t = tank_id(tag)
            inv = ws.cell(r, 3).value or 0
            cc = ws.cell(r + 1, 3).value
            crude = norm(cc) if cc else None
            if crude is None and inv:
                a = norm(P["ALLOW"].get(t, ""))
                if a in P["PROC"]: crude = a
            init[t] = {"crude": crude, "ton": float(inv)}
            r += 6 if t in gtanks or t.startswith("G") else 3
        else:
            r += 1
    P["INIT"] = init

    # 非计划国内库存油纳入炼制（须在 DENOM/配方筛选/build_units 之前，使其可见）
    detect_extra_domestic(P)

    # 批次预估分母：vol>small_ref 用大罐参考容量，否则小罐（detect 之后构建，自然包含 JZ）
    small_ref = P["AVAIL_CAP"].get("G151A", DEF_TCAP)
    big_ref   = P["AVAIL_CAP"].get("G1101", DEF_GCAP)
    P["BATCH_REF"] = (small_ref, big_ref)
    P["DENOM"] = {c: (big_ref if vol > small_ref else small_ref)
                  for c, vol in P["PROC"].items()}

    # 配方筛选：只保留 主力油+全部掺炼油 都在当月加工计划(PROC) 的配方
    kept = {}
    for rid, rc in P.get("RECIPES_RAW", {}).items():
        crudes = [rc["main_crude"]] + [b["crude"] for b in rc["blends"]]
        missing = [c for c in crudes if c not in P["PROC"]]
        if missing:
            print(f"  配方{rid} 含当月无关油种{missing}，已跳过")
        elif not rc["blends"] or any(not b["cands"] for b in rc["blends"]):
            print(f"  配方{rid} 掺炼负荷可取值缺失，已跳过")
        else:
            kept[rid] = rc
    P["RECIPES"] = kept

    # 到港计划（靠泊日在当月）
    ws = wb["原油到港计划"]; arr = []
    demurrage_included = 0.0   # 小计行 H列「预计滞期天数」：当前靠泊日期已包含的滞期天数
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, 2).value
        if not nm: continue
        if norm(nm) == "小计":
            hv = ws.cell(r, 8).value   # 预计滞期天数 小计
            if hv is not None:
                try: demurrage_included = float(hv)
                except (ValueError, TypeError): pass
            continue
        qty = ws.cell(r, 4).value; berth = ws.cell(r, 7).value
        if qty is None or berth is None: continue
        try:
            bday = berth.day if (berth.year == P["YEAR"] and berth.month == P["MONTH"]) else None
        except AttributeError:
            bday = None
        arr.append({"crude": norm(nm), "ton": float(qty) * WT2T, "berth_day": bday})
    P["ARRIVALS"] = arr
    # 已含滞期：靠泊日期已反映这些天数，故 tank_fill 的额外等待预算 = 限制 − 已含
    P["DEMURRAGE_INCLUDED"] = demurrage_included

    return P

# ============================================================
# 2. 构建加工单元（单炼 / 掺炼）
# ============================================================
def unload_days(ton, tph): return max(1, math.ceil(ton / (tph * 24)))

def unload_hours(ton, tph):
    """卸油小时数（向上取整到整点）：ceil(ton / tph)；tph<=0 返回 0。
    取整使卸油时长为整数小时→feed_ready/灰块落整点（避免 31200/1800=17.33 的非整数小时）。"""
    return math.ceil(ton / tph) if tph and tph > 0 else 0.0

def plan_unload_split(C, Q, tanks_state, pre_tank, allow_mix, max_tanks=2):
    """一船(油种C, Q吨)卸油：在六档优先级下找【≤max_tanks 个罐】的组合。

    六档(**T 储罐优先于 G 供料罐**，把 G 罐留出来供料)：
      ① 同种T ② 同种G ③ 空T ④ 空G ⑤ 混T ⑥ 混G；预分配罐(干净可容)为 Tier0。
      混油档必须 `c' ∈ allow_mix[C]`；异种非白名单罐不可用。
    选罐逻辑（组合搜索，非逐罐贪心）：
      - 优先返回**能整装整船**的最优 ≤K 罐组合（`remaining≈0`）；优选键=(罐数少, 组合最差档小
        →避免混油/落T靠后, 档位和, sum_far 近优先, −sum_space 大罐优先)。
      - 若**无 ≤K 组合能整装** → 返回**捕获空间最大**的 ≤K 罐组合（partial, `remaining>0`），
        供调用层滞期等待(延长靠泊)/超预算强卸时尽量少丢货。
    返回 (splits=[(tank,qty),...], remaining)。"""
    import itertools

    def space(t):
        s = tanks_state[t]
        # 收油 G 罐靠泊当日"先供后卸"腾出的空间(berth_feed)计入可容量：卸油发生在先供之后
        return s["cap"] - s["ton"] + s.get("berth_feed", 0.0)

    def is_empty(s):
        return (s["ton"] - s["heel"]) <= 1e-6

    def tier_of(t):
        s = tanks_state[t]
        if not s["allow_ok"] or space(t) <= 1e-6:
            return None
        is_g = s["is_g"]
        if t == pre_tank and (s["crude"] == C or is_empty(s)):
            return 0                                       # Tier0：预分配罐且干净可容
        if s["crude"] == C:         return 2 if is_g else 1   # 同种：T=1, G=2
        if is_empty(s):             return 4 if is_g else 3   # 空  ：T=3, G=4
        if s["crude"] in allow_mix: return 6 if is_g else 5   # 混  ：T=5, G=6（需白名单）
        return None                                        # 异种非白名单 → 不可用

    usable = [(t, tier_of(t), space(t)) for t in tanks_state]
    usable = [x for x in usable if x[1] is not None]
    if not usable:
        return [], float(Q)

    far = lambda t: tanks_state[t]["far"]

    def key_whole(combo):
        trs = [tr for _, tr, _ in combo]
        return (len(combo), max(trs), sum(trs),
                sum(far(t) for t, _, _ in combo),
                -sum(sp for _, _, sp in combo))

    # 找【能整装整船】的最少罐组合（罐数从少到多，首个有整装解的规模即最优规模）
    best = None
    for size in range(1, min(max_tanks, len(usable)) + 1):
        whole = [list(c) for c in itertools.combinations(usable, size)
                 if sum(sp for _, _, sp in c) >= Q - 1e-6]
        if whole:
            best = min(whole, key=key_whole)
            break
    if best is None:
        # 无 ≤K 组合能整装 → 取空间最大的 ≤K 罐（partial，尽量少丢货）
        best = sorted(usable, key=lambda x: -x[2])[:max_tanks]

    # 按档顺序灌装（高优先罐先灌满再下一个）；同档近罐/大罐先
    splits, remaining = [], float(Q)
    for t, tr, sp in sorted(best, key=lambda x: (x[1], far(x[0]), -x[2])):
        if remaining <= 1e-6:
            break
        q = min(remaining, sp)
        if q > 1e-6:
            splits.append((t, q)); remaining -= q
    return splits, remaining

def _overlap_by_day(lo, hi, valfn):
    """把小时区间 [lo, hi] 按每日 [(d-1)*24, d*24] 的重叠切分，返回 {day: valfn(重叠小时, day)}。"""
    out = {}
    d = int(lo // 24) + 1
    while (d - 1) * 24 < hi - 1e-9:
        ov = min(hi, d * 24) - max(lo, (d - 1) * 24)
        if ov > 1e-9:
            out[d] = valfn(ov, d)
        d += 1
    return out

def receiving_windows(splits_sf, unload_tph, unload_prep_h, feed_prep_h):
    """按串行卸油计算各收油罐的不供料时间窗口（小时级、逐日精确）。
    splits_sf: [(tank, qty, stop_feed_h), ...]，stop_feed_h=该罐【实际】最后供料时刻
      （数据驱动：当日无该罐供料→当日0点；供到9点→9点；封顶 STOP_FEED_HOD）。
    串行卸油：首罐停供后 +卸油准备 起卸，后续罐接续卸（单管道）。每罐：
      unload_start = 本罐起卸时刻；unload_end = +ceil(qty/tph)；feed_ready = +供料准备。
    - **不供料窗 = [unload_start − 卸油准备, feed_ready]**（本罐卸油准备+卸油+供料准备）；
      **不含串行等卸期**——第2+顺位罐在等前罐卸油期间可供料（其 stop_feed_h 之后到本罐起卸前的等待不标停料）。
    - **到库量按【卸油速度×当日卸油小时】分摊，最后一天=剩余未卸量**（`recv_by_day`），
      逐小时按 unload_tph 灌，末日补足剩余以保证 Σ=qty（总量守恒）。
    返回 per_tank[t] 含 no_feed_hours / recv_by_day；rel_h=末罐 feed_ready。"""
    per_tank = {}
    if not splits_sf:
        return {"per_tank": per_tank, "rel_h": 0.0}
    # 串行卸油以首罐（预分配/主罐）停供时刻为起点推进卸油管
    cursor = splits_sf[0][2] + unload_prep_h
    rel_h = splits_sf[0][2]
    for tank, qty, sf in splits_sf:
        us = cursor
        ue = us + unload_hours(qty, unload_tph)
        fr = ue + feed_prep_h
        cursor = ue                            # 下一罐接续卸
        rel_h = max(rel_h, fr)
        # 不供料窗下界 = 本罐起卸前的卸油准备开始（=us−准备时间），排除串行等卸期。
        # 首罐：us−准备=sf（含其卸油准备）；后续罐：远晚于 sf，等卸期归为可供料。
        lo = us - unload_prep_h
        no_feed_hours = _overlap_by_day(lo, fr, lambda ov, d: round(ov, 2))
        # 到库量按【卸油速度×当日卸油小时】分摊，最后一天=剩余未卸量（保证 Σ=qty）。
        # 卸油以恒速 unload_tph 进行：非末日到库=unload_tph×当日卸油小时；末日=之前几天未卸剩余。
        ov_by_day = _overlap_by_day(us, ue, lambda ov, d: ov)   # {day: 当日卸油小时}
        recv_by_day = {}
        assigned = 0.0
        days_sorted = sorted(ov_by_day)
        for i, d in enumerate(days_sorted):
            if i == len(days_sorted) - 1:
                recv_by_day[d] = qty - assigned                 # 末日 = 之前几天未卸剩余
            else:
                v = min(unload_tph * ov_by_day[d], qty - assigned)
                recv_by_day[d] = v
                assigned += v
        per_tank[tank] = {
            "stop_feed_h": sf, "unload_start_h": us,
            "unload_end_h": ue, "feed_ready_h": fr,
            "no_feed_hours": no_feed_hours, "recv_by_day": recv_by_day,
        }
    return {"per_tank": per_tank, "rel_h": rel_h}

def should_regrade(q, pre_ton, heel):
    """罐重命名(regrade)判据：卸量 q 大于卸前可用油(pre_ton-heel) 则罐存油种改为卸油油种。
    仅决定罐中存什么油种，与是否发生混油(commingling)无关——后者见 unload_commingled。"""
    return q > max(0.0, pre_ton - heel) + 1e-9

def arrival_rel_h(berth_day, ton, P):
    """到港油最早可加工时刻(整点小时) = feed_ready 向上取整：
    stop_feed_h + UNLOAD_PREP_H + 卸油小时 + FEED_PREP_H。"""
    stop_feed_h = (berth_day - 1) * 24 + P.get("STOP_FEED_HOD", STOP_FEED_HOD_DEF)
    fr = (stop_feed_h + P.get("UNLOAD_PREP_H", UNLOAD_PREP_H_DEF)
          + unload_hours(ton, P["UNLOAD_TPH"])
          + P.get("FEED_PREP_H", FEED_PREP_H_DEF))
    return int(math.ceil(fr))


def available_by_crude(P):
    """各油种本月可得总量 = 初始库存(扣底油) + 本月内到港量。
    **到港油最早可供料加工时刻(卸油准备+卸油+供料准备后)越月末(arrival_rel_h > H=次月1日0点)
    的船本月排不了，不计入**（与 avail_sources 剔除口径一致）。供 cp_schedule 加工量硬上限共用。

    【piece 生成的总量上限】返回 {油种: 可得吨数}，在 cp_schedule 中作为
    加工量硬约束 `pe ≤ avail` 的右端——piece 的 Σ this_h 折算吨数不能超过此值。
    与 avail_sources(按来源切分) 口径一致：月初库存 + 本月到港(越月末丢弃)。"""
    # 变量定义:
    #   proc    (dict): 加工计划 {油种: 目标吨数}
    #   H       (int):  月度总小时 = DAYS×24
    #   init_by (dict): 各油种月初可用库存(吨) = 库存 - 底油
    proc = P["PROC"]
    H = P["DAYS"] * 24
    init_by = defaultdict(float)
    for t, d in P["INIT"].items():
        if d["crude"]:
            init_by[d["crude"]] += d["ton"] - P["HEEL"].get(t, 0.0)
    # 变量定义: avail (dict): {油种: 可得总量}，初始为月初库存
    avail = {c: init_by.get(c, 0.0) for c in proc}
    for a in P["ARRIVALS"]:
        if not (a["berth_day"] and a["crude"] in avail):
            continue   # 作用: 无靠泊日或非计划油种，跳过
        # 变量定义: rel (float): 该船油最早可供料时刻
        rel = arrival_rel_h(a["berth_day"], a["ton"], P)
        if rel > H or rel + P.get("EDGE_MIN_BATCH_H", 24) > H:
            continue  # 到港太晚，剩余时间放不下最短批次（幽灵可用量）  # 作用: 越月末或到港后剩余<边批下限则丢弃
        avail[a["crude"]] += a["ton"]   # 作用: 本月可加工的到港量累加进可得总量
    return avail


def arrival_cum_segments(P):
    """各油种的到港可供料累计量分段：返回 {crude: [(可供料时刻 T_k, 累计到港量 cum), ...]}。
    供 cp_schedule 批次级可用量约束（候选1方式1）判定 st_p 落在哪个到港段。
    cum[k] = Σ(到港时刻 ≤ T_k 的到港量)，T_k 按时刻升序。
    与 available_by_crude(全月总量) 互补——后者 = 初始库存 + 最后一段 cum。"""
    proc = P["PROC"]; H = P["DAYS"] * 24
    events_by_crude = defaultdict(list)
    for a in P["ARRIVALS"]:
        if not (a.get("berth_day") and a["crude"] in proc):
            continue
        rel = arrival_rel_h(a["berth_day"], a["ton"], P)
        if rel > H:   # 越月末，本月排不了
            continue
        events_by_crude[a["crude"]].append((rel, a["ton"]))
    segments = {}
    for c in proc:
        evs = sorted(events_by_crude.get(c, []))
        cum = 0.0; segs = []
        for rel, ton in evs:
            cum += ton
            segs.append((rel, int(round(cum))))
        segments[c] = segs
    return segments


def g_init_availability(P):
    """各油种在G罐区(含T罐初始，T→G由tank_fill执行)的初始可用量(扣底油)。
    供 cp_schedule 批次级可用量约束（候选1方式1）的 g_init(c) 项。
    与 GPOOL_MODEL 的 g_init 口径一致（行 1050-1061）。"""
    HEEL = P.get("HEEL", {})
    g_init = defaultdict(float)
    # G 罐初始库存
    for t in P["GTANKS"]:
        d = P["INIT"].get(t, {})
        c = d.get("crude")
        if c:
            g_init[c] += d.get("ton", 0.0) - HEEL.get(t, 0.0)
    # T 罐初始库存也纳入（T→G 传输由 tank_fill 执行，CP-SAT 侧视为可用）
    for t in P["TTANKS"]:
        d = P["INIT"].get(t, {})
        c = d.get("crude")
        if c and c in P["PROC"]:
            g_init[c] += d.get("ton", 0.0) - HEEL.get(t, 0.0)
    return dict(g_init)


def greedy_batch_priority(P):
    """
    动态剩余批次预估贪心：每步在「剩余>0 且 当前可用≥最小批量」的油种里，
    挑 剩余批次预估(=剩余/分母) 最高者排下一批，扣减后重算，时间轴按到港事件推进。
    期初库存让油种从 hour 0 参与竞争。
    返回 (weights, order)：
      weights[crude] = [w1, w2, ...]  非递增，wk = 排第k批时的 剩余/分母
      order = [(crude, k, wk), ...]   全局排批顺序
    """
    proc = P["PROC"]; denom = P["DENOM"]; rate_hr = P["RATE_HR"]
    MINH = P.get("MIN_BATCH_H", MIN_BATCH_H)
    UTPH = P.get("UNLOAD_TPH", DEF_UNLOAD_TPH)

    init_by = defaultdict(float)
    for t, d in P["INIT"].items():
        if d["crude"]: init_by[d["crude"]] += d["ton"] - P["HEEL"].get(t, 0.0)
    avail = {c: init_by.get(c, 0.0) for c in proc}
    remaining = {c: proc[c] for c in proc}

    def min_batch_ton(c):
        return MINH * rate_hr.get(c, DEF_RATE_HR)

    # 到港事件：(可供油日, 油种, 吨)，按时间排序
    H = P["DAYS"] * 24
    events = []
    for a in P["ARRIVALS"]:
        if a["berth_day"] and a["crude"] in proc:
            # 最早可供料加工时刻越月末(>H=次月1日0点)的到港，本月排不了 → 不参与本月批次优先级/动态可用量
            if arrival_rel_h(a["berth_day"], a["ton"], P) > H:
                continue
            rel = a["berth_day"] + unload_days(a["ton"], UTPH)
            events.append((rel, a["crude"], a["ton"]))
    events.sort(key=lambda x: x[0])

    weights = defaultdict(list); order = []

    def schedulable():
        return [c for c in proc
                if remaining[c] > 1e-6 and avail[c] >= min_batch_ton(c) - 1e-6]

    ei = 0
    while True:
        cand = schedulable()
        while cand:
            # 最落后优先；同值按分母降序（大档优先）兜底
            cstar = max(cand, key=lambda c: (remaining[c] / denom[c], denom[c]))
            est = remaining[cstar] / denom[cstar]
            batch = min(denom[cstar], avail[cstar], remaining[cstar])
            weights[cstar].append(est)
            order.append((cstar, len(weights[cstar]), est))
            remaining[cstar] -= batch
            avail[cstar] -= batch
            cand = schedulable()
        if ei >= len(events):
            break
        cur_h = events[ei][0]
        while ei < len(events) and events[ei][0] == cur_h:
            _, c, ton = events[ei]
            avail[c] += ton
            ei += 1
    return weights, order

def build_units(P):
    """
    单元 = 单炼（1种油）或掺炼（主力油+掺炼油，共用同一CDU时段）。
    comps 列表：[{"crude":..., "load_hr":...(模板原值)}]  ← 绝对负荷
    rate = Σload_hr * 24（吨/天）
    """
    proc = dict(P["PROC"]); rate_hr = P["RATE_HR"]
    units = []; blended = set()

    for rid, rc in P["RECIPES"].items():
        main = rc["main_crude"]
        cap = rc["total_cap_hr"]              # 总负荷上限 = 主力油标准负荷
        comps = [{"crude": main, "is_main": True}]   # 主力油负荷=cap-Σ掺炼,由求解器定
        for b in rc["blends"]:
            comps.append({"crude": b["crude"], "is_main": False, "cands": b["cands"]})
        ton_est = sum(proc.get(c["crude"], 0) for c in comps)   # 片段cap参考
        units.append({"uid": rid, "comps": comps, "ton": ton_est,
                      "rate": cap * 24, "rate_hr": cap, "total_cap_hr": cap,
                      "is_blend": True})
        for c in comps: blended.add(c["crude"])

    can_single = P["CAN_SINGLE"]
    for c, v in proc.items():
        if not can_single.get(c, True):
            if c not in blended:
                print(f"  ⚠ {c} 不可单炼且不在当月任何掺炼配方中 → 本月无加工出路")
            continue                       # 不可单炼:只能走掺炼,不建单炼单元
        hr = rate_hr.get(c, DEF_RATE_HR)   # 可单炼(含掺炼主力油BZ):仍建单炼单元
        units.append({"uid": c,
                      "comps": [{"crude": c, "load_hr": hr, "is_main": True}],
                      "ton": v, "rate": hr * 24, "rate_hr": hr,
                      "is_blend": False})
    return units

def avail_sources(unit, P):
    """
    可得性来源：(release_hour, avail_ton)
      - 初始库存：release_hour=0
      - 到港：release_hour = (berth_day + 卸船天数) * 24（卸完次日0点可供油）

    【piece 生成的输入源】返回该单元各油种的"可得油量"按可得时刻切分成的来源列表。
    每个 (rel, vol) 来源后续在 cp_schedule 中折算成 cap_h=vol/产能 工时，并生成 piece。
    越月末(rel>H)的到港丢弃——本月排不了，作下月结转。
    """
    # 变量定义: init_by_crude (dict): 各油种月初可用库存(吨) = 库存 - 底油，按罐汇总
    init_by_crude = defaultdict(float)
    for t, d in P["INIT"].items():
        if d["crude"]: init_by_crude[d["crude"]] += d["ton"] - P["HEEL"].get(t, 0.0)

    H = P["DAYS"] * 24
    crudes = [c["crude"] for c in unit["comps"]]
    if not unit["is_blend"]:
        # ── 【单炼来源】主力油：月初库存(rel=0) + 各到港(rel=卸完时刻) ──
        c = crudes[0]
        # 变量定义: src (list): [(rel,vol),...] 来源列表，首个为月初库存
        src = [(0, init_by_crude.get(c, 0.0))]
        for a in sorted([x for x in P["ARRIVALS"] if x["crude"] == c and x["berth_day"]],
                        key=lambda x: x["berth_day"]):
            # 变量定义: rel (float): 该船油最早可供料时刻 = (靠泊日+卸船天数)×24
            rel = arrival_rel_h(a["berth_day"], a["ton"], P)
            if rel > H:   # 卸船完成越月末，本月无法加工，作下月结转（与 pipeline_arrivals_by_day 一致）
                continue   # 作用: 越月末到港丢弃，不生成来源
            src.append((rel, a["ton"]))
        return src

    # ── 【掺炼来源】主力油+掺炼组分合并：月初库存优先，不足部分等最迟组分到港 ──
    # 简化：取最迟组分的到港时间作为 rest 的可得时刻
    # 变量定义:
    #   init0 (float): 各组分月初可用库存之和
    #   rest  (float): 月初库存不足以覆盖单元目标量时，需等到港补足的缺口
    init0 = sum(init_by_crude.get(c, 0.0) for c in crudes)
    src = [(0, min(init0, unit["ton"]))]   # 作用: 月初可用的部分，rel=0 立即可加工
    rest = max(0.0, unit["ton"] - init0)   # 作用: 月初不够的缺口，需等到港
    if rest > 0:
        # 变量定义: arr_times (list): 本月内各组分到港的可供时刻集合
        arr_times = []
        for c in crudes:
            for a in [x for x in P["ARRIVALS"] if x["crude"] == c and x["berth_day"]]:
                rel = arrival_rel_h(a["berth_day"], a["ton"], P)
                if rel > H:   # 越月末到港不计入本月可供
                    continue
                arr_times.append(rel)
        if arr_times:   # 仅当本月内有到港可供 rest 时才加该来源
            src.append((max(arr_times), rest))   # 作用: 缺口 rest 等最迟组分到齐后可供，rel取最晚到港时刻
    return src

# ============================================================
# 2.5  Phase 0：预处理来船分配 + 生成腾容约束
# ============================================================
def preprocess_arrivals(P):
    """
    逐艘来船按靠泊日顺序分配目标罐，返回：
      ship_assignments: {ship_index: {"tank": str, "is_g": bool}}
      cp_constraints:   [(crude, need_tons, deadline_hour), ...]
    """
    G = P["GTANKS"]; T = P["TTANKS"]
    CAP = P["CAP"]; ALLOW = P["ALLOW"]
    rate_hr = P["RATE_HR"]
    can_single = P["CAN_SINGLE"]
    SW = P.get("SWITCH_RESID_TON", SWITCH_RESID_TON)  # 换油残余阈值：旧油可用残余≤SW即可换油
    CLEAR_MARGIN = P.get("CLEAR_MARGIN", CLEAR_MARGIN_TON)  # 腾容安全余量：约束多清此量防零余量装不下

    def with_margin(base, clearable_cap, c, limit_h):
        """腾容约束取最小需求 base，封顶到可清量、且不超清罐期限(小时)；否则回退 base。"""
        N = min(base, clearable_cap)
        return N if hours_to_clear(c, N) <= limit_h else base

    g_est = {g: {"crude": P["INIT"].get(g, {}).get("crude"),
                 "ton":   P["INIT"].get(g, {}).get("ton", 0.0)}
             for g in G}

    def gE(g, c):
        a = str(ALLOW.get(g, "任意")).strip()
        if a == "进口": return P["ORIGIN"].get(c) == "imp"
        if a == "国内": return P["ORIGIN"].get(c) == "dom"
        return a in ("任意", "") or norm(c) == norm(a)

    def eT(t, c):  # T罐国内/进口归属校验（与 tank_fill.eT 同源）
        a = str(ALLOW.get(t, "任意"))
        if a == "进口": return P["ORIGIN"].get(c) == "imp"
        if a == "国内": return P["ORIGIN"].get(c) == "dom"
        return True

    def hours_to_clear(c, tons):
        """清 tons 吨该油所需【小时数】= 加工工时 tons/负荷；与 CP-SAT 侧 en≤deadline_h 同为小时口径。"""
        r = rate_hr.get(c, DEF_RATE_HR)
        return tons / r

    g_est_T = {t: {"crude": P["INIT"].get(t, {}).get("crude"),
                   "ton":   P["INIT"].get(t, {}).get("ton", 0.0)}
               for t in T}

    def find_t_tank(c, ton):
        t1 = [t for t in T if eT(t, c) and g_est_T.get(t, {}).get("crude") == c
              and CAP[t] - g_est_T[t]["ton"] >= ton - 1e-6]
        if t1: return t1[0]
        t2 = [t for t in T if eT(t, c) and g_est_T.get(t, {}).get("ton", 0) < 1e-6
              and CAP[t] >= ton - 1e-6]
        if t2: return t2[0]
        return None  # 无满足条件的T罐（含国内/进口归属不符）

    ship_assignments = {}
    cp_constraints = []
    no_process_windows = []  # [(crude, start_h, end_h)] 专用罐卸油停产窗口
    g_cleared = defaultdict(float)  # 每个G罐因P2约束累计清除的旧油量

    skipped = [a["crude"] for a in P["ARRIVALS"] if not a["berth_day"]]
    if skipped:
        print(f"  Phase0 跳过无靠泊日来船: {skipped}")

    indexed = [(i, a) for i, a in enumerate(P["ARRIVALS"]) if a["berth_day"]]
    indexed.sort(key=lambda x: x[1]["berth_day"])

    for ship_idx, a in indexed:
        C = a["crude"]; Q = a["ton"]; D = a["berth_day"]
        is_blend_only = not can_single.get(C, True)

        # 靠泊当日供料到 STOP_FEED_HOD 停料，旧油最晚供到此刻；之后准备+卸油。
        # 腾容须在停料前完成 → deadline=stop_feed_h(小时)。
        stop_feed_h = (D - 1) * 24 + P.get("STOP_FEED_HOD", STOP_FEED_HOD_DEF)
        deadline_h = stop_feed_h
        # 腾容判定按【小时】：need/负荷 ≤ deadline_h（=stop_feed_h），与 CP-SAT 侧腾容约束的
        # en ≤ deadline_h 口径一致；也与 tank_fill「先供后卸」在靠泊当日仍供料到 STOP_FEED_HOD
        # 一致（旧油可在 [0, stop_feed_h] 内炼掉，含靠泊当日先供的零头小时）。旧版按 //24 整天
        # 判定会丢当日零头小时、偏保守，致预分配误判"腾不出"（如第2日 CFD 需2罐时降级失败）。

        primary   = [g for g in G if g in P["BLEND_TANKS"]] if is_blend_only \
                    else [g for g in G if g in P["MAIN_TANKS"]]
        secondary = [g for g in G if g not in primary]

        assigned = None
        for is_primary, search_list in [(True, primary), (False, secondary)]:
            eligible = [g for g in search_list if gE(g, C)]

            # P1: 同种油 + 空间够（先遍历所有罐再降级，避免 P4 抢占空罐机会）
            for g in eligible:
                est = g_est[g]; free = CAP[g] - est["ton"]
                if est["crude"] == C and free >= Q - 1e-6:
                    assigned = {"tank": g, "is_g": True}
                    g_est[g]["ton"] += Q
                    break
            if assigned: break

            # P2: 同种油 + 需腾容 + 时间够 + 可清除量(到底油)足够
            for g in eligible:
                est = g_est[g]; free = CAP[g] - est["ton"]
                if est["crude"] == C and free < Q - 1e-6:
                    need = Q - free
                    clearable = est["ton"] - P["HEEL"].get(g, 0.0)   # 只能处理到底油
                    if need <= clearable + 1e-6 and hours_to_clear(C, need) <= deadline_h:
                        N = with_margin(need, clearable, C, deadline_h)  # 加安全余量
                        cp_constraints.append((C, N, deadline_h))
                        g_est[g]["ton"] -= need   # 空间记账仍按实际需求(余量只提约束下限)
                        g_cleared[g] += need
                        g_est[g]["ton"] += Q
                        assigned = {"tank": g, "is_g": True}
                        break
            if assigned: break

            # P3: 空罐 + allow 允许 + 容量够
            for g in eligible:
                est = g_est[g]
                if est["ton"] < 1e-6 and CAP[g] >= Q - 1e-6:
                    assigned = {"tank": g, "is_g": True}
                    g_est[g]["crude"] = C
                    g_est[g]["ton"] = Q
                    break
            if assigned: break

            # P4: 存其他油 C'，估算能在 D-1 前加工完
            # 仅在 primary 列表触发，避免跨类型（掺炼罐↔主力罐）换油
            if is_primary:
                for g in eligible:
                    est = g_est[g]
                    if est["crude"] and est["crude"] != C and est["ton"] > 1e-6:
                        c_prime = est["crude"]
                        # 换油只需把旧油可用残余清到换油阈值 SW（非清到底油），
                        # 与 tank_fill 换油判定（可用残余≤SW 即可接新油）一致。
                        # 漏减 SW 会过量要求腾容（need 比可炼上限还高 → 硬约束不可行）。
                        need_switch = (est["ton"] + g_cleared.get(g, 0)
                                       - P["HEEL"].get(g, 0.0) - SW)
                        if need_switch <= 0:
                            # 旧油估计量已低于阈值，无需额外约束
                            assigned = {"tank": g, "is_g": True, "switch": True,
                                        "old_crude": c_prime}
                            g_est[g]["crude"] = C
                            g_est[g]["ton"] = Q
                            break
                        if hours_to_clear(c_prime, need_switch) <= deadline_h:
                            # 换油可清上限=清到底油(而非仅到SW)，故余量封顶用到底油的量
                            clearable_sw = (est["ton"] + g_cleared.get(g, 0)
                                            - P["HEEL"].get(g, 0.0))
                            N = with_margin(need_switch, clearable_sw, c_prime, deadline_h)
                            cp_constraints.append((c_prime, N, deadline_h))
                            assigned = {"tank": g, "is_g": True, "switch": True,
                                        "old_crude": c_prime}
                            g_est[g]["crude"] = C
                            g_est[g]["ton"] = Q
                            break
                if assigned: break

        if not assigned:
            t = find_t_tank(C, Q)
            if t is None:
                print(f"  ⚠ Phase0 降级失败：{C} 船({Q:.0f}吨) 无G罐也无可用T罐")
            assigned = {"tank": t, "is_g": False}
            if t:
                g_est_T[t]["crude"] = C
                g_est_T[t]["ton"] = g_est_T[t].get("ton", 0) + Q

        ship_assignments[ship_idx] = assigned

        # 卸油停产窗口：收油罐在卸油/准备期间不能供料。窗口油种取该罐【当前所供油种 c'】：
        # 换油场景=被顶掉的旧油 old_crude；否则=船油种 C（与旧口径一致，保守安全）。
        # 船油种 C 本身由 CP-SAT rel(=feed_ready) 自动挡住。窗口区间=灰块[stop_feed_h, feed_ready_h]。
        if assigned.get("is_g"):
            c_block = assigned.get("old_crude") or C
            feed_ready_h = int(math.ceil(
                stop_feed_h + P.get("UNLOAD_PREP_H", UNLOAD_PREP_H_DEF)
                + unload_hours(Q, P.get("UNLOAD_TPH", DEF_UNLOAD_TPH))
                + P.get("FEED_PREP_H", FEED_PREP_H_DEF)))
            no_process_windows.append((c_block, stop_feed_h, feed_ready_h))

    return ship_assignments, cp_constraints, no_process_windows

def pipeline_pools(P):
    """油种按 can_single 归池，返回 (主力油集, 掺炼油集, 主力池容量, 掺炼池容量)。"""
    can_single = P.get("CAN_SINGLE", {})
    crudes = list(P["PROC"].keys())
    main = {c for c in crudes if can_single.get(c, True)}
    blend = {c for c in crudes if not can_single.get(c, True)}
    main_cap = sum(P["AVAIL_CAP"].get(t, 0.0) for t in P.get("MAIN_TANKS", set()))
    blend_cap = sum(P["AVAIL_CAP"].get(t, 0.0) for t in P.get("BLEND_TANKS", set()))
    return main, blend, main_cap, blend_cap


def pipeline_initials(P):
    """月初各油种 G池/T池 库存（净底油，可为负）。只计 PROC 中油种。"""
    heel = P.get("HEEL", {})
    gset = set(P["GTANKS"]); tset = set(P["TTANKS"])
    gp0 = defaultdict(float); tp0 = defaultdict(float)
    for t, d in P["INIT"].items():
        c = d.get("crude")
        if not c or c not in P["PROC"]:
            continue
        net = d.get("ton", 0.0) - heel.get(t, 0.0)
        if t in gset:
            gp0[c] += net
        elif t in tset:
            tp0[c] += net
    return dict(gp0), dict(tp0)


def pipeline_arrivals_by_day(P):
    """到港量按卸船完成日(release_day=berth_day+unload_days)聚合，release_day>DAYS 的剔除。"""
    out = defaultdict(lambda: defaultdict(float))
    for a in P["ARRIVALS"]:
        if not a.get("berth_day") or a["crude"] not in P["PROC"]:
            continue
        rel = a["berth_day"] + unload_days(a["ton"], P["UNLOAD_TPH"])
        if rel > P["DAYS"]:
            continue
        out[a["crude"]][rel] += a["ton"]
    return {c: dict(days) for c, days in out.items()}


def pipeline_feed_vars(m, days, pieces, combo_vars):
    """把连续加工片段按天切，返回 (feed, ohu, fh)。
    feed[crude][day] 当天加工量(吨); ohu[uid][day] 单元当天工时(h); fh[crude][day] 油种当天供料工时(h)。
    """
    H = days * 24
    # (a) 片段×日 重叠工时 ovh[piece_idx][day]，并锚定 Σ_d ovh == dur
    # ── 【片段×日重叠工时】把每个连续加工片段按天切片，计算它当天实际占用的工时 ──
    ovh = defaultdict(dict)
    for pi, p in enumerate(pieces):
        st, en, dur = p["st"], p["en"], p["dur"]   # st/en/dur: 片段开工/完工时刻/时长(均为 CP-SAT 变量)
        day_ovs = []
        for d in range(1, days + 1):
            L, R = (d - 1) * 24, d * 24              # L/R: 第 d 天起止时刻(常数)
            # 变量定义:
            #   lo (IntVar[0,H]): 片段开工与当天0点之较大者 = 当天占用起点
            #   hi (IntVar[0,H]): 片段完工与当天24点之较小者 = 当天占用终点
            #   raw (IntVar[-H,H]): 当天占用原始时长(可为负=当天不占用)
            #   ov  (IntVar[0,24]): 当天实际占用工时(负值钳为0)
            lo = m.NewIntVar(0, H, f"plo_{pi}_{d}"); m.AddMaxEquality(lo, [st, L])   # 作用: lo=max(st,L)，片段开工可能晚于当天0点
            hi = m.NewIntVar(0, H, f"phi_{pi}_{d}"); m.AddMinEquality(hi, [en, R])   # 作用: hi=min(en,R)，片段完工可能早于当天24点
            raw = m.NewIntVar(-H, H, f"praw_{pi}_{d}"); m.Add(raw == hi - lo)        # 作用: raw=当天占用终点-起点(未裁剪)
            ov = m.NewIntVar(0, 24, f"povh_{pi}_{d}"); m.AddMaxEquality(ov, [raw, 0]) # 作用: ov=max(raw,0)，负时长(不占用当天)归零
            ovh[pi][d] = ov; day_ovs.append(ov)
        # ── 【工时锚定】当天占用工时之和必须等于片段总时长，防止压低 ovh 逃避日级约束 ──
        m.Add(sum(day_ovs) == dur)   # 锚定：防止压低 ovh 逃避约束              # 作用: Σ_d ov == dur，保证片段工时全部分摊到各天

    # (b) 单元每日工时 ohu[uid][day]
    pieces_by_uid = defaultdict(list)
    for pi, p in enumerate(pieces):
        pieces_by_uid[p["u"]["uid"]].append(pi)
    ohu = defaultdict(dict)
    for uid, pidxs in pieces_by_uid.items():
        for d in range(1, days + 1):
            # 变量定义: oh (IntVar[0,24]): 单元 uid 在第 d 天的总工时
            oh = m.NewIntVar(0, 24, f"ohu_{uid}_{d}")
            # ── 【单元日工时汇总】单元当日工时 = 其所有片段当天重叠工时之和(同单元片段不重叠) ──
            m.Add(oh == sum(ovh[pi][d] for pi in pidxs))   # 作用: 聚合该单元各片段当天占用工时
            ohu[uid][d] = oh

    # (c) feed[crude][day] 与 fh[crude][day]
    feed_terms = defaultdict(lambda: defaultdict(list))   # crude->day->[线性项]
    fh_terms = defaultdict(lambda: defaultdict(list))     # crude->day->[工时项]
    for uid, pidxs in pieces_by_uid.items():
        u = pieces[pidxs[0]]["u"]
        if not u["is_blend"]:
            c = u["comps"][0]["crude"]; lh = u["comps"][0]["load_hr"]
            for d in range(1, days + 1):
                feed_terms[c][d].append(lh * ohu[uid][d])
                fh_terms[c][d].append(ohu[uid][d])
        else:
            cvars = combo_vars[uid]   # [(pk, loadmap)]
            # 该单元喂到的油种集合 = 任一 combo 的 loadmap 键（各 combo 键相同）
            unit_crudes = set()
            for _pk, loadmap in cvars:
                unit_crudes.update(loadmap.keys())
            for d in range(1, days + 1):
                # w[k][d] = ohu if pk_k else 0
                # ── 【blend档位工时门控】用 big-M 把"选中档位的工时"提取出来；仅当选中档 pk=1 时 w 取单元工时，否则 w=0 ──
                wkd = []
                for ki, (pk, loadmap) in enumerate(cvars):
                    # 变量定义: w (IntVar[0,24]): 第 ki 档候选配方选中时该单元当天工时；pk(BoolVar): 该档是否选中
                    w = m.NewIntVar(0, 24, f"w_{uid}_{ki}_{d}")
                    m.Add(w <= 24 * pk)                              # 作用: pk=0 时上界为0 → w=0(档未选不占工时)
                    m.Add(w <= ohu[uid][d])                          # 作用: w 不超过单元当日总工时
                    m.Add(w >= ohu[uid][d] - 24 * (1 - pk))          # 作用: pk=1 时下界逼 w≥ohu → w=ohu(选中档承担全部工时)
                    wkd.append((w, loadmap))
                for c in unit_crudes:
                    feed_terms[c][d].append(sum(loadmap.get(c, 0) * w for w, loadmap in wkd))
                fh_for_unit = sum(w for w, _ in wkd)   # = ohu（恰选一档）
                for c in unit_crudes:
                    fh_terms[c][d].append(fh_for_unit)

    # 计算各油种最大日加工量上界（用于 IntVar 域）
    # 单个单元每天最多 24h，最大 load_hr 取片段最大值；多单元同油累加
    max_lh_per_crude = defaultdict(float)
    n_units_per_crude = defaultdict(int)
    for uid, pidxs in pieces_by_uid.items():
        u = pieces[pidxs[0]]["u"]
        if not u["is_blend"]:
            c = u["comps"][0]["crude"]; lh = u["comps"][0]["load_hr"]
            max_lh_per_crude[c] = max(max_lh_per_crude[c], lh)
            n_units_per_crude[c] += 1
        else:
            for _pk, loadmap in combo_vars.get(uid, []):
                for c, lh in loadmap.items():
                    max_lh_per_crude[c] = max(max_lh_per_crude[c], lh)
                    n_units_per_crude[c] += 1

    feed = defaultdict(dict); fh = defaultdict(dict)
    all_crudes = set(feed_terms) | set(fh_terms)
    for c in all_crudes:
        max_feed_day = int(max_lh_per_crude.get(c, 1000) * 24 * max(n_units_per_crude.get(c, 1), 1))
        max_fh_day = 24 * max(n_units_per_crude.get(c, 1), 1)
        for d in range(1, days + 1):
            # 变量定义: fv (IntVar[0,max_feed_day]): 油种 c 当天加工量(吨)
            fv = m.NewIntVar(0, max_feed_day, f"feed_{c}_{d}")
            # ── 【日加工量汇总】油种当日加工量 = 各单元对 c 的当日贡献之和 ──
            m.Add(fv == sum(feed_terms[c][d]) if feed_terms[c][d] else 0)   # 作用: 把各单元当天对 c 的加工项聚合成总量
            feed[c][d] = fv
            # 变量定义: fhv (IntVar[0,max_fh_day]): 油种 c 当天供料工时(h)
            fhv = m.NewIntVar(0, max_fh_day, f"fh_{c}_{d}")
            # ── 【日供料工时汇总】油种当日供料工时 = 各单元对 c 的当日工时之和 ──
            m.Add(fhv == sum(fh_terms[c][d]) if fh_terms[c][d] else 0)      # 作用: 聚合各单元当天对 c 的供料工时
            fh[c][d] = fhv
    return dict(feed), {k: dict(v) for k, v in ohu.items()}, dict(fh)


def _intexpr(v):
    """常数→int；已是 CP-SAT 变量/表达式→原样返回。"""
    if isinstance(v, (int,)):
        return v
    if isinstance(v, float):
        return int(round(v))
    return v


def add_pipeline_constraints(m, P, pieces, combo_vars):
    """按天、油种级多商品流约束（管道时序/能力 + 同罐供受串行）。策略X：全硬。"""
    days = P["DAYS"]
    main_crudes, blend_crudes, main_cap, blend_cap = pipeline_pools(P)
    gp0, tp0 = pipeline_initials(P)
    arr = pipeline_arrivals_by_day(P)
    pipe_cap_day = int(round(P.get("TRANSFER_TPH", 700) * 24))
    crudes = list(P["PROC"].keys())

    feed, ohu, fh = pipeline_feed_vars(m, days, pieces, combo_vars)

    BIG = days * 24 * 1000
    Gp = defaultdict(dict); Tp = defaultdict(dict)
    pin = defaultdict(dict); dArr = defaultdict(dict); tArr = defaultdict(dict)
    gpos = defaultdict(dict)  # max(0, Gp)，用于"供料不超前一日可用"（可用可为负）

    def gp_prev(c, d):
        return gp0.get(c, 0.0) if d == 1 else Gp[c][d - 1]
    def gpos_prev(c, d):
        # 前一日可用的非负部分；d==1 用 max(0, gp0)
        if d == 1:
            return max(0.0, gp0.get(c, 0.0))
        return gpos[c][d - 1]

    for c in crudes:
        for d in range(1, days + 1):
            a_cd = int(round(arr.get(c, {}).get(d, 0.0)))   # a_cd: 油种 c 第 d 天到港量(常数)
            # 变量定义:
            #   da (IntVar[0,a_cd]): 当天到港直卸入G罐的量(吨)
            #   ta (IntVar[0,a_cd]): 当天到港入T罐暂存的量(吨)
            da = m.NewIntVar(0, a_cd, f"dArr_{c}_{d}")
            ta = m.NewIntVar(0, a_cd, f"tArr_{c}_{d}")
            # ── 【到港量分流】当日到港量必须全部分配到直卸(da)或暂存(ta)，无遗漏 ──
            m.Add(da + ta == a_cd)                            # 作用: da+ta=到港总量，保证到港量守恒
            # 变量定义:
            #   pn (IntVar[0,pipe_cap_day]): 当天 T→G 管道输送量(吨)
            #   g  (IntVar[-BIG,BIG]):       G 罐区净库存(可负，表示缺口)
            #   t  (IntVar[0,BIG]):          T 罐区库存
            #   gp (IntVar[0,BIG]):          G 罐区可用库存的非负部分 = max(g,0)
            pn = m.NewIntVar(0, pipe_cap_day, f"pin_{c}_{d}")
            g = m.NewIntVar(-BIG, BIG, f"Gp_{c}_{d}")
            t = m.NewIntVar(0, BIG, f"Tp_{c}_{d}")
            gp = m.NewIntVar(0, BIG, f"gpos_{c}_{d}")
            dArr[c][d] = da; tArr[c][d] = ta; pin[c][d] = pn
            Gp[c][d] = g; Tp[c][d] = t; gpos[c][d] = gp
            # ── 【G罐库存递推】G罐当日净库存 = 前日可用 + 直卸 + 管道输送 − 当天加工 ──
            m.Add(g == _intexpr(gp_prev(c, d)) + da + pn - feed[c][d])   # 作用: G罐库存逐日滚动，扣加工、增收油
            # ── 【T罐库存递推】T罐当日库存 = 前日库存 + 暂存量 − 管道输出量 ──
            m.Add(t == _intexpr(tp0.get(c, 0.0) if d == 1 else Tp[c][d - 1]) + ta - pn)  # 作用: T罐库存逐日滚动，暂存进、输送出
            # ── 【可用库存非负化】gp = max(g, 0)：把可负的净库存钳为非负可用量 ──
            m.AddMaxEquality(gp, [g, 0])                        # 作用: 取 G 罐净库存与0较大者，供"不超前"约束引用
            # ── 【供料不超前】当天加工 ≤ 前一日可用(非负) + 当天直卸，不能消耗尚未到港的油 ──
            m.Add(feed[c][d] <= _intexpr(gpos_prev(c, d)) + da)  # 作用: 加工量受限于已可得油量(前日结余+当日直卸)
            # ── 【同罐供受串行】卸油(fh)与管道输送(pin)共用罐/管道，工时不能叠加超管道日能力 ──
            m.Add(700 * fh[c][d] + pn <= pipe_cap_day)          # 作用: 供料工时×管速 + 输送量 ≤ 日能力，保证同罐供受不冲突

    # 共享：池容量
    for d in range(1, days + 1):
        # ── 【主料池容量】当日主力油种在G罐总库存 ≤ 主料罐区总容量 ──
        if main_crudes:
            m.Add(sum(Gp[c][d] for c in main_crudes) <= int(round(main_cap)))   # 作用: 主力油种聚合库存不超主料罐区容量
        # ── 【掺炼池容量】当日掺炼油种在G罐总库存 ≤ 掺炼罐区总容量 ──
        if blend_crudes:
            m.Add(sum(Gp[c][d] for c in blend_crudes) <= int(round(blend_cap))) # 作用: 掺炼油种聚合库存不超掺炼罐区容量
        # 共享：单管道日能力
        # ── 【单管道日能力】所有油种当日管道输送总量 ≤ 管道日输送能力 ──
        m.Add(sum(pin[c][d] for c in crudes) <= pipe_cap_day)   # 作用: 共用管道各油种输送量之和受日能力上限约束

    return {"feed": feed, "Gp": Gp, "Tp": Tp, "pin": pin, "dArr": dArr, "tArr": tArr}


def add_gpool_constraints(m, P, pieces, units, combo_vars, ship_assignments, no_process_windows):
    """G 池日级库存模型：按油种按天追踪 G 罐区聚合库存（变量化），确保 CP-SAT 排出的批次
    在 tank_fill 仿真中无供料缺口。

    核心约束（均为 CP-SAT 变量约束）：
      g_inv[c][d+1] = g_inv[c][d] - feed[c][d] + g_arrival[c][d]
      feed[c][d] ≤ g_inv[c][d]                          不能抽超库存
      feed[c][d] ≤ (24 - no_feed_h[c][d]) × load_hr_c   no_feed 限容
      feed[c][d] == Σ 各单元对 c 的日加工贡献            精确关联到 piece
    """
    G = P["GTANKS"]
    CAP = P["CAP"]; ALLOW = P.get("ALLOW", {}); HEEL = P.get("HEEL", {})
    ORIGIN = P.get("ORIGIN", {}); RATE_HR = P.get("RATE_HR", {})
    DAYS = P["DAYS"]; H = DAYS * 24
    proc = P["PROC"]

    # ---- 1. 预计算常量 ----
    g_init = defaultdict(float)
    for t in G:
        d = P["INIT"].get(t, {})
        c = d.get("crude")
        if c:
            g_init[c] += d.get("ton", 0.0) - HEEL.get(t, 0.0)
    # T 罐初始库存也纳入 G 池（T→G 传输由 tank_fill 执行，CP-SAT 侧视为可用）
    for t in P["TTANKS"]:
        d = P["INIT"].get(t, {})
        c = d.get("crude")
        if c and c in P["PROC"]:
            g_init[c] += d.get("ton", 0.0) - HEEL.get(t, 0.0)

    def gE(t, c):
        a = str(ALLOW.get(t, "任意")).strip()
        if a == "进口": return ORIGIN.get(c) == "imp"
        if a == "国内": return ORIGIN.get(c) == "dom"
        return a in ("任意", "") or norm(c) == norm(a)
    g_pool_cap = {}
    for c in proc:
        g_pool_cap[c] = int(sum(CAP[t] for t in G if gE(t, c)))

    g_arrival = defaultdict(lambda: defaultdict(float))
    for idx, a in enumerate(P["ARRIVALS"]):
        if not a.get("berth_day") or a["crude"] not in proc:
            continue
        # G 罐到港直接计入；T 罐到港也计入（T→G 传输由 tank_fill 执行）
        g_arrival[a["crude"]][a["berth_day"]] += a["ton"]

    no_feed_h = defaultdict(lambda: defaultdict(float))
    for c, s, e in (no_process_windows or []):
        s = max(0, int(s)); e = min(H, int(e))
        d = s // 24 + 1
        while d <= DAYS and (d - 1) * 24 < e:
            ov = min(e, d * 24) - max(s, (d - 1) * 24)
            if ov > 0:
                no_feed_h[c][d] += ov
            d += 1

    # ---- 2. piece×day 重叠变量 ----
    pieces_by_uid = defaultdict(list)
    for pi, p in enumerate(pieces):
        pieces_by_uid[p["u"]["uid"]].append(pi)

    ovh = {}
    for pi, p in enumerate(pieces):
        st, en, dur = p["st"], p["en"], p["dur"]   # st/en/dur: 片段开工/完工时刻/时长(均为 CP-SAT 变量)
        day_ovs = []
        for d in range(1, DAYS + 1):
            L, R = (d - 1) * 24, d * 24              # L/R: 第 d 天起止时刻(常数)
            # 变量定义:
            #   lo  (IntVar[0,H]):  片段开工与当天0点之较大者=当天占用起点
            #   hi  (IntVar[0,H]):  片段完工与当天24点之较小者=当天占用终点
            #   raw (IntVar[-H,H]): 当天占用原始时长(可为负)
            #   ov  (IntVar[0,24]): 当天实际占用工时(负值钳为0)
            lo = m.NewIntVar(0, H, f"glo_{pi}_{d}"); m.AddMaxEquality(lo, [st, L])    # 作用: lo=max(st,L)
            hi = m.NewIntVar(0, H, f"ghi_{pi}_{d}"); m.AddMinEquality(hi, [en, R])    # 作用: hi=min(en,R)
            raw = m.NewIntVar(-H, H, f"graw_{pi}_{d}"); m.Add(raw == hi - lo)         # 作用: raw=当天占用终点-起点
            ov = m.NewIntVar(0, 24, f"govh_{pi}_{d}"); m.AddMaxEquality(ov, [raw, 0]) # 作用: ov=max(raw,0)，负值归零
            ovh[(pi, d)] = ov
            day_ovs.append(ov)
        # ── 【工时锚定】各天占用工时之和必须等于批次时长，防止压低 ovh 逃避约束 ──
        m.Add(sum(day_ovs) == dur)   # 锚定：Σ 日重叠 == 批次时长             # 作用: Σ_d ov == dur，工时全分摊到各天

    ohu = {}
    for uid, pidxs in pieces_by_uid.items():
        ohu[uid] = {}
        for d in range(1, DAYS + 1):
            # 变量定义: oh (IntVar[0,24]): 单元 uid 在第 d 天的总工时
            oh = m.NewIntVar(0, 24, f"gohu_{uid}_{d}")
            # ── 【单元日工时汇总】单元当日工时 = 其所有片段当天重叠工时之和 ──
            m.Add(oh == sum(ovh[(pi, d)] for pi in pidxs))   # 作用: 聚合该单元各片段当天占用工时
            ohu[uid][d] = oh

    # ---- 3. feed[c][d] == Σ 各单元贡献（精确等式）----
    # 先收集各油种各天的贡献项列表，最后建等式
    feed_terms = defaultdict(lambda: defaultdict(list))  # c -> d -> [线性项]

    for u in units:
        uid = u["uid"]
        if not u.get("is_blend"):
            c = u["comps"][0]["crude"]
            lh = int(u["comps"][0].get("load_hr") or RATE_HR.get(c, DEF_RATE_HR))
            for d in range(1, DAYS + 1):
                feed_terms[c][d].append(lh * ohu[uid][d])
        else:
            cvars = combo_vars.get(uid, [])
            for d in range(1, DAYS + 1):
                for ki, (pk, lm) in enumerate(cvars):
                    # 变量定义: w (IntVar[0,24]): 第 ki 档选中时该单元当天工时; pk(BoolVar): 该档是否选中
                    w = m.NewIntVar(0, 24, f"gw_{uid}_{ki}_{d}")
                    # ── 【blend档位工时门控】big-M 把选中档工时提取出来；pk=1 时 w=ohu，pk=0 时 w=0 ──
                    m.Add(w <= 24 * pk)                              # 作用: pk=0 时上界为0 → w=0(未选档不占工时)
                    m.Add(w <= ohu[uid][d])                          # 作用: w 不超过单元当日总工时
                    m.Add(w >= ohu[uid][d] - 24 * (1 - pk))          # 作用: pk=1 时下界逼 w≥ohu → w=ohu(选中档承担全部工时)
                    for c, lv in lm.items():
                        if lv > 0:
                            feed_terms[c][d].append(int(lv) * w)

    feed = {}
    for c in proc:
        feed[c] = {}
        for d in range(1, DAYS + 1):
            # 变量定义: fv (IntVar[0,g_pool_cap]): 油种 c 当天加工量(吨)
            fv = m.NewIntVar(0, g_pool_cap.get(c, H * 1000), f"gfeed_{c}_{d}")
            terms = feed_terms[c][d]
            # ── 【加工量下界】油种当日加工量 ≥ 各单元贡献之和(松下界，配合上限容由库存约束收紧) ──
            m.Add(fv >= sum(terms) if terms else 0)   # 作用: fv 至少等于各单元对 c 的当日加工贡献和
            feed[c][d] = fv

    # ---- 4. G 池库存变量 + 递推 + 限容约束 ----
    for c in proc:
        cap_c = g_pool_cap.get(c, 0)
        if cap_c <= 0:
            continue
        lh_c = int(RATE_HR.get(c, DEF_RATE_HR))
        init_c = int(round(g_init.get(c, 0.0)))        # init_c: 油种 c 月初 G 罐可用库存(常数)

        g_inv = {}
        for d in range(1, DAYS + 1):
            # 变量定义: g_inv[d] (IntVar[0,cap_c]): 油种 c 第 d 天 G 罐库存(吨)
            g_inv[d] = m.NewIntVar(0, cap_c, f"ginv_{c}_{d}")

        # 月初库存
        # ── 【月初库存锚定】第1天库存 = 月初可用库存(初始库存-底油) ──
        m.Add(g_inv[1] == init_c)                      # 作用: 固定首日库存为已知月初量

        for d in range(1, DAYS + 1):
            # ── 【不抽超库存】当天加工量 ≤ 当天库存，不能加工超过现有库存的油 ──
            m.Add(feed[c][d] <= g_inv[d])              # 作用: 核心约束，防止抽空G罐库存
            # 库存递推
            if d < DAYS:
                arr = int(round(g_arrival[c].get(d, 0.0)))   # arr: 第 d 天到港入G罐量(常数)
                # ── 【库存逐日递推】次日库存 = 当日库存 − 当天加工 + 当天到港 ──
                m.Add(g_inv[d + 1] == g_inv[d] - feed[c][d] + arr)  # 作用: 库存逐日滚动(扣加工、收到港)



def domestic_total_target(P):
    """国内可单炼油总量硬约束目标 = 各计划国内可单炼油目标之和，
    排除 EXTRA_DOMESTIC（补充库存油计入供给但不抬高门槛）。"""
    extra = P.get("EXTRA_DOMESTIC", set())
    return int(round(sum(
        tgt for c, tgt in P["PROC"].items()
        if P["ORIGIN"].get(c, "dom") != "imp"
        and P["CAN_SINGLE"].get(c, True)
        and c not in extra)))


# ============================================================
# 3. CP-SAT 排程
# ============================================================
def charge_cap(P, c):
    """油种 c 的【最大供料罐可用容量】= allow 兼容主力罐 usable 的最大值（无匹配退回全部主力罐）。
    工艺硬上限：供料期不中途补料 ⇒ **一段连续供料 ≤ 一罐可用容量** = 此值。
    供 `estimate_batches`(批次数下限) 与 `cp_schedule`(单批时长按罐容收紧) 共用，口径统一。

    【piece 生成的输入】返回值用于 cp_schedule 计算 max_h_u = 此值/负荷，
    即单批时长上限——保证一段连续供料不超过一个罐的可用容量，
    避免 tank_fill 仿真因供料期中途无法补料而亏油。"""
    # 变量定义:
    #   G     (list):  G 供料罐列表
    #   ALLOW (dict):  罐→允许油种限制("进口"/"国内"/"任意"/具体油种)
    #   ORIGIN(dict):  油种→产地("imp"/"dom")
    #   avail (dict):  罐→可用容量(预计算)，缺省取 罐容-底油
    #   CAP   (dict):  罐→额定容量
    #   heel  (dict):  罐→底油(不可用)
    #   blend (set):   掺炼专用罐集合(供料期不直接用)
    G = P["GTANKS"]; ALLOW = P.get("ALLOW", {}); ORIGIN = P.get("ORIGIN", {})
    avail = P.get("AVAIL_CAP", {}); CAP = P["CAP"]; heel = P.get("HEEL", {})
    blend = P.get("BLEND_TANKS", set())

    # ── 【罐可用容量】优先取预计算 AVAIL_CAP，缺省取 罐容-底油 ──
    def usable(g):
        return avail.get(g, CAP.get(g, 0.0) - heel.get(g, 0.0))   # 作用: 返回罐 g 的实际可用容量

    # ── 【罐油种兼容判定】罐的 ALLOW 限制是否允许收油种 cc ──
    def gok(g, cc):
        a = str(ALLOW.get(g, "任意")).strip()
        if a == "进口": return ORIGIN.get(cc) == "imp"                     # 作用: 罐限进口 → 仅进口油可用
        if a == "国内": return ORIGIN.get(cc) == "dom"                     # 作用: 罐限国内 → 仅国内油可用
        return a in ("任意", "") or norm(cc) == norm(a)                   # 作用: "任意"或匹配具体油种名 → 可用

    # 变量定义:
    #   main_tanks (list): 主力罐(非掺炼专用罐)列表
    #   cand       (list): 兼容油种 c 的主力罐(无匹配则退回全部主力罐)
    main_tanks = [g for g in G if g not in blend]
    cand = [g for g in main_tanks if gok(g, c)] or main_tanks   # 作用: 筛兼容罐，无则兜底用全部主力罐
    return max((usable(g) for g in cand), default=0.0)          # 作用: 取兼容罐中可用容量最大者=一段连续供料上限


def warm_start_hints(pieces, greedy_order, MIN_H, H):
    """按贪心批次顺序(greedy_batch_priority 的 order)推算每个 piece 的 warm-start 初值。

    返回 [(piece, pres, dur, st), ...]，覆盖**全部** pieces：
      - 被贪心选中的 piece：pres=1，st=max(rel, 同unit游标)，dur=max_h（超月末则裁剪，裁到<MIN_H则弃）
      - 未选中的 piece：pres=0, dur=0, st=rel
    纯函数：只读 piece 的 crude/rel/max_h/uid，不碰 CP 变量，便于单测。
    greedy_order 元素为 (crude, k, wk)。"""
    crude_pieces = defaultdict(list)
    for p in pieces:
        crude_pieces[p["u"]["comps"][0]["crude"]].append(p)
    for c in crude_pieces:
        crude_pieces[c].sort(key=lambda p: p["rel"])   # 同油种按可供油时刻排

    unit_cursor = defaultdict(int)   # {unit_id: 该罐最早可用时刻}（同 unit 批次不重叠）
    cnt = defaultdict(int)           # {crude: 已消费的 piece 数}
    decided = {}                     # id(piece) -> (piece, pres, dur, st)
    for crude, _k, _w in greedy_order:
        ps = crude_pieces.get(crude, [])
        idx = cnt[crude]; cnt[crude] += 1
        if idx >= len(ps):
            continue
        p = ps[idx]
        if id(p) in decided:
            continue
        uid = p["u"]["uid"]
        st = max(p["rel"], unit_cursor[uid])
        dur = p["max_h"]
        if st + dur > H:                       # 超月末：裁剪到月末
            dur = H - st
            if dur < MIN_H:                    # 放不下 → 标记不存在
                decided[id(p)] = (p, 0, 0, p["rel"])
                continue
        decided[id(p)] = (p, 1, dur, st)
        unit_cursor[uid] = st + dur
    return [decided.get(id(p), (p, 0, 0, p["rel"])) for p in pieces]


def objective_weights(late_max, pri_max):
    """按运行期上界自底向上算三级字典序权重：达标 ≫ 批次优先 ≫ 到港早开工（2026-07-15）。
    返回 (W_short, W_pri, W_late)：
      - 到港档权重 = 1（最低）；
      - 批次优先档权重 = 到港项最大可能值 + 1（保证批次优先每改善一个最小单位都盖过到港档全部波动）；
      - 达标档权重 = 批次优先档权重 × 批次优先项最大可能值 + 到港项最大可能值 + 1（保证欠 1 吨达标盖过其下两档全部波动）。
    late_max/pri_max 分别为「到港晚开工」「批次优先」两项的最大可能值（常数、建模期可得）。
    注：为压小权重量级、避免大系数拖慢求解，批次优先项按【天】计开工时刻（见 cp_schedule）。"""
    late_max = int(late_max); pri_max = int(pri_max)
    w_late = 1
    w_pri = late_max + 1
    w_short = w_pri * pri_max + late_max + 1
    return w_short, w_pri, w_late


class _CompliantStop(cp_model.CpSolverSolutionCallback):
    """达标即停回调（2026-07-15）：一旦某可行解「所有软达标油种缺口之和 = 0」（CP 计划层全部达标）
    立即停搜，不再为到港/批次优先继续优化。缺口列表为空（全硬约束）时，第一个可行解即停。
    判据直接读缺口变量，与目标权重无关。"""
    def __init__(self, shorts):
        super().__init__()
        self._shorts = list(shorts)

    def on_solution_callback(self):
        if sum(self.Value(s) for s in self._shorts) == 0:
            self.StopSearch()


def cp_schedule(P, ship_assignments=None, cp_constraints=None, time_limit=240,
                random_seed=None, no_process_windows=None, batch_cap=None,
                warm_start=False, hint_search=False, stop_on_compliant=False,
                benders_cuts=None):
    if cp_constraints is None:
        cp_constraints = []
    if no_process_windows is None:
        no_process_windows = []
    if benders_cuts is None:
        benders_cuts = []
    import itertools
    MIN_H = P.get("MIN_BATCH_H", MIN_BATCH_H)
    MAX_H = P.get("MAX_BATCH_H", MAX_BATCH_H)
    EDGE_MIN = P.get("EDGE_MIN_BATCH_H", EDGE_MIN_BATCH_H)  # 月初(st==0)/月末(en==H)边批最短时长
    units = build_units(P)
    H = P["DAYS"] * 24
    m = cp_model.CpModel()
    pieces = []
    proc = P["PROC"]

    # 各油种可得总量(初始库存+本月内到港) = 加工量上限(略超不可超可得)。
    # 越月末(次月才能供料加工)的到港不计入——见 available_by_crude。
    avail = available_by_crude(P)

    proc_terms = defaultdict(list)   # crude -> [线性加工量项]
    combo_vars = {}                  # blend uid -> [(present_bool, {crude:load_hr})]

    # 非禁连油种连续供料段限容开关：需 warm-start 引导方可高效求解（见下方约束块注释），
    # 故与 warm_start 耦合。关闭时下方 mtons/run 全不建模，模型与本特性引入前完全一致。
    stretch_on = bool(warm_start) and P.get("STRETCH_CAP", True)

    # ════════════════════════════════════════════════════════════════════
    # piece 生成：逐单元把"可得油量"折算成候选加工批次(piece)
    #   流程：单元 → 可得来源(初始库存+到港) → 来源合并 → 按罐容拆子批次 → 每子批一个 piece
    #   piece 的 dur 上界 this_h 在此静态钉死(按油量算)，CP-SAT 只能在 [0,this_h] 内选时长
    # ════════════════════════════════════════════════════════════════════
    for u in units:
        # 变量定义:
        #   u (dict): 加工单元，含 uid/comps(油种+负荷)/rate(吨/天)/is_blend
        #   rate_hr_u (float): 该单元每小时加工吨数 = 日产能/24
        rate_hr_u = u["rate"] / 24.0
        # 方案A（2026-07-05）：单批时长按【一罐可用容量】收紧——一段连续供料 ≤ 一罐容
        # （主料 tons = dur×main_load ≤ 主料罐可用容量 charge_cap）。避免 CP-SAT 排出超容段致仿真末日亏油。
        # 变量定义:
        #   main_c    (str):   该单元主力油种
        #   main_load (float): 主力油负荷(吨/h)，无则取单元均速
        #   cc        (float): 主力油最大供料罐可用容量(吨)——charge_cap 算
        #   max_h_u   (int):   单批时长上限(h)=min(MAX_H, 罐容/负荷)，保证单批不超一罐容
        main_c = u["comps"][0]["crude"]
        main_load = u["comps"][0].get("load_hr") or rate_hr_u
        cc = charge_cap(P, main_c)
        max_h_u = (max(MIN_H, min(MAX_H, int(cc / main_load)))   # 作用: 单批上限取罐容折算与全局上限的较小者
                   if (main_load > 0 and cc > 0) else MAX_H)      # 作用: 罐容/负荷无效时退回全局上限 MAX_H
        upieces = []   # 收集本单元生成的 piece（供后续 proc_terms/mtons 绑定）
        # ── 【可得来源】从 avail_sources 取该单元各油种的(可得时刻, 油量) ──
        # 来源合并：不足 MIN_H 的小来源（如到港残量）并入本单元最大来源，
        # 作为某个合法 48-156h 批次的一部分加工，避免被整段丢弃。
        # rel 取并入双方较早值（保证可加工、不延后；总量上限另由 avail 硬约束兜底）。
        # 变量定义:
        #   raw_src (list): [(rel,vol),...] 可得时刻rel(0=月初库存,>0=到港卸完后)与油量vol
        #   big     (list): 折算≥MIN_H 的来源(够独立成批)
        #   small   (list): 折算<MIN_H 的来源(太小,需并入大来源)
        raw_src = [(rel, vol) for rel, vol in avail_sources(u, P) if vol > 0]
        big   = [[rel, vol] for rel, vol in raw_src if round(vol / rate_hr_u) >= MIN_H]   # 作用: 按油量/产能折算工时,≥48h 为大来源
        small = [(rel, vol) for rel, vol in raw_src if round(vol / rate_hr_u) <  MIN_H]   # 作用: <48h 为小来源,单独成批非法
        # ── 【来源合并策略】按三种情况决定 sources(后续每个 source 生成 1+ piece) ──
        if big and small:
            # 情况1：有大有小 → 小来源并入容量最大的大来源(贪心)
            # 变量定义: tgt (list): [rel,vol] 容量最大的大来源(可变,被小来源累加)
            tgt = max(big, key=lambda x: x[1])          # 并入容量最大的来源
            for rel, vol in small:
                tgt[1] += vol; tgt[0] = min(tgt[0], rel)   # 作用: 油量累加、可得时刻取较早(保证可加工不延后)
            sources = [(r, v) for r, v in big]              # 作用: 合并后只剩大来源列表
        elif big:
            # 情况2：只有大来源 → 直接用
            sources = [(r, v) for r, v in big]
        elif raw_src and round(sum(v for _, v in raw_src) / rate_hr_u) >= EDGE_MIN:
            # 情况3：无大源但合计 ≥ 边批下限 → 合成一股(只能当月初/月末边批)
            # 无大源但合计 ≥ 边批下限 → 合成一股（this_h<MIN_H 时只能当月初/末边批，见下方 dur 条件约束）
            sources = [(min(r for r, _ in raw_src), sum(v for _, v in raw_src))]   # 作用: 合成单一来源,rel取最早
        else:
            # 情况4：合计 < 边批下限 → 无法成批,跳过(油量浪费)
            if raw_src:
                tot_h = round(sum(v for _, v in raw_src) / rate_hr_u)
                print(f"  ⚠ {u['uid']} 全部来源合计仅{tot_h}h<边批下限{EDGE_MIN}h，无法成批，跳过")
            sources = []
        # ── 【逐来源构造 piece】每个合并后的来源折算成工时，超单批上限则拆子批次 ──
        for si, (rel, vol) in enumerate(sources):
            # 变量定义:
            #   si  (int):   来源序号
            #   rel (float): 该来源油最早可得时刻(0=月初库存,>0=到港卸完次日0点)
            #   vol (float): 该来源油量(吨)
            #   cap_h (int): 该来源可加工工时 = 油量/小时产能
            cap_h = round(vol / rate_hr_u)
            # 超过【单批时长上限 max_h_u=min(MAX_H, 罐容/负荷)】则拆分为多个子批次（方案A）
            # 变量定义:
            #   n_sub  (int): 拆分成的子批次数 = ceil(工时/单批上限)
            #   base_h (int): 每个子批次基准工时(整除)
            n_sub = math.ceil(cap_h / max_h_u)   # 作用: 工时超单批上限则向上取整拆分
            base_h = cap_h // n_sub              # 作用: 均分基准工时
            # ── 【逐子批次创建 piece】每个子批次 = 一个候选加工批次，建 CP-SAT 变量 ──
            for sub_si in range(n_sub):
                # 变量定义: this_h (int): 该子批的 dur 上界(工时)。前 n-1 个取 base_h，末个取余数
                this_h = base_h if sub_si < n_sub - 1 else cap_h - base_h * (n_sub - 1)
                if this_h < EDGE_MIN: break      # 作用: 拆出的子批太小(<边批下限)则丢弃剩余
                uid_tag = f"{u['uid']}_{si}_{sub_si}"
                # 变量定义:
                #   pres (BoolVar):        该子批次是否实际排产(present)
                #   dur  (IntVar[0,this_h]): 该子批次加工时长(h)——上界 this_h 按油量静态钉死
                #   st   (IntVar[rel,H]):   开工时刻(h)，不早于油源可得时刻 rel
                #   en   (IntVar[0,H]):     完工时刻(h)
                #   iv   (OptionalInterval): pres 门控的可选区间变量(供 NoOverlap)
                pres = m.NewBoolVar(f"pre_{uid_tag}")
                dur  = m.NewIntVar(0, this_h, f"dur_{uid_tag}")
                st   = m.NewIntVar(rel, H, f"st_{uid_tag}")
                en   = m.NewIntVar(0, H, f"en_{uid_tag}")
                iv   = m.NewOptionalIntervalVar(st, dur, en, pres, f"iv_{uid_tag}")
                # ── 【缺席归零】未排产的子批次时长必须为0 ──
                m.Add(dur == 0).OnlyEnforceIf(pres.Not())   # 作用: pres=0 时 dur 强制为0，区间变量随之不存在
                # 最短时长：**月初(st==0)或月末(en==H)边批**可短到 EDGE_MIN，其余批次守 MIN_H。
                if EDGE_MIN < MIN_H:
                    # 变量定义:
                    #   at_start (BoolVar): 该子批是否为月初边批(st==0)
                    #   at_end   (BoolVar): 该子批是否为月末边批(en==H)
                    at_start = m.NewBoolVar(f"start_{uid_tag}")
                    # ── 【月初边批判定】at_start 真时 st=0，假时 st≥1 ──
                    m.Add(st == 0).OnlyEnforceIf(at_start)          # 作用: at_start=1 → 开工在月初0点(月初边批)
                    m.Add(st >= 1).OnlyEnforceIf(at_start.Not())    # 作用: at_start=0 → 开工非0点(非月初边批)
                    at_end = m.NewBoolVar(f"end_{uid_tag}")
                    # ── 【月末边批判定】at_end 真时 en=H，假时 en≤H−1 ──
                    m.Add(en == H).OnlyEnforceIf(at_end)            # 作用: at_end=1 → 完工在月末H点(月末边批)
                    m.Add(en <= H - 1).OnlyEnforceIf(at_end.Not())  # 作用: at_end=0 → 完工非月末(非月末边批)
                    # ── 【边批最短时长】排产的边批(月初或月末)时长可短至 EDGE_MIN ──
                    m.Add(dur >= EDGE_MIN).OnlyEnforceIf(pres)                       # 作用: pres=1 时 dur≥边批下限(覆盖所有排产批)
                    # ── 【非边批最短时长】非月初非月末的中间批时长须守 MIN_H ──
                    m.Add(dur >= MIN_H).OnlyEnforceIf(pres, at_start.Not(), at_end.Not())  # 作用: 排产且非边批 → dur≥常规下限
                else:
                    # ── 【统一下限】EDGE_MIN≥MIN_H 时退化为所有排产批守 MIN_H ──
                    m.Add(dur >= MIN_H).OnlyEnforceIf(pres)   # EDGE_MIN≥MIN_H：退化为统一下限  # 作用: pres=1 → dur≥MIN_H
                # 主料吨数(线性)：单炼绑 load*dur；掺炼在下方 combo 门控绑 main_load*dur。
                # 供“非禁连油种连续供料段限容”累加（一段相接连续供料 ≤ 一罐容 charge_cap）。
                # 仅特性开启时建模，关闭时 mtons=None、不向模型加任何变量。
                if stretch_on:
                    # 变量定义:
                    #   mtons   (IntVar): 该子批主料加工吨数(供连续段限容累加)
                    #   mt_load (float):  主料负荷(掺炼取总负荷,单炼取主力油负荷)
                    mt_load = u["total_cap_hr"] if u["is_blend"] else u["comps"][0]["load_hr"]
                    mtons = m.NewIntVar(0, int(math.ceil(mt_load * this_h)) + 1, f"mt_{uid_tag}")
                else:
                    mtons = None
                # ── 【piece 装载】把子批次所有变量与元信息打包存入 pieces ──
                # 变量定义: pd (dict): 一个 piece 的完整记录(单元/来源/变量/上限)
                pd = dict(u=u, si=si, sub_si=sub_si, rel=rel, is_arr=(rel > 0),
                          pres=pres, dur=dur, st=st, en=en, iv=iv, rate=u["rate"],
                          max_h=this_h, mtons=mtons)
                pieces.append(pd); upieces.append(pd)   # 作用: 全局 pieces + 本单元 upieces 各存一份
        # 变量定义: tot_dur (线性表达式): 该单元所有子批 dur 之和 = 单元总工时
        tot_dur = sum(p["dur"] for p in upieces)   # 该单元总工时(线性)

        if not u["is_blend"]:
            c = u["comps"][0]["crude"]
            lh = u["comps"][0]["load_hr"]
            proc_terms[c].append(lh * tot_dur)   # 工时×标准负荷
            if stretch_on:
                # ── 【单炼主料吨数】子批主料吨数 = 标准负荷 × 工时（供连续段限容累加）──
                for p in upieces:                # 主料吨数 = 负荷×工时（常数×变量，线性）
                    m.Add(p["mtons"] == lh * p["dur"])   # 作用: 线性绑定单炼子批主料加工量
        else:
            cap = u["total_cap_hr"]; main_c = u["comps"][0]["crude"]
            # 变量定义: td (IntVar[0,H]): 该掺炼单元总工时
            td = m.NewIntVar(0, H, f"td_{u['uid']}"); m.Add(td == tot_dur)   # 作用: 把单元总工时变量化，供各组分门控引用
            blend_comps = [c for c in u["comps"] if not c.get("is_main")]
            combos = list(itertools.product(*[c["cands"] for c in blend_comps])) or [()]
            cvars = []
            for ki, combo in enumerate(combos):
                main_load = cap - sum(combo)
                if main_load <= 0: continue          # 剔除掺炼超总负荷的非法组合
                # 变量定义: pk (BoolVar): 是否选中第 ki 档掺炼配方组合
                pk = m.NewBoolVar(f"cb_{u['uid']}_{ki}")
                loadmap = {main_c: main_load}
                for cc, lv in zip(blend_comps, combo): loadmap[cc["crude"]] = lv
                for cr, lv in loadmap.items():        # 选中该组合: 加工量=负荷常数×工时(线性)
                    # 变量定义: vv (IntVar): 选中该 combo 时组分 cr 的加工量
                    vv = m.NewIntVar(0, lv * H, f"v_{u['uid']}_{ki}_{cr}")
                    # ── 【掺炼组分加工门控】选中该档时组分加工量=负荷×工时，未选时为0 ──
                    m.Add(vv == lv * td).OnlyEnforceIf(pk)    # 作用: pk=1 → 该组分加工量=负荷×单元总工时
                    m.Add(vv == 0).OnlyEnforceIf(pk.Not())    # 作用: pk=0 → 该组分不加工(吨数为0)
                    proc_terms[cr].append(vv)
                if stretch_on:                   # 选中该 combo 时，主料吨数 = main_load×工时
                    # ── 【掺炼主料吨数门控】选中该 combo 时各子批主料吨数 = 主料负荷×工时 ──
                    for p in upieces:
                        m.Add(p["mtons"] == main_load * p["dur"]).OnlyEnforceIf(pk)  # 作用: pk=1 → 主料吨数=主料负荷×工时
                cvars.append((pk, loadmap))
            if not cvars:
                raise RuntimeError(f"配方{u['uid']} 所有负荷组合都超主力油总负荷,请检查候选值")
            # ── 【恰选一档】掺炼单元必须恰好选中一个配方组合 ──
            m.Add(sum(pk for pk, _ in cvars) == 1)    # 恰选一档   # 作用: 各候选档布尔和=1，保证选定唯一配方
            combo_vars[u["uid"]] = cvars

    if P.get("PIPELINE_MODEL", False):   # 选配，默认关（与 read_all 默认一致）
        add_pipeline_constraints(m, P, pieces, combo_vars)
    if P.get("GPOOL_MODEL", False):   # G 池日级库存模型，默认关（模型膨胀大，待优化求解时间）
        add_gpool_constraints(m, P, pieces, units, combo_vars, ship_assignments, no_process_windows)

    # 达标约束：进口油必须完成(硬)；国内主力油可互换但总量必达(硬)；掺炼专用油软约束
    shorts = []
    dom_main_pe = []   # 国内主力原油加工量项（用于总量硬约束）
    dom_main_tgt = domestic_total_target(P)   # 总量门槛（排除 EXTRA_DOMESTIC，不被其抬高）
    for c, tgt in proc.items():
        pe = sum(proc_terms[c]) if proc_terms[c] else 0   # pe: 油种 c 全月加工量(线性项之和)
        # ── 【加工量不超可得量】全月加工量 ≤ 该油种可得总量(初始库存+本月到港) ──
        m.Add(pe <= int(round(max(0.0, avail.get(c, 0)))))   # 作用: 防止加工超过实际可得油量
        origin_c = P["ORIGIN"].get(c, "dom")
        if origin_c == "imp":
            # 进口原油：硬约束，必须完成
            if avail.get(c, 0) >= tgt - 1:
                # ── 【进口油硬达标】可得量充足时加工量必须 ≥ 计划目标(硬) ──
                m.Add(pe >= int(round(tgt)))             # 作用: 强制完成进口油计划量，无缺口
            else:
                print(f"  ⚠ 进口油{c} 可得量{avail.get(c,0):.0f}<计划{tgt:.0f}，退为软约束")
                # 变量定义: sh (IntVar[0,tgt]): 进口油缺口量(可得不足时退软)
                sh = m.NewIntVar(0, int(tgt), f"short_{c}")
                # ── 【进口油软达标】可得不足时允许缺口 sh，加工量 ≥ 目标 − sh ──
                m.Add(pe >= int(round(tgt)) - sh)        # 作用: 缺口计入目标函数(大权重)，尽量少缺
                shorts.append(sh)
        elif P["CAN_SINGLE"].get(c, True):
            # 国内主力原油：个体软约束，但参与总量硬约束（允许油种间互换补量）
            # 变量定义: sh (IntVar[0,tgt]): 国内主力油个体缺口量
            sh = m.NewIntVar(0, int(tgt), f"short_{c}")
            # ── 【国内主力油软达标】单油种允许缺口，加工量 ≥ 目标 − sh ──
            m.Add(pe >= int(round(tgt)) - sh)            # 作用: 个体允许欠产，缺口计入目标函数
            shorts.append(sh)
            dom_main_pe.extend(proc_terms[c])
        else:
            # 国内掺炼专用油：软约束（不可单炼，无互换能力）
            # 变量定义: sh (IntVar[0,tgt]): 掺炼专用油缺口量
            sh = m.NewIntVar(0, int(tgt), f"short_{c}")
            # ── 【掺炼专用油软达标】加工量 ≥ 目标 − sh，缺口计入目标 ──
            m.Add(pe >= int(round(tgt)) - sh)            # 作用: 掺炼油允许欠产，尽量达标
            shorts.append(sh)
    # 国内主力油总量必须完成（允许油种间替代，但总量不减）
    if dom_main_pe and dom_main_tgt > 0:
        # ── 【国内主力总量硬达标】各主力油加工量之和 ≥ 总目标(允许油种间互换补量) ──
        m.Add(sum(dom_main_pe) >= dom_main_tgt)          # 作用: 个体可互换，但总量不可少(硬)

    # 连续不中断
    # ── 【批次不重叠】所有子批区间互不重叠，保证 CDU 连续加工不中断 ──
    m.AddNoOverlap([p["iv"] for p in pieces])            # 作用: 可选区间变量互斥，同一 CDU 同时刻只加工一批
    # 变量定义: mk (IntVar[0,H]): 全部批次的完工时刻(makespan)
    mk = m.NewIntVar(0, H, "makespan")
    # ── 【makespan 上界1】每批完工 ≤ makespan ──
    for p in pieces: m.Add(p["en"] <= mk)                # 作用: makespan 至少为最晚完工时刻
    # ── 【makespan 上界2】各批工时之和 ≤ makespan（无并行时串行总工时为下界）──
    m.Add(sum(p["dur"] for p in pieces) <= mk)           # 作用: makespan 至少为累计加工工时
    # ── 【makespan 月内】makespan ≤ 月度总小时，保证全月在月内排完 ──
    m.Add(mk <= H)                                       # 作用: 排产不能超出计划月长度

    # ---- 腾容约束（由 preprocess_arrivals 生成）----
    # ---- 腾容约束（由 preprocess_arrivals 生成）----
    # 腾容约束确保到港船卸油前，收油罐中的旧油已被加工腾出足够空间。
    # 这是硬约束：deadline 前必须累计加工 ≥ need 吨。
    # 如需临时禁用（如2月模板到港太早导致INFEASIBLE），注释整个 for 循环即可。

    def crude_of_piece(p):
        return p["u"]["comps"][0]["crude"]

    pieces_of_unit = defaultdict(list)
    for p in pieces:
        pieces_of_unit[id(p["u"])].append(p)

    for (C, N, T_deadline) in cp_constraints:
        contribs = []
        # (a) C 作为【单炼】单元主力油的片段：dur × 单炼标准负荷
        # ── 【腾容·单炼贡献门控】统计 deadline 前完工的单炼片段对油种 C 的加工量 ──
        for p in pieces:
            if crude_of_piece(p) != C:
                continue
            if p["u"].get("is_blend"):
                continue
            load_hr_C = p["u"]["comps"][0].get("load_hr") or P["RATE_HR"].get(C, p["u"]["rate_hr"])
            p_tag = f"{p['u']['uid']}_{p['si']}_{p.get('sub_si', 0)}_{T_deadline}"
            # 变量定义: early (BoolVar): 该片段是否在 deadline 前完工(en≤T_deadline)
            early = m.NewBoolVar(f"early_{p_tag}")
            # ── 【deadline 前完工判定】early 真则 en≤deadline，假则 en>deadline ──
            m.Add(p["en"] <= T_deadline).OnlyEnforceIf(early)      # 作用: early=1 → 该片段完工早于腾容截止
            m.Add(p["en"] > T_deadline).OnlyEnforceIf(early.Not())  # 作用: early=0 → 该片段完工晚于截止(不计入腾容)
            # 变量定义: contrib (IntVar): 该片段对腾容的贡献加工量
            contrib = m.NewIntVar(0, H * int(load_hr_C) + 1, f"contrib_{p_tag}")
            # ── 【单炼贡献门控】deadline 前完工则贡献=工时×负荷，否则为0 ──
            m.Add(contrib == p["dur"] * int(load_hr_C)).OnlyEnforceIf(early)  # 作用: early=1 → 贡献=该批加工吨数
            m.Add(contrib == 0).OnlyEnforceIf(early.Not())                    # 作用: early=0 → 贡献为0(不计入腾容)
            contribs.append(contrib)
        # (b) C 出现在掺炼单元（作主力油 或 掺炼组分）
        # ── 【腾容·掺炼贡献门控】统计 deadline 前完工的掺炼片段对油种 C 的加工量 ──
        for u in units:
            if not u.get("is_blend"):
                continue
            if not any(c["crude"] == C for c in u["comps"]):
                continue
            up = pieces_of_unit.get(id(u), [])
            if not up:
                continue
            td_early_terms = []
            for p in up:
                p_tag = f"{u['uid']}_{p['si']}_{p.get('sub_si', 0)}_{C}_{T_deadline}"
                # 变量定义: early (BoolVar): 该掺炼片段是否在 deadline 前完工
                early = m.NewBoolVar(f"bearly_{p_tag}")
                # ── 【deadline 前完工判定】同上：early 真则 en≤deadline ──
                m.Add(p["en"] <= T_deadline).OnlyEnforceIf(early)      # 作用: early=1 → 片段在腾容截止前完工
                m.Add(p["en"] > T_deadline).OnlyEnforceIf(early.Not())  # 作用: early=0 → 片段晚于截止(不贡献)
                # 变量定义: de (IntVar[0,H]): deadline 前完工时该片段的工时
                de = m.NewIntVar(0, H, f"bde_{p_tag}")
                # ── 【工时门控】deadline 前完工则 de=工时，否则为0 ──
                m.Add(de == p["dur"]).OnlyEnforceIf(early)   # 作用: early=1 → de=该片段工时
                m.Add(de == 0).OnlyEnforceIf(early.Not())    # 作用: early=0 → de=0(不计工时)
                td_early_terms.append(de)
            td_early = sum(td_early_terms)   # td_early: 该掺炼单元 deadline 前累计工时
            for ki, (pk, loadmap) in enumerate(combo_vars.get(u["uid"], [])):
                lv = int(loadmap.get(C, 0))   # lv: 该 combo 下 C 的负荷
                if lv <= 0:
                    continue
                # 变量定义: cuk (IntVar): 该 combo 下 C 的腾容贡献
                cuk = m.NewIntVar(0, lv * H, f"bcontrib_{u['uid']}_{ki}_{C}_{T_deadline}")
                # ── 【掺炼贡献门控】选中该 combo 时贡献=负荷×deadline前工时，否则0 ──
                m.Add(cuk == lv * td_early).OnlyEnforceIf(pk)   # 作用: pk=1 → 贡献=C负荷×deadline前累计工时
                m.Add(cuk == 0).OnlyEnforceIf(pk.Not())         # 作用: pk=0 → 该 combo 不贡献
                contribs.append(cuk)
        if contribs:
            # ── 【腾容硬约束】deadline 前累计加工 C 的量 ≥ 需腾空量 N(到港前需腾出罐容) ──
            m.Add(sum(contribs) >= int(round(N)))   # 作用: 保证到港前已加工足够旧油腾出收油空间
        else:
            print(f"  Warning: clearance({C},{N}t,{T_deadline}h) no pieces, skipped")

    # ---- 批次级可用量约束（候选1方式1）：每个 piece 开工时G罐可用量 ≥ 本批加工量 ----
    # g_avail(c, st_p) = g_init(c) + arrival_cum(c, st_p) − processed_cum(c, st_p)
    # 约束: q_p + processed_cum(c, st_p) ≤ g_init(c) + arrival_cum(c, st_p)
    # 预防版的 Benders 割：让 CP-SAT 一次就知道G罐可用量时序，减少迭代缺口。
    # 复用腾容约束的 early/contrib 建模模式（O(n²) pair，st_p 是变量无法静态排序）。
    # 开关：P["SEG_AVAIL_ON"]，默认 True；模板参数"分段可用量(是/否)"可关。
    if P.get("SEG_AVAIL_ON", True):
        _g_init = P.get("_G_INIT_AVAIL", {})
        _arr_segs = P.get("_ARRIVAL_CUM_SEGS", {})
        # 按油种分组 piece（用主力油种 crude_of_piece 归类）
        pieces_of_crude = defaultdict(list)
        for p in pieces:
            pieces_of_crude[crude_of_piece(p)].append(p)
        for c, c_pieces in pieces_of_crude.items():
            if len(c_pieces) < 2:
                continue   # 仅1个piece：由全月avail约束覆盖，无需pair
            g_init_c = int(round(_g_init.get(c, 0.0)))
            segs_c = _arr_segs.get(c, [])
            for p in c_pieces:
                # --- (1) processed_cum(c, st_p): Σ 同油种其它piece在st_p前完工的贡献 ---
                # ── 【批次可用量·已加工累计门控】统计 p 开工前同油种其它片段已完工的加工量 ──
                accum_terms = []
                for p2 in c_pieces:
                    if p2 is p:
                        continue
                    if p2["u"].get("is_blend"):
                        # 掺炼piece：每个combo门控累加该油种贡献（复用腾容 bcontrib 模式）
                        for ki, (pk, loadmap) in enumerate(combo_vars.get(p2["u"]["uid"], [])):
                            lv = int(loadmap.get(c, 0))
                            if lv <= 0:
                                continue
                            tag = f"sa_{p['u']['uid']}_{p['si']}_{p2['u']['uid']}_{p2['si']}_{ki}_{c}"
                            # 变量定义: early (BoolVar): p2 是否在 p 开工前完工(en≤p.st)
                            early = m.NewBoolVar(f"sae_{tag}")
                            # ── 【前序完工判定】early 真则 p2 在 p 开工前完工 ──
                            m.Add(p2["en"] <= p["st"]).OnlyEnforceIf(early)     # 作用: early=1 → p2 早于 p 开工完工
                            m.Add(p2["en"] > p["st"]).OnlyEnforceIf(early.Not()) # 作用: early=0 → p2 晚于 p 开工(不累计)
                            # 变量定义: cv (IntVar): p2 对已加工累计的贡献
                            cv = m.NewIntVar(0, lv * H, f"sac_{tag}")
                            # ── 【掺炼已加工门控】p2 在 p 前完工且该 combo 选中时贡献=负荷×工时 ──
                            m.Add(cv == lv * p2["dur"]).OnlyEnforceIf([early, pk])  # 作用: early∧pk → 贡献=C负荷×p2工时
                            m.Add(cv == 0).OnlyEnforceIf([early.Not(), pk])         # 作用: p2 未在p前完工 → 贡献0
                            m.Add(cv == 0).OnlyEnforceIf(pk.Not())                  # 作用: 该 combo 未选 → 贡献0
                            accum_terms.append(cv)
                    else:
                        # 单炼piece：dur × load_hr
                        load_hr_c = p2["u"]["comps"][0].get("load_hr") or P["RATE_HR"].get(c, p2["u"]["rate_hr"])
                        tag = f"sa_{p['u']['uid']}_{p['si']}_{p2['u']['uid']}_{p2['si']}_{c}"
                        # 变量定义: early (BoolVar): p2 是否在 p 开工前完工
                        early = m.NewBoolVar(f"sae_{tag}")
                        # ── 【前序完工判定】同上 ──
                        m.Add(p2["en"] <= p["st"]).OnlyEnforceIf(early)     # 作用: early=1 → p2 在 p 开工前完工
                        m.Add(p2["en"] > p["st"]).OnlyEnforceIf(early.Not()) # 作用: early=0 → p2 未在 p 前完工
                        # 变量定义: cv (IntVar): p2 对已加工累计的贡献
                        cv = m.NewIntVar(0, H * int(load_hr_c) + 1, f"sac_{tag}")
                        # ── 【单炼已加工门控】p2 在 p 前完工则贡献=工时×负荷，否则0 ──
                        m.Add(cv == p2["dur"] * int(load_hr_c)).OnlyEnforceIf(early)  # 作用: early=1 → 贡献=p2加工吨数
                        m.Add(cv == 0).OnlyEnforceIf(early.Not())                     # 作用: early=0 → 贡献0
                        accum_terms.append(cv)
                # --- (2) arrival_cum(c, st_p): 分段常数线性化 ---
                # arrival_cum = Σ (cum_k - cum_{k-1}) × 1[st_p ≥ T_k]
                # ── 【批次可用量·到港累计线性化】把分段常数到港累计量按 st_p 落点线性化 ──
                arrival_contrib = 0
                prev_cum = 0
                for T_k, cum_k in segs_c:
                    delta = cum_k - prev_cum   # delta: 第 k 段到港增量(常数)
                    prev_cum = cum_k
                    if delta <= 0:
                        continue
                    tag = f"sar_{p['u']['uid']}_{p['si']}_{c}_{T_k}"
                    # 变量定义: early (BoolVar): p 开工是否晚于到港时刻 T_k(st≥T_k)
                    early = m.NewBoolVar(f"sare_{tag}")
                    # ── 【到港时刻判定】early 真则 p 开工不早于 T_k ──
                    m.Add(p["st"] >= T_k).OnlyEnforceIf(early)      # 作用: early=1 → st≥T_k(开工时该到港已发生)
                    m.Add(p["st"] < T_k).OnlyEnforceIf(early.Not()) # 作用: early=0 → st<T_k(该到港尚未发生)
                    # 变量定义: dv (IntVar[0,delta]): 该段到港增量是否计入 p 开工时可用量
                    dv = m.NewIntVar(0, delta, f"sard_{tag}")
                    # ── 【到港增量门控】开工晚于 T_k 则计入该段到港增量，否则不计 ──
                    m.Add(dv == delta).OnlyEnforceIf(early)   # 作用: early=1 → dv=该段到港增量(已到港)
                    m.Add(dv == 0).OnlyEnforceIf(early.Not()) # 作用: early=0 → dv=0(未到港)
                    arrival_contrib += dv
                # --- (3) 本piece加工量 q_p ---
                # ── 【批次可用量·本批加工量】计算 p 本身的加工量(掺炼门控 / 单炼线性) ──
                if p["u"].get("is_blend"):
                    q_terms = []
                    for ki, (pk, loadmap) in enumerate(combo_vars.get(p["u"]["uid"], [])):
                        lv = int(loadmap.get(c, 0))
                        if lv <= 0:
                            continue
                        # 变量定义: qv (IntVar): p 在该 combo 下对 c 的加工量
                        qv = m.NewIntVar(0, lv * H, f"saq_{p['u']['uid']}_{p['si']}_{ki}_{c}")
                        # ── 【掺炼本批加工门控】选中该 combo 时 qv=负荷×工时，否则0 ──
                        m.Add(qv == lv * p["dur"]).OnlyEnforceIf(pk)   # 作用: pk=1 → 本批对 c 加工=负荷×工时
                        m.Add(qv == 0).OnlyEnforceIf(pk.Not())         # 作用: pk=0 → 本批不加工 c
                        q_terms.append(qv)
                    q_p = sum(q_terms) if q_terms else 0   # q_p: 本批对 c 的加工量
                else:
                    load_hr_c = p["u"]["comps"][0].get("load_hr") or P["RATE_HR"].get(c, p["u"]["rate_hr"])
                    q_p = p["dur"] * int(load_hr_c)        # q_p: 单炼本批加工量=工时×负荷(线性)
                # --- (4) 汇总约束: q_p + accum ≤ g_init_c + arrival_contrib ---
                # ── 【批次可用量硬约束】本批加工量 + 开工前已加工量 ≤ 月初库存 + 开工前到港量 ──
                lhs = q_p + (sum(accum_terms) if accum_terms else 0)   # lhs: 本批+已加工累计
                rhs = g_init_c + arrival_contrib                       # rhs: 月初库存+到港累计
                m.Add(lhs <= rhs)   # 作用: 开工时 G 罐可用量必须 ≥ 本批加工量(预防版 Benders 割)

    # ---- 时间分段可用量约束（已移除）----
    # 原由：每个到港释放时刻加 early/contrib 约束，但变量数爆炸导致 CP-SAT 超时。
    # 替代：腾容约束（下限，保证到港前加工够）+ Benders 割平面（上限，防止超排）
    # 共同覆盖了时间分段的语义，无需单独建模。

    # ---- Benders 割平面（由子问题 tank_fill 缺口检测自动生成）----
    for bi, (C, T_deadline, max_tons) in enumerate(benders_cuts):
        contribs = []
        for p in pieces:
            if crude_of_piece(p) != C or p["u"].get("is_blend"):
                continue
            lh = int(p["u"]["comps"][0].get("load_hr") or P["RATE_HR"].get(C, p["u"]["rate_hr"]))
            tag = f"bcut_{bi}_{C}_{p['u']['uid']}_{p['si']}_{p.get('sub_si',0)}"
            # 变量定义: early (BoolVar): 该片段是否在 deadline 前完工
            early = m.NewBoolVar(f"be_{tag}")
            # ── 【Benders割·单炼deadline前完工判定】early 真则 en≤deadline ──
            m.Add(p["en"] <= T_deadline).OnlyEnforceIf(early)      # 作用: early=1 → 片段在割平面截止前完工
            m.Add(p["en"] > T_deadline).OnlyEnforceIf(early.Not())  # 作用: early=0 → 片段晚于截止(不计入上限)
            # 变量定义: cv (IntVar): 该片段对割平面的贡献加工量
            cv = m.NewIntVar(0, H * lh + 1, f"bc_{tag}")
            # ── 【单炼贡献门控】deadline 前完工则贡献=工时×负荷，否则0 ──
            m.Add(cv == p["dur"] * lh).OnlyEnforceIf(early)   # 作用: early=1 → 贡献=该批加工吨数
            m.Add(cv == 0).OnlyEnforceIf(early.Not())         # 作用: early=0 → 贡献0
            contribs.append(cv)
        for u in units:
            if not u.get("is_blend") or not any(c["crude"] == C for c in u["comps"]):
                continue
            up = pieces_of_unit.get(id(u), [])
            if not up: continue
            td_terms = []
            for p in up:
                tag = f"bcut_b_{bi}_{C}_{u['uid']}_{p['si']}_{p.get('sub_si',0)}"
                # 变量定义: early (BoolVar): 该掺炼片段是否在 deadline 前完工
                early = m.NewBoolVar(f"bbe_{tag}")
                # ── 【Benders割·掺炼deadline前完工判定】同上 ──
                m.Add(p["en"] <= T_deadline).OnlyEnforceIf(early)      # 作用: early=1 → 片段在截止前完工
                m.Add(p["en"] > T_deadline).OnlyEnforceIf(early.Not())  # 作用: early=0 → 片段晚于截止
                # 变量定义: de (IntVar[0,H]): deadline 前完工时该片段工时
                de = m.NewIntVar(0, H, f"bbd_{tag}")
                # ── 【工时门控】deadline 前完工则 de=工时，否则0 ──
                m.Add(de == p["dur"]).OnlyEnforceIf(early)   # 作用: early=1 → de=该片段工时
                m.Add(de == 0).OnlyEnforceIf(early.Not())    # 作用: early=0 → de=0
                td_terms.append(de)
            td_e = sum(td_terms)   # td_e: 该掺炼单元 deadline 前累计工时
            for ki, (pk, lm) in enumerate(combo_vars.get(u["uid"], [])):
                lv = int(lm.get(C, 0))
                if lv <= 0: continue
                # 变量定义: cuk (IntVar): 该 combo 下 C 的割平面贡献
                cuk = m.NewIntVar(0, lv * H, f"bbc_{bi}_{C}_{u['uid']}_{ki}")
                # ── 【掺炼贡献门控】选中该 combo 时贡献=负荷×deadline前工时，否则0 ──
                m.Add(cuk == lv * td_e).OnlyEnforceIf(pk)   # 作用: pk=1 → 贡献=C负荷×deadline前累计工时
                m.Add(cuk == 0).OnlyEnforceIf(pk.Not())     # 作用: pk=0 → 贡献0
                contribs.append(cuk)
        if contribs:
            # ── 【Benders割硬约束】deadline 前累计加工 C 的量 ≤ 上限 max_tons(防止超排致缺口) ──
            m.Add(sum(contribs) <= int(max_tons))   # 作用: 上界割，禁止截止前加工超过可得量上限

    # ---- 卸油停产窗口（专用罐）：该油种主力片段不得与卸油窗口重叠 ----
    # 注：仅挡主力油种 piece（crude_of_piece）。掺炼组分不在 CP-SAT 挡，
    # 由 tank_fill 侧多罐接力（pick_G 跨类型兜底）保证组分供应。
    for wi, (C, s, e) in enumerate(no_process_windows):
        e = min(e, H)
        if e <= s:
            continue
        grp = [p["iv"] for p in pieces if crude_of_piece(p) == C]
        if not grp:
            continue
        # 变量定义: block (FixedSizeInterval): 油种 C 第 wi 个卸油停产窗口[s,e]的固定占用区间
        block = m.NewFixedSizeIntervalVar(s, e - s, f"noproc_{C}_{wi}")
        # ── 【卸油停产窗口】油种 C 的主力片段不得与卸油窗口重叠(卸油期间该罐不能供料加工) ──
        m.AddNoOverlap([block] + grp)   # 作用: 卸油窗口区间与该油种片段区间互斥

    # 禁连两批：同油种(主力油)片段不得首尾相接
    cru_pieces = defaultdict(list)
    for p in pieces: cru_pieces[p["u"]["comps"][0]["crude"]].append(p)
    for c in P["NO_CONSEC"]:
        ps = cru_pieces.get(c, [])
        for i in range(len(ps)):
            for j in range(len(ps)):
                if i == j: continue
                # 变量定义: adj (BoolVar): ps[i] 是否紧接在 ps[j] 之前(en_i==st_j)
                adj = m.NewBoolVar(f"adj_{c}_{i}_{j}")
                # ── 【禁连·邻接判定】adj 真则 ps[i] 完工=ps[j] 开工(首尾相接) ──
                m.Add(ps[i]["en"] == ps[j]["st"]).OnlyEnforceIf(adj)      # 作用: adj=1 → 两批相接
                m.Add(ps[i]["en"] != ps[j]["st"]).OnlyEnforceIf(adj.Not()) # 作用: adj=0 → 两批不相接
                # ── 【禁连硬约束】禁止禁连油种两批相接：至少一个不排产，或不相接 ──
                m.AddBoolOr([ps[i]["pres"].Not(), ps[j]["pres"].Not(), adj.Not()])  # 作用: 两批都排产时强制不相接(链长≤1)

    # 非禁连油种连续供料段限容：首尾相接(en==st)的同主料 piece 链，主料总吨数 ≤ 一罐可用容量。
    # 补 Part A(限单批)缺口——多批相接成超一罐容的连续段，tank_fill 续喂不中途补料 → 末日亏油。
    # 禁连油种已由上方反邻接(链长≤1)覆盖，此处跳过；上限取 charge_cap(一个最大罐，恒与罐占用无关)。
    #
    # 【与 warm_start 耦合】此约束的 O(n²) 时间邻接反射(en!=st，弱传播)显著加大求解难度：实测
    # 无 warm_start 时真实模板 180s 仍返回 UNKNOWN；有贪心 warm-start 引导则 ~120s 收敛可行解。
    # 故仅在 warm_start=True 时启用——生产调度(multi_round_schedule 默认 WARM_START=True)全程启用；
    # 轻量单测(warm_start=False，只求快速拿可行解验证其它功能)跳过，避免退化成 UNKNOWN。
    if stretch_on:
        for c, ps in cru_pieces.items():
            if c in P["NO_CONSEC"] or len(ps) < 2:
                continue
            cap_i = int(charge_cap(P, c))
            # 全月可得量 ≤ 一罐容 → 任何连续段（子集）都不会超容，无需约束（省模型、加速求解）
            if cap_i <= 0 or avail.get(c, 0.0) <= cap_i:
                continue
            run = {}
            for p in ps:                         # run = 该 piece 所在连续段累计主料吨数，封顶一罐
                # 变量定义: rv (IntVar[0,cap_i]): p 所在连续段累计主料吨数(上界=一罐容)
                rv = m.NewIntVar(0, cap_i, f"run_{c}_{p['u']['uid']}_{p['si']}_{p['sub_si']}")
                # ── 【段起点下界】连续段累计 ≥ 本 piece 吨数 ──
                m.Add(rv >= p["mtons"])          # 段起点：至少含本 piece 吨数  # 作用: 段累计至少等于该 piece 主料吨数
                run[id(p)] = rv
            for i in range(len(ps)):
                for j in range(len(ps)):
                    if i == j:
                        continue
                    # 变量定义: radj (BoolVar): ps[i] 是否紧接在 ps[j] 之前(相接)
                    radj = m.NewBoolVar(f"radj_{c}_{i}_{j}")
                    # ── 【相接判定】radj 真则 ps[i] 完工=ps[j] 开工 ──
                    m.Add(ps[i]["en"] == ps[j]["st"]).OnlyEnforceIf(radj)      # 作用: radj=1 → 两批相接成连续段
                    m.Add(ps[i]["en"] != ps[j]["st"]).OnlyEnforceIf(radj.Not()) # 作用: radj=0 → 两批不相接
                    # 相接则累加前驱段吨数；run 上界=cap_i，超一罐容的链无解 → 该相接被禁
                    # ── 【连续段累加】两批都排产且相接时，后继段累计 ≥ 前驱段累计 + 后继批吨数 ──
                    m.Add(run[id(ps[j])] >= ps[j]["mtons"] + run[id(ps[i])]).OnlyEnforceIf(
                        [radj, ps[i]["pres"], ps[j]["pres"]])  # 作用: 相接则累计吨数叠加，超一罐容(cap_i)则该相接被禁

    # 目标优先级：达标 ≫ 批次优先（填满全月已由硬约束 sum(dur)==H 保证，到港早已删除）
    pri_on = P.get("W2_PRIORITY", 20) > 0
    SCALE = P.get("W2_WEIGHT_SCALE", 10)
    nbatch = sum(p["pres"] for p in pieces)   # nbatch: 实际排产批次总数(线性项)
    # 批次硬上限：present 片段总数 ≤ batch_cap（从 estimate_batches 预估起、逐步放宽到+2）。
    # 强制 CDU 用少而长的批次、消除碎片；太紧则本轮 INFEASIBLE（由上层递增 cap 兜底）。
    if batch_cap is not None:
        # ── 【批次数硬上限】排产批次总数 ≤ batch_cap，强制少而长的批次消除碎片 ──
        m.Add(nbatch <= int(batch_cap))   # 作用: present 片段数不超过上限，避免碎片化排产

    # 批次优先项（第二优先级）：按同油种 piece 创建序号取递减权重，present门控开工时刻
    weights_seq = P.get("_BATCH_WEIGHTS")
    if weights_seq is None:
        weights_seq, _ = greedy_batch_priority(P)
    # 开工时刻按【天】计（con_day = 开工小时 // 24 ∈ [0, DAYS]）——压小批次优先项量级，
    # 避免字典序权重过大拖慢求解（批次优先本是粗粒度先后，天级足够）。
    DMAX = H // 24
    pri_terms = []; pri_wsum = 0
    if pri_on:
        crude_piece_idx = defaultdict(int)
        for p in pieces:
            c = p["u"]["comps"][0]["crude"]
            seq = weights_seq.get(c) or [0.0]
            j = crude_piece_idx[c]; crude_piece_idx[c] += 1
            w = seq[min(j, len(seq) - 1)]
            wi = int(round(w * SCALE))   # wi: 该 piece 的批次优先权重(按创建序递减)
            if wi <= 0:
                continue
            # 变量定义: con (IntVar[0,H]): 该 piece 的开工时刻(present 门控)
            con = m.NewIntVar(0, H, f"pcon_{p['u']['uid']}_{p['si']}_{p['sub_si']}")
            # ── 【开工时刻门控】排产时 con=开工时刻，未排产时 con=0 ──
            m.Add(con == p["st"]).OnlyEnforceIf(p["pres"])      # 作用: pres=1 → con 取真实开工时刻
            m.Add(con == 0).OnlyEnforceIf(p["pres"].Not())      # 作用: pres=0 → con=0(不参与优先项)
            # 变量定义: con_day (IntVar[0,DMAX]): 开工所在天 = con // 24
            con_day = m.NewIntVar(0, DMAX, f"pcond_{p['u']['uid']}_{p['si']}_{p['sub_si']}")
            # ── 【开工时刻按天取整】con_day = con ÷ 24(整除)，压小优先项量级 ──
            m.AddDivisionEquality(con_day, con, 24)   # 作用: 开工时刻转天级，供批次优先项加权
            pri_terms.append(wi * con_day); pri_wsum += wi

    # 目标：达标 ≫ 批次优先（到港早已删除，填满由后处理保证）
    W_SHORT = 10**7
    W_PRI = 1000
    # ── 【目标函数】(非约束)最小化：大权重×达标缺口 + 小权重×批次优先项，达标远优先于批次顺序 ──
    m.Minimize(W_SHORT * sum(shorts) + W_PRI * sum(pri_terms))   # 作用: 先极小化缺口(达标)，再优化批次先后

    # 贪心 warm-start：按 greedy_batch_priority 的批次顺序推算每个 piece 的 pres/dur/st 初值，
    # 用 AddHint 注入作为搜索起点（软提示，不可行自动回退）。HINT_SEARCH 另受 hint_search 控制。
    if warm_start:
        order = P.get("_BATCH_ORDER")
        if order is None:
            _, order = greedy_batch_priority(P)
        for p, pres_v, dur_v, st_v in warm_start_hints(pieces, order, MIN_H, H):
            # ── 【warm-start 软提示】(非约束)为 pres/dur/st 注入贪心初值作为搜索起点 ──
            m.AddHint(p["pres"], pres_v)   # 作用: 提示 pres 初值，加速收敛(不可行自动回退)
            m.AddHint(p["dur"], dur_v)     # 作用: 提示 dur 初值
            m.AddHint(p["st"], st_v)       # 作用: 提示 st 初值

    sol = cp_model.CpSolver()
    sol.parameters.max_time_in_seconds = time_limit
    sol.parameters.num_search_workers = P.get("NUM_WORKERS", 8)   # 并行搜索，加速劣质FEASIBLE→更优解收敛
    if random_seed is not None:
        sol.parameters.random_seed = random_seed
    # HINT_SEARCH：让分支优先沿 hint 下潜（更快命中贪心解，但削弱多 worker/多轮多样性），
    # 故做成可控——默认关，仅由上层在指定轮次(如第1轮)开启。
    if warm_start and hint_search:
        sol.parameters.search_branching = sol.parameters.HINT_SEARCH
    # 达标即停：命中首个「所有软达标油种缺口=0」的可行解即停搜（2026-07-15，默认关）
    if stop_on_compliant:
        stt = sol.Solve(m, _CompliantStop(shorts))
    else:
        stt = sol.Solve(m)
    if stt not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError(f"CP-SAT 无可行解[{sol.StatusName(stt)}]：检查加工量、到港量、负荷候选是否矛盾")

    # 解后:写回掺炼单元选中档的各组分负荷 → comps load_hr (供 tank_fill)
    for u in units:
        if not u["is_blend"]: continue
        for pk, loadmap in combo_vars[u["uid"]]:
            if sol.Value(pk):
                for comp in u["comps"]: comp["load_hr"] = loadmap[comp["crude"]]
                u["chosen_loads"] = dict(loadmap); break

    seq = []
    for p in pieces:
        if sol.Value(p["pres"]):
            seq.append({"u": p["u"], "start_h": sol.Value(p["st"]),
                        "dur_h": sol.Value(p["dur"]), "si": p["si"]})
    seq.sort(key=lambda x: x["start_h"])
    return seq, sol.Value(nbatch), sol.ObjectiveValue(), stt

# ============================================================
# 4. 罐物流仿真（CDU计划已定）
#    - 到港→T罐；T罐→G罐一次性整批传输（在批次开始前完成）
#    - 预排产sheet到库量行：按靠泊日期填入T罐
# ============================================================
@dataclass
class Tk:
    crude: object = None
    ton: float = 0.0

def order_transfer_sources(c, gt, G, T, crude_of, avail_of, ggt, tgt):
    """向 G 罐 gt 输油种 c 时的**源罐顺序**（2026-07-05 引入 G→G）：
    ① 其它同种 G 罐（`ggt`=G→G管速）→ ② 同种 T 罐（`tgt`=T→G管速）；各按可用油升序（尽快腾空小罐）。
    crude_of(t)→该罐当前油种；avail_of(t)→可用油(ton−底油)。返回 [(src_tank, tph), ...]。
    纯函数，便于单测；G→G 与 T→G 共用同一根管道，仅管速不同。"""
    g_srcs = sorted((g for g in G if g != gt and crude_of(g) == c and avail_of(g) > 1e-6),
                    key=avail_of)
    t_srcs = sorted((t for t in T if crude_of(t) == c and avail_of(t) > 1e-6),
                    key=avail_of)
    return [(g, ggt) for g in g_srcs] + [(t, tgt) for t in t_srcs]


def tank_fill(P, seq, ship_assignments=None):
    G, T = P["GTANKS"], P["TTANKS"]
    CAP = P["CAP"]; ALLOW = P["ALLOW"]; FAR = P["FARNESS"]
    heel = P.get("HEEL", {})  # Backward compat: default to empty dict if not in P
    SW = P.get("SWITCH_RESID_TON", SWITCH_RESID_TON)  # G罐换油残余阈值（吨）：可用残余≤SW即视为可换油
    DAYS = P["DAYS"]
    tanks = {t: Tk(P["INIT"].get(t, {}).get("crude"),
                   P["INIT"].get(t, {}).get("ton", 0.0)) for t in G + T}
    grid = {t: {d: {} for d in range(1, DAYS + 1)} for t in G + T}
    warn = []; clog = []
    unload_to_t_count = 0  # 最终卸油落入T罐的船次（直落+改投），后续需T→G倒运 → 评分维度
    unload_commingle_count = 0  # 卸油混油(commingling)次数：卸前罐内异种旧油残余>SW → 评分维度

    # ---- 工具函数 ----
    def eT(t, c):
        a = str(ALLOW.get(t, "任意"))
        if a == "进口": return P["ORIGIN"].get(c) == "imp"
        if a == "国内": return P["ORIGIN"].get(c) == "dom"
        return True

    def gE(t, c):
        a = str(ALLOW.get(t, "任意")).strip()
        if a == "进口": return P["ORIGIN"].get(c) == "imp"
        if a == "国内": return P["ORIGIN"].get(c) == "dom"
        return a in ("任意", "") or norm(c) == norm(a)

    def pick_G(c, day=None, exclude=None):
        """为油种c选进料G罐：同类型罐优先（主力油→主力罐，掺炼专用油→掺炼罐）。
        搜索顺序：同类型有同种油 > 同类型空罐 > 跨类型有同种油（兜底）。
        day 非空时，排除当日处于卸油停产窗口的罐（不得从收船罐进料）。
        exclude：排除的罐集合（多罐接力时排除已用/已选的接力罐）。"""
        can_single = P.get("CAN_SINGLE", {})
        is_blend_only = not can_single.get(c, True)
        excl = set(exclude or ())
        def feedable(t):
            # 边界天(部分不供料)仍可供料，仅整天不供料(可供小时≤0)才排除
            return t not in excl and (day is None or feed_free_h(t, day) > 1e-6)
        type_G  = [t for t in G if (t in P["BLEND_TANKS"]) == is_blend_only and feedable(t)]
        other_G = [t for t in G if (t in P["BLEND_TANKS"]) != is_blend_only and feedable(t)]
        # 1. 同类型罐已有同种油（多油优先）
        same = [t for t in type_G if tanks[t].crude == c and gE(t, c) and tanks[t].ton > 1e-6]
        same.sort(key=lambda t: (0 if norm(ALLOW.get(t, "")) == norm(c) else 1, -tanks[t].ton))
        if same: return same[0]
        # 2. 同类型空罐（专用优先）：crude 未占用且可用残余(ton−底油)≤换油阈值即视为可接新油，
        #    停在底油位的罐(ton==heel, crude=None)同样可接收新油。
        empty = [t for t in type_G
                 if tanks[t].crude is None
                 and tanks[t].ton - heel.get(t, 0.0) <= SW and gE(t, c)]
        empty.sort(key=lambda t: 0 if norm(ALLOW.get(t, "")) == norm(c) else 1)
        if empty: return empty[0]
        # 2.5. 同类型旧油已抽至换油阈值内（可用残余≤阈值，允许接收新油）
        low = [t for t in type_G
               if tanks[t].crude is not None and tanks[t].crude != c
               and (tanks[t].ton - heel.get(t, 0.0)) <= SW
               and gE(t, c) and CAP[t] - tanks[t].ton >= 1]
        low.sort(key=lambda t: tanks[t].ton)   # 残余最少的优先
        if low: return low[0]
        # 3. 跨类型已有同种油（兜底，避免 cargo 丢失）
        same_other = [t for t in other_G if tanks[t].crude == c and gE(t, c) and tanks[t].ton > 1e-6]
        same_other.sort(key=lambda t: -tanks[t].ton)
        if same_other: return same_other[0]
        return None

    def transfer_to_G(c, need, day, gt, allow_g_src=True):
        """向 G 罐 gt 一次性整批输入 need 吨油种 c。
        allow_g_src=False 时只从 T 罐搬油（预传输用，保留 G 罐作 no_feed 备用）。
        源优先级：① 其它同种 G 罐(GG_TPH) → ② 同种 T 罐(TRANSFER_TPH)。
        G→G 与 T→G 共用同一根管道，每条 raw_move 记各自 tph。"""
        space = CAP[gt] - tanks[gt].ton; mv = 0
        GGT = P.get("GG_TPH", GG_TPH_DEF); TGT = P.get("TRANSFER_TPH", 700)
        # 源顺序：① 其它同种 G 罐(GGT) → ② 同种 T 罐(TGT)；各按可用油升序（纯函数，见 order_transfer_sources）
        srcs = order_transfer_sources(c, gt, G, T,
                                      lambda t: tanks[t].crude,
                                      lambda t: tanks[t].ton - heel.get(t, 0.0), GGT, TGT)
        if not allow_g_src:
            srcs = [(t, tph) for t, tph in srcs if t in T]
        for s, tph in srcs:
            if mv >= need - 1e-6 or space <= 1e-6: break
            avail_s = tanks[s].ton - heel.get(s, 0.0)
            q = min(avail_s, need - mv, space)
            if q <= 0: continue
            tanks[s].ton -= q
            tanks[gt].crude = c; tanks[gt].ton += q; space -= q; mv += q
            # oil_avail_h=0.0：仿真只搬运源罐里当时已存在的油，源油可用性恒满足；
            # release_floor 绑定目标 G 罐上一段供料结束时刻（见 build_sessions）。
            raw_moves.append({"src_t": s, "dst_g": gt, "crude": c, "tons": q,
                              "trigger_h": (day - 1) * 24, "oil_avail_h": 0.0, "tph": tph})
        return mv

    # ---- 预计算每日加工量（按组分） ----
    # {day: [(crude, q_ton, load_hr)]}
    day_draw = defaultdict(list)
    day_span = {}   # (day, crude) -> [min_start_h, max_end_h]
    day_segs = defaultdict(list)   # day -> [{crude, tons, load, seg_start, seg_end}]（逐段真实配方，供逐段供料）
    for b in seq:
        u = b["u"]; t = b["start_h"]
        remain = {c["crude"]: b["dur_h"] * c["load_hr"] for c in u["comps"]}
        cr_hr = {c["crude"]: c["load_hr"] for c in u["comps"]}
        while any(v > 1e-6 for v in remain.values()) and t < DAYS * 24:
            day = int(t // 24) + 1; hleft = 24 - (t % 24)
            for c in u["comps"]:
                k = c["crude"]
                if remain[k] <= 1e-6: continue
                q = min(remain[k], cr_hr[k] * hleft)
                day_draw[day].append((k, q, c["load_hr"]))
                seg_start = t
                seg_end = t + (q / cr_hr[k] if cr_hr[k] else 0)
                day_segs[day].append({"crude": k, "tons": q, "load": c["load_hr"],
                                      "seg_start": seg_start, "seg_end": seg_end})
                span = day_span.get((day, k))
                if span is None:
                    day_span[(day, k)] = [seg_start, seg_end]
                else:
                    span[0] = min(span[0], seg_start)
                    span[1] = max(span[1], seg_end)
                remain[k] -= q
            t += hleft

    raw_moves = []          # 原始输油动作（每源 T 罐一条）：T→G 排期输入
    feed_pieces = []        # 每 (G罐, 油种, 日) 的供料片段，后续合并成供料段
    fp_by_key = {}          # (g_tank, crude, day) -> feed_piece：同罐同油同日多段合并为一条

    # ---- 逐日仿真 ----
    charge = {}  # crude -> 当前进料G罐
    charge_chain = {}  # crude -> [有序接力供料罐]（一段连续供料 > 一罐容时多罐接力，2026-07-05）
    future_recv = defaultdict(list)  # day -> [(tank, crude, tons, regrade_now)] 按实际卸油日的到库量
    pending_ships = []       # [(ship_idx, a, days_waited)] 等候容量的到港船
    total_demurrage_days = 0.0  # 月累计滞期天数（本仿真额外追加部分）
    # 可追加滞期 = 月度限制 − 靠泊计划已含滞期（避免重复计算）
    demurrage_remaining = max(0.0, P.get("DEMURRAGE_BUDGET", 0.0)
                              - P.get("DEMURRAGE_INCLUDED", 0.0))
    no_feed_hours = defaultdict(dict)  # tank -> {day: 不供料小时数}（逐日精确，边界天取部分）
    def feed_free_h(t, day):
        """该罐当天可供料小时数 = 24 − 不供料小时（整天不供料→0）。"""
        return 24.0 - no_feed_hours.get(t, {}).get(day, 0.0)

    # 预计算：批次开始「前一天」触发 T→G 整批传输
    # 月初(第1天)批次也在第1天预传输：主力油初始多在G大罐(shortfall≤0自动跳过)，
    # 但掺炼专用油初始常在T罐，必须先 T→G 才能供料，否则头批 starve。
    pre_transfer = defaultdict(list)
    for b in seq:
        batch_start_day = int(b["start_h"] // 24) + 1
        transfer_day = max(1, batch_start_day - 1)
        for c in b["u"]["comps"]:
            crude = c["crude"]
            need = b["dur_h"] * c["load_hr"]
            pre_transfer[transfer_day].append((crude, need, b))

    for day in range(1, DAYS + 1):
        day_start_phys = {t: tanks[t].ton for t in G + T}
        todays_unloads = []   # 当日接卸的船 [(crude, splits, days_waited, pre_tank)]，加工后再算窗口/到库

        # 新到港船加入等候队列
        for ship_idx, a in enumerate(P["ARRIVALS"]):
            if a["berth_day"] == day:
                pending_ships.append([ship_idx, a, 0])

        # 逐船尝试接卸：一船可分卸多罐(plan_unload_split)。放不下全部→滞期等待；
        # 超预算→强卸(放能放的，余量货物丢失)。落罐/重命名由 split + should_regrade 决定。
        next_pending = []
        for entry in pending_ships:
            ship_idx, a, days_waited = entry
            C = a["crude"]; Q = a["ton"]
            pre_tank = (ship_assignments.get(ship_idx, {}).get("tank")
                        if ship_assignments else None)
            # 靠泊当日"先供后卸"：收油 G 罐当天会先供 CDU、腾出空间，分卸空间应计入这部分
            # （否则用 0 点库存低估空间、把本可进 G 罐的油多分给 T 罐）。
            STOP_HOD = P.get("STOP_FEED_HOD", STOP_FEED_HOD_DEF)
            def _berth_free(t):
                if t not in G:
                    return 0.0
                xc = tanks[t].crude
                if xc is None or charge.get(xc) != t:   # 仅当前正供该罐油种的 G 罐计入，避免多罐重复
                    return 0.0
                planned = sum(q for c2, q, lh in day_draw.get(day, []) if c2 == xc)
                if planned <= 1e-6:
                    return 0.0
                load = next((lh for c2, q, lh in day_draw.get(day, []) if c2 == xc and lh), 0) \
                    or P["RATE_HR"].get(xc, DEF_RATE_HR)
                return min(planned, STOP_HOD * load, max(0.0, tanks[t].ton - heel.get(t, 0.0)))
            # 当前罐态快照（供分卸决策；allow_ok: G查gE、T查eT；berth_feed: 靠泊当日先供腾出空间）
            state = {t: {"crude": tanks[t].crude, "ton": tanks[t].ton, "cap": CAP[t],
                         "heel": heel.get(t, 0.0), "is_g": t in G,
                         "allow_ok": (gE(t, C) if t in G else eT(t, C)),
                         "far": FAR.get(t, 9), "berth_feed": _berth_free(t)}
                     for t in G + T}
            splits, remaining = plan_unload_split(
                C, Q, state, pre_tank, P.get("ALLOW_MIX", {}).get(C, set()),
                P.get("MAX_UNLOAD_TANKS", 2))

            if remaining > 1e-6:
                # 放不下全部：滞期预算内继续等待
                if total_demurrage_days + 1 <= demurrage_remaining:
                    if days_waited == 0:
                        clog.append(f"第{day}日 {C}船靠泊，罐容量不足，进入等待")
                    total_demurrage_days += 1
                    next_pending.append([ship_idx, a, days_waited + 1])
                    clog.append(f"第{day}日 {C}船滞期第{days_waited+1}天，"
                                f"月累计{total_demurrage_days:.0f}天")
                    continue
                # 超预算强卸：放能放的 splits，剩余货物丢失
                warn.append(f"第{day}日 {C}船等待{days_waited}天已超可追加滞期"
                            f"(限制{P.get('DEMURRAGE_BUDGET',0):.0f}−已含"
                            f"{P.get('DEMURRAGE_INCLUDED',0):.0f}=可追加"
                            f"{demurrage_remaining:.0f}天)，强制接卸")
                warn.append(f"第{day}日 {C} 强制接卸仍有{remaining:.0f}吨无合规罐，货物丢失")
                if not splits:
                    continue
                # 有 splits → 落到下方 commit（强卸把能放的放掉）

            # commit：登记分卸；收油罐靠泊当日**最多供料到 STOP_FEED_HOD**（provisional 上限，
            # 供加工块封顶），真正的 stop_feed_h/时间窗口/到库量在【加工后】按实际供料工时算（先供后卸）。
            prov_nofeed = 24.0 - P.get("STOP_FEED_HOD", STOP_FEED_HOD_DEF)
            for t, q in splits:
                if t in T:
                    unload_to_t_count += 1
                no_feed_hours[t][day] = max(no_feed_hours[t].get(day, 0.0), prov_nofeed)
            todays_unloads.append((C, splits, days_waited, pre_tank))
            # 预分配罐因旧油未清(异种油漂移)被 Tier0 跳过 → 告警，区分落T(降级)/落其它G(换罐)
            if pre_tank is not None and all(t != pre_tank for t, _ in splits):
                ps = state.get(pre_tank)
                if ps and ps["crude"] is not None and ps["crude"] != C and splits:
                    avail_old = ps["ton"] - ps["heel"]
                    t_dest = next((t for t, _ in splits if t in T), None)
                    if t_dest:
                        warn.append(f"第{day}日 T罐降级：{pre_tank}有旧油{ps['crude']}"
                                    f"(可用旧油{avail_old:.0f}t未抽尽)，{C}→{t_dest}")
                    else:
                        warn.append(f"第{day}日 换罐接卸：{pre_tank}有旧油{ps['crude']}"
                                    f"(可用旧油{avail_old:.0f}t未抽尽)，{C}→{splits[0][0]}")
        pending_ships = next_pending

        # 当日到库（按实际卸油日分布，标在真正卸油那天）；重命名(regrade)标志随首个卸油日
        for t, c, tons, regrade_now in future_recv.pop(day, []):
            if regrade_now:
                tanks[t].crude = c            # 卸量>卸前可用 → 罐存油种重命名为卸油油种
            elif tanks[t].crude is None:
                tanks[t].crude = c            # 空罐(无油种标记)挂名
            # 否则保持原 crude（混油、按多数命名）
            tanks[t].ton += tons
            grid[t][day]["recv"] = grid[t][day].get("recv", 0) + tons
            grid[t][day]["unload_recv"] = grid[t][day].get("unload_recv", 0) + tons
        # 当日卸完后查容量（G/T都查）—— 此块保持原样，紧随其后
        for t in P["GTANKS"] + P["TTANKS"]:
            if tanks[t].ton > CAP[t] + 1:
                warn.append(f"第{day}日 {t} 卸船后超罐容{CAP[t]:.0f}（当前{tanks[t].ton:.0f}）")

        # T→G 整批传输（批次开始前一天触发，一次性输油）
        for (crude, need, b) in pre_transfer.get(day, []):
            gt = charge.get(crude)
            if gt is None or not gE(gt, crude) or feed_free_h(gt, day) <= 1e-6:
                gt = pick_G(crude, day)
            if gt is None:
                warn.append(f"第{day}日 {crude} 无可用进料G罐（整批传输）"); continue
            charge[crude] = gt
            # 换油保护：若目标 G 罐旧油可用残余>换油阈值，跳过本次传输
            if (tanks[gt].crude is not None
                    and tanks[gt].crude != crude
                    and (tanks[gt].ton - heel.get(gt, 0.0)) > SW):
                warn.append(
                    f"第{day}日 {gt} 换油等待：旧油{tanks[gt].crude}"
                    f" 可用旧油{tanks[gt].ton-heel.get(gt,0.0):.0f}吨未抽尽，"
                    f"新油{crude}暂缓传入"
                )
                continue
            # ── 连续供料期一次装够（所有供料罐通用：主力罐+掺炼罐）──
            # 供料罐供料期不受油，一段连续供料只有开炼前这一次装油机会，故须一次预装
            # 「整段总抽量 + 换油残余」，封顶到罐可用容量；续喂批次处于供料期内，跳过。
            batch_start_day = int(b["start_h"] // 24) + 1
            # ① 续喂判定：前一天已在喂同种油 → 处于连续供料期内，禁止供料期补油，跳过
            if any(_c == crude for _c, _q, _lh in day_draw.get(batch_start_day - 1, [])):
                continue
            # ② 整段连续供料期总抽量：自开炼日逐日累加，遇「该油种当天不加工」即到补油窗口止
            stretch_draw = 0.0; dd = batch_start_day
            while dd <= DAYS and any(_c == crude for _c, _q, _lh in day_draw.get(dd, [])):
                stretch_draw += sum(_q for _c, _q, _lh in day_draw.get(dd, []) if _c == crude)
                dd += 1
            # ③ 整段总需求 = 整段抽量 + 换油残余；用【多罐接力】覆盖（B-完整，2026-07-05）：
            #    首罐 gt 装满，溢出量再选接力罐(同种/空 G 罐)逐个预装，登记有序接力链 charge_chain[crude]。
            need_total = stretch_draw + SW
            chain = charge_chain.setdefault(crude, [])
            if gt not in chain:
                chain.insert(0, gt)   # 首罐入链首位
            def _same_avail(t):
                return max(0.0, tanks[t].ton - heel.get(t, 0.0)) if tanks[t].crude == crude else 0.0
            covered = sum(_same_avail(t) for t in chain)
            guard = 0
            while covered < need_total - 1e-6 and guard < len(G) + 2:
                guard += 1
                # 链上还没装满的罐优先；否则新选一个接力罐
                rgt = next((t for t in chain
                            if (CAP[t] - heel.get(t, 0.0)) - _same_avail(t) > 1e-6), None)
                if rgt is None:
                    rgt = pick_G(crude, day, exclude=chain)
                    if rgt is None:
                        warn.append(f"第{day}日 {crude} 接力罐不足，整段供料缺约{need_total-covered:.0f}吨")
                        break
                    chain.append(rgt)
                r_usable = CAP[rgt] - heel.get(rgt, 0.0)
                r_cur = _same_avail(rgt)
                sf = max(0.0, min(r_usable, r_cur + (need_total - covered)) - r_cur)
                if sf <= 1e-6:
                    break
                mv = transfer_to_G(crude, sf, day, rgt, allow_g_src=False)
                covered += mv
                if mv < sf - 1:
                    warn.append(f"第{day}日 {crude} 整段预输缺货{sf-mv:.0f}吨（{rgt}）")
                    break

            # 额外从 T 罐填充一个 BLEND_TANK 作 no_feed 备份（不在 chain 中，
            # 仅当主力罐被 no_feed 封住时由 pick_G 跨类型兜底发现）
            backup_g = None
            for bg in G:
                if bg in chain: continue
                if bg in P.get("MAIN_TANKS", set()): continue  # 只要 BLEND_TANK
                if not gE(bg, crude): continue
                if tanks[bg].crude is not None and tanks[bg].crude != crude: continue
                if tanks[bg].crude is not None and (tanks[bg].ton - heel.get(bg, 0.0)) > SW: continue
                backup_g = bg; break
            if backup_g:
                space_bg = CAP[backup_g] - tanks[backup_g].ton
                if space_bg > 1e-6:
                    transfer_to_G(crude, space_bg, day, backup_g, allow_g_src=False)

        # 当日加工（从G罐抽油）——按批次段 day_segs 逐段供料，每段用【真实配方负荷】。
        # 单机CDU串行：段按开工时刻 seg_start 时序供料；同种油当天多段时，charge/charge_chain
        # 在段间自然延续（后段接前段罐续抽，罐见底才切）。不再用「日等效负荷」抹平真实负荷/缺口。
        for seg in sorted(day_segs.get(day, []), key=lambda s: s["seg_start"]):
            c = seg["crude"]; seg_load = seg["load"]
            # ── 多罐接力供料（B-完整，2026-07-05）：一段连续供料 > 一罐容时，见底切下一预装罐 ──
            # 硬性不重叠规则：供料期不受油、单罐不中途补料；但**一段连续供料可由多个预装罐顺序接力**
            # （CDU 不停、阀门切换）。只有确无任何同种油可供时才亏油。
            remaining_q = seg["tons"]; used = []; win_short = False; no_tank = False
            chain0 = charge_chain.get(c, [])
            def _feedable_now(t):
                return (gE(t, c) and tanks[t].crude in (None, c)
                        and feed_free_h(t, day) > 1e-6
                        and tanks[t].ton - heel.get(t, 0.0) > 1e-6 and t not in used)
            while remaining_q > 1e-6:
                gt = next((t for t in chain0 if _feedable_now(t)), None)       # 链上优先
                if gt is None and charge.get(c) and _feedable_now(charge[c]):
                    gt = charge[c]
                if gt is None:
                    gt = pick_G(c, day, exclude=used)                          # 兜底新选
                if gt is None:
                    # 所有 G 罐均无此油种 → 尝试从 T 罐实时补料到空 G 罐接力
                    t_src = next((t for t in T if tanks[t].crude == c
                                  and tanks[t].ton - heel.get(t, 0.0) > 1e-6), None)
                    if t_src:
                        # 找一个空的、兼容的 G 罐接收（不限 MAIN/BLEND）
                        for eg in G:
                            if eg in used: continue
                            if not gE(eg, c): continue
                            if feed_free_h(eg, day) <= 1e-6: continue
                            if tanks[eg].crude is not None and tanks[eg].crude != c: continue
                            if tanks[eg].crude is not None and (tanks[eg].ton - heel.get(eg, 0.0)) > SW:
                                continue
                            space_eg = CAP[eg] - tanks[eg].ton
                            avail_t = tanks[t_src].ton - heel.get(t_src, 0.0)
                            qmv = min(avail_t, remaining_q, space_eg)
                            if qmv > 1e-6:
                                tanks[t_src].ton -= qmv
                                tanks[eg].crude = c
                                tanks[eg].ton += qmv
                                gt = eg
                                break
                if gt is None:
                    no_tank = True; break
                avail_gt = max(0.0, tanks[gt].ton - heel.get(gt, 0.0))
                free_h = feed_free_h(gt, day)
                if seg_load > 0:
                    # 确保加工时间为整数：按整小时抽取，不足整小时的零头留罐内，
                    # 剩余需求由接力罐/T→G补足，保障整月连续生产、时刻满负荷。
                    max_h_oil = int(avail_gt // seg_load)
                    max_h_win = int(free_h) if free_h < 24.0 - 1e-9 else 999
                    need_h = int(math.ceil(remaining_q / seg_load - 1e-9))
                    actual_h = min(max_h_oil, max_h_win, need_h)
                    take = actual_h * seg_load if actual_h > 0 else 0.0
                    if take > remaining_q + 1e-6:
                        actual_h -= 1
                        take = actual_h * seg_load if actual_h > 0 else 0.0
                    if free_h < 24.0 - 1e-9 and actual_h > 0 and actual_h >= max_h_win:
                        win_short = True
                else:
                    take = 0.0; actual_h = 0
                used.append(gt)
                if take <= 1e-6:
                    continue                                                   # 此罐供不了整小时→换下一罐
                tanks[gt].crude = c; tanks[gt].ton -= take; remaining_q -= take
                t_hours = actual_h                                             # 整数加工时间
                gd = grid[gt][day]
                gd["crude"] = c
                gd["time"] = round(gd.get("time", 0.0) + t_hours, 1)
                gd["proc"] = gd.get("proc", 0.0) + take
                gd["load"] = round(gd["proc"] / gd["time"]) if gd["time"] > 1e-6 else 0
                # 逐段明细：单段罐 proc/time 即真实配方负荷；跨批混合罐由此明细挂批注（write_back）
                gd.setdefault("feeds", []).append({"crude": c, "load": round(seg_load),
                                                   "hours": t_hours, "tons": take})
                # 供料片段：同 (罐,油,日) 多段合并为一条（min起/max止/累加吨），保持下游合并假设
                fpk = (gt, c, day)
                fp = fp_by_key.get(fpk)
                if fp is None:
                    fp = {"g_tank": gt, "crude": c, "day": day,
                          "start_h": seg["seg_start"], "end_h": seg["seg_end"], "tons": take}
                    fp_by_key[fpk] = fp; feed_pieces.append(fp)
                else:
                    fp["start_h"] = min(fp["start_h"], seg["seg_start"])
                    fp["end_h"] = max(fp["end_h"], seg["seg_end"])
                    fp["tons"] += take
                if tanks[gt].ton - heel.get(gt, 0.0) < 1e-6:
                    charge.pop(c, None)
            if remaining_q > 1:
                if seg["tons"] - remaining_q < 1e-6 and no_tank:               # 完全没供上
                    warn.append(f"第{day}日 {c} 卸油停产窗口内无替代进料G罐，暂停加工"
                                if any(feed_free_h(t, day) <= 1e-6 for t in G)
                                else f"第{day}日 {c} 无进料G罐")
                else:
                    kind = "掺炼罐" if (used and used[-1] in P["BLEND_TANKS"]) else "供料罐"
                    short = "卸油窗口占用" if win_short else "供料期不受油"
                    warn.append(f"第{day}日 {c} 进料不足({kind}{short}，缺{remaining_q:.0f}吨)")
            if used:
                charge[c] = used[-1]                                           # 指向最后供料罐

        # ---- 加工后：按收油罐【当日实际供料工时】定 stop_feed_h → 时间窗口/到库量/停料 ----
        # 先供后卸：靠泊当日先供料（上面加工块，封顶 STOP_FEED_HOD），再据实际供料工时定停供时刻，
        # 之后卸油准备/卸油/供料准备。stop_feed_h = 当日0点 + min(STOP_FEED_HOD, 实际供料工时)。
        cap_h = (day - 1) * 24 + P.get("STOP_FEED_HOD", STOP_FEED_HOD_DEF)
        for C, splits, days_waited, pre_tank in todays_unloads:
            splits_sf = [(t, q, min(cap_h, (day - 1) * 24
                                    + grid[t].get(day, {}).get("time", 0.0)))
                         for t, q in splits]
            win = receiving_windows(splits_sf, P["UNLOAD_TPH"],
                                    P.get("UNLOAD_PREP_H", UNLOAD_PREP_H_DEF),
                                    P.get("FEED_PREP_H", FEED_PREP_H_DEF))
            label = f"等待{days_waited}天后接卸" if days_waited > 0 else "靠泊"
            for t, q in splits:
                pt = win["per_tank"][t]
                regrade_tank = should_regrade(q, tanks[t].ton, heel.get(t, 0.0))  # 卸前(=当日供料后)库存
                # 卸油混油(commingling)判定：卸前罐内为【异种】旧油且可用残余>换油残余阈值 SW。
                # 必须在下方 recv 循环【之前】取值——循环内会改写 tanks[t].crude / 累加 tanks[t].ton。
                resid_crude = tanks[t].crude                   # 卸前罐存油种(crude grade)，None=空罐/无标记
                resid_ton   = tanks[t].ton - heel.get(t, 0.0)  # 卸前可用残余（扣底油）
                unload_commingled = (resid_crude is not None and resid_crude != C
                                     and resid_ton > SW + 1e-9)
                if unload_commingled:
                    unload_commingle_count += 1
                recv_days = sorted(pt["recv_by_day"])
                for d in recv_days:
                    regrade_now = regrade_tank and d == recv_days[0]
                    if d == day:
                        # 靠泊当日到库（先供后卸）：加工后直接入罐
                        if regrade_now:
                            tanks[t].crude = C
                        elif tanks[t].crude is None:
                            tanks[t].crude = C
                        tanks[t].ton += pt["recv_by_day"][d]
                        grid[t][day]["recv"] = grid[t][day].get("recv", 0) + pt["recv_by_day"][d]
                        grid[t][day]["unload_recv"] = grid[t][day].get("unload_recv", 0) + pt["recv_by_day"][d]
                    else:
                        future_recv[d].append([t, C, pt["recv_by_day"][d], regrade_now])
                for d, hrs in pt["no_feed_hours"].items():
                    if d == day:
                        no_feed_hours[t][day] = hrs        # 覆盖 provisional，用实际窗口值
                    else:
                        no_feed_hours[t][d] = min(24.0, no_feed_hours[t].get(d, 0.0) + hrs)
                clog.append(f"{C}@第{day}日{label}→{t}[{q:.0f}t]"
                            + (f"（卸油混油：卸前{resid_crude}残余{resid_ton:.0f}t）"
                               if unload_commingled else ""))
        # 靠泊当日到库后再查一次容量
        for t in P["GTANKS"] + P["TTANKS"]:
            if tanks[t].ton > CAP[t] + 1:
                warn.append(f"第{day}日 {t} 卸船后超罐容{CAP[t]:.0f}（当前{tanks[t].ton:.0f}）")

        # 灰块：当天处于停料准备/卸油/供料准备的收油罐，记【当天实际】不供料小时（边界天取部分）
        for t in G:
            h = no_feed_hours.get(t, {}).get(day, 0.0)
            if h > 1e-6:
                grid[t][day]["no_feed_h"] = round(h, 1)

        # 库存口径：当日0点可用库存 = 0点物理 − 底油（=前一日24点可用），可为负
        for t in G + T:
            grid[t][day].setdefault("crude", tanks[t].crude)
            grid[t][day]["inv"] = round(day_start_phys[t] - heel.get(t, 0.0), 1)

    # ---- T→G 输油管道排期（仿真后独立 pass）----
    # 合并供料片段为供料段：同 (G罐, 油种) 连续日合并
    feed_segments = []
    by_key = defaultdict(list)
    for fp in feed_pieces:
        by_key[(fp["g_tank"], fp["crude"])].append(fp)
    for (g_tank, crude), pieces in by_key.items():
        pieces.sort(key=lambda p: p["day"])
        cur = None
        for p in pieces:
            if cur is not None and p["day"] == cur["_last_day"] + 1:
                # feed_start_h 保持首日不变，仅延伸结束时刻与累加吨数
                cur["feed_end_h"] = p["end_h"]
                cur["tons"] += p["tons"]
                cur["_last_day"] = p["day"]
            else:
                if cur is not None:
                    feed_segments.append(cur)
                cur = {"g_tank": g_tank, "crude": crude,
                       "feed_start_h": p["start_h"], "feed_end_h": p["end_h"],
                       "tons": p["tons"], "_last_day": p["day"]}
        if cur is not None:
            feed_segments.append(cur)
    for fs in feed_segments:
        fs.pop("_last_day", None)

    # 供料窗（按 G 罐：任何油种的供料都占用该罐，供避"边供边受"前移）
    fseg_by_g = defaultdict(list)
    for fs in feed_segments:
        fseg_by_g[fs["g_tank"]].append((fs["feed_start_h"], fs["feed_end_h"]))
    # 卸船窗（按 T 罐，逐日：当日有船卸入即整日占用，供避"边卸边输"前移）
    unload_windows = defaultdict(list)
    for t in P["TTANKS"]:
        for day, cell in grid[t].items():
            if cell.get("unload_recv", 0) > 1e-6:
                unload_windows[t].append(((day - 1) * 24.0, day * 24.0))

    sessions, w_build = build_sessions(raw_moves, feed_segments)
    scheduled, w_sched = schedule_transfers(sessions, feed_windows=fseg_by_g,
                                            unload_windows=unload_windows)
    warn.extend(w_build); warn.extend(w_sched)

    # 「受油不供料」违规扫描（小时级真重叠）：某 G 罐的 T→G 输油窗口与该罐供料窗口
    # 在小时上真重叠 = 边供边受。约束感知前移已尽量避让；剩余多为真·无间隙的高负荷
    # 连续供料罐（floor 紧贴 deadline、无空闲时段可避）→ 显式告警标注供人工审阅。
    for s in scheduled:
        g = s["dst_g"]
        for (fa, fb) in fseg_by_g.get(g, []):
            lo = max(s["start_h"], fa); hi = min(s["end_h"], fb)
            if hi - lo > 1e-6:   # 小时窗真重叠
                kind = "掺炼罐" if g in P["BLEND_TANKS"] else "主力罐"
                warn.append(
                    f"第{int(s['start_h']//24)+1}日 {kind}{g} 供料期受油{s['tons']:.0f}吨"
                    f"（{s['src_t']}→{g} {s['crude']} 与供料窗口重叠{hi-lo:.1f}h，边供边受）"
                    + ("：高负荷连续供料单管道物理受限，需人工审阅" if kind == "主力罐" else ""))
                break
    # 「边卸边输」违规扫描：某 T 储罐的 T→G 输油窗口与该罐卸船窗口小时级真重叠
    # = 同时接船卸油 + 向 G 罐输油（单管道/计量冲突）。约束感知前移已尽量避让，残留告警供人工审阅。
    for s in scheduled:
        for (ua, ub) in unload_windows.get(s["src_t"], []):
            lo = max(s["start_h"], ua); hi = min(s["end_h"], ub)
            if hi - lo > 1e-6:
                warn.append(
                    f"第{int(s['start_h']//24)+1}日 储罐{s['src_t']}边卸边输"
                    f"（{s['src_t']}→{s['dst_g']} {s['crude']} 输油{s['tons']:.0f}吨"
                    f"与卸船窗口重叠{hi-lo:.1f}h），需人工审阅")
                break
    for s in sorted(scheduled, key=lambda s: s["start_h"]):
        clog.append(f"{s['src_t']}→{s['dst_g']} "
                    f"{s['start_h']/24:.2f}d~{s['end_h']/24:.2f}d "
                    f"输入{s['crude']} {s['tons']:.0f}吨")

    # T→G 每日输油量按源 T 罐回填 grid（写回时以负值填入"到库量"行）
    for t, by_day in transfer_outflow_by_day(scheduled, DAYS).items():
        for day, tons in by_day.items():
            grid[t][day]["transfer_out"] = grid[t][day].get("transfer_out", 0) + tons
    # T→G 每日到库量按目标 G 罐回填 grid（正值填入 G 罐"到库量"行，与 T 罐输油逐日镜像）
    spread_in = transfer_inflow_by_day(scheduled, DAYS)
    for g, by_day in spread_in.items():
        for day, tons in by_day.items():
            grid[g][day]["recv"] = grid[g][day].get("recv", 0) + tons

    # inv 对齐：仿真按"瞬时整批"在 transfer_day 一次性搬运，期末库存因此跳变；
    # 此处只用"摊开口径 - 瞬时口径"的累计差微调 inv，使其与逐日到库/输油一致，
    # 其余流量（船卸、加工）沿用已验证仿真不动。窗口结束后差值归零。
    spread_out = transfer_outflow_by_day(scheduled, DAYS)
    inst_out = defaultdict(lambda: defaultdict(float))  # {src_t:{transfer_day:tons}}
    inst_in = defaultdict(lambda: defaultdict(float))   # {dst_g:{transfer_day:tons}}
    for s in scheduled:
        td = int(s["trigger_h"] // 24) + 1
        inst_out[s["src_t"]][td] += s["tons"]
        inst_in[s["dst_g"]][td] += s["tons"]

    def _align_inv(tank, inst_by_td, spread_by_day, sign):
        # 源罐流出 sign=-1，目标罐流入 sign=+1；inv_new = inv_old + sign*(累计摊开 - 累计瞬时)
        cum_inst = cum_spread = 0.0
        for d in range(1, DAYS + 1):
            cell = grid.get(tank, {}).get(d)
            if cell is not None and "inv" in cell:
                cell["inv"] = round(cell["inv"] + sign * (cum_spread - cum_inst), 1)
            cum_inst += inst_by_td.get(d, 0.0)
            cum_spread += spread_by_day.get(d, 0.0)

    for t, by_day in spread_out.items():
        _align_inv(t, inst_out[t], by_day, -1)
    for g, by_day in spread_in.items():
        _align_inv(g, inst_in[g], by_day, +1)

    clog.append({"metrics": {"unload_to_t_count": unload_to_t_count,
                             "unload_commingle_count": unload_commingle_count}})
    return grid, warn, clog

# ============================================================
# 4.5  多轮调度评分
# ============================================================
def count_compliant(P, plan, real, tol=1):
    """达标油种数（逐油种计数，对齐 CP-SAT 可调剂约束）。

    - 进口油（ORIGIN==imp）：各自达标——plan 与 real 各自 ≥ 自身需求(留容差)，互不调剂。
    - 计划国内油（非进口、非 EXTRA_DOMESTIC）：不逐个判，只看全局总加工量是否达小计(D列)。
      plan 层与 real 层总量都达标 → 计划国内油全部记达标；否则全部不达标。
      无 PROC_TOTAL 时退回逐油种严格判定（兜底，旧口径）。
    - 非计划库存油（EXTRA_DOMESTIC）：产量计入总量供给侧，但不参与达标计数（不判达标）。
    """
    proc = P["PROC"]
    origin = P.get("ORIGIN", {})
    proc_total = P.get("PROC_TOTAL")
    extra = P.get("EXTRA_DOMESTIC", set())

    # 计划国内油（排除非计划库存油）
    dom_crudes = [c for c in proc if origin.get(c) != "imp" and c not in extra]
    imp_compliant = sum(1 for c, req in proc.items()
                        if origin.get(c) == "imp"
                        and plan.get(c, 0) >= req - tol
                        and real.get(c, 0) >= req - tol)

    if proc_total and proc_total > 0:
        # 总量检查包含 EXTRA_DOMESTIC 产量（sum 覆盖全部 plan/real 键）
        dom_ok = (sum(plan.values()) >= proc_total - tol
                  and sum(real.values()) >= proc_total - tol)
        dom_compliant = len(dom_crudes) if dom_ok else 0
    else:
        dom_compliant = sum(1 for c in dom_crudes
                            if plan.get(c, 0) >= proc[c] - tol
                            and real.get(c, 0) >= proc[c] - tol)

    return imp_compliant + dom_compliant


def _transfer_flow_by_day(scheduled, days, key):
    """把已排期的 T→G 输油会话按小时比例分摊到每日，按 key 字段汇总输油量。

    scheduled: list[{src_t, dst_g, start_h, end_h, tons, ...}]（schedule_transfers 输出）
    days: 计划天数；第 d 天覆盖小时窗 [(d-1)*24, d*24]，越界部分截断。
    key: "src_t"（源罐流出）或 "dst_g"（目标罐流入）。
    返回 {tank: {day: tons}}，仅含有输油量的 (罐, 日)。
    """
    out = defaultdict(lambda: defaultdict(float))
    for s in scheduled:
        start_h, end_h, tons = s["start_h"], s["end_h"], s["tons"]
        dur = end_h - start_h
        if dur <= 1e-9 or tons <= 1e-9:
            continue
        t = s[key]
        d = int(start_h // 24) + 1
        while d <= days and (d - 1) * 24 < end_h - 1e-9:
            lo = max(start_h, (d - 1) * 24)
            hi = min(end_h, d * 24)
            ov = hi - lo
            if ov > 1e-9:
                out[t][d] += tons * ov / dur
            d += 1
    return {t: dict(days_) for t, days_ in out.items()}


def transfer_outflow_by_day(scheduled, days):
    """按源 T 罐汇总每日输油量（写回时以负值填入 T 罐"到库量"行）。返回 {src_t: {day: tons}}。"""
    return _transfer_flow_by_day(scheduled, days, "src_t")


def transfer_inflow_by_day(scheduled, days):
    """按目标 G 罐汇总每日到库量（写回时以正值填入 G 罐"到库量"行，与 T 罐输油逐日镜像）。
    返回 {dst_g: {day: tons}}。"""
    return _transfer_flow_by_day(scheduled, days, "dst_g")


def _latest_fit(dur_h, lo, hi, blocked):
    """在 [lo, hi] 内找一段长 dur_h、右端尽量靠 hi、且不与任何 blocked 区间相交的
    放置 [start, end]。找不到（区间被占满到 lo 以下）返回 None。
    blocked: list[(a, b)]，闭开半区间，相交判定 start<b 且 a<end。
    贪心：从 end=hi 起，若 [end-dur,end] 撞到某些 blocked，则把 end 退到这些相交
    区间里最早的起点（必须在它之前输完），再重扫；end 单调左移，至多 len(blocked)+1 轮。
    """
    end = hi
    for _ in range(len(blocked) + 2):
        start = end - dur_h
        if start < lo - 1e-6:
            return None
        hits = [(a, b) for (a, b) in blocked if start < b - 1e-6 and a < end - 1e-6]
        if not hits:
            return (start, end)
        end = min(a for (a, b) in hits)   # 退到最早相交区间之前
    return None


def schedule_transfers(sessions, feed_windows=None, unload_windows=None):
    """单管道 T→G 输油逆向排期（右对齐 EDD + 约束感知前移避让）。
    sessions: list[dict]，每个含 dur_h, deadline, release_floor（src_t/dst_g/crude/tons 透传）。
    feed_windows:   {g_tank: [(start_h, end_h), ...]} 目标供料罐供料窗——避"边供边受"。
    unload_windows: {t_tank: [(start_h, end_h), ...]} 源储罐卸船窗——避"边卸边输"。
    就地写入 start_h/end_h，返回 (ordered_sessions, warnings)。
    语义：每段默认贴 deadline（最晚、最少提前囤油）；为避开本罐供料窗 / 源罐卸船窗 /
    其它会话占用的管道时段而整体前移；在 [release_floor, deadline] 内找不到能避让的
    空闲时段时，退回"只避管道争用"的放置（边供边受由事后扫描另行告警）；连管道都排不开
    → 告警并钳到 release_floor。同 deadline 的会话按输入顺序处理（stable sort）。
    """
    feed_windows = feed_windows or {}
    unload_windows = unload_windows or {}
    warnings = []
    ordered = sorted(sessions, key=lambda s: s["deadline"], reverse=True)
    occupied = []   # 已排会话占用的管道时段 [(start, end), ...]
    for s in ordered:
        hi, lo, dur = s["deadline"], s["release_floor"], s["dur_h"]
        avoid = (list(feed_windows.get(s["dst_g"], []))          # 目标G供料窗：边供边受
                 + list(feed_windows.get(s["src_t"], []))         # 源G供料窗：G→G 源输出↔供料互斥(T源为空)
                 + list(unload_windows.get(s["src_t"], [])))      # 源罐卸船窗：边卸边输
        slot = _latest_fit(dur, lo, hi, occupied + avoid)
        if slot is None:
            # 避不开供料/卸船窗 → 退回只避管道争用（保持旧右对齐行为）
            slot = _latest_fit(dur, lo, hi, occupied)
        if slot is None:
            # 连管道都排不开 → 钳到 release_floor 下界并告警
            start = s["release_floor"]
            end = start + dur
            short_h = end - min(hi, occupied[-1][0] if occupied else hi)
            warnings.append(
                f"管道排不开：{s['crude']} 需在 {hi/24:.2f}d 前输完，"
                f"缺约 {max(0.0, short_h):.1f}h（源油未就绪/目标罐未空闲/管道占满）")
            slot = (start, end)
        s["start_h"], s["end_h"] = slot
        occupied.append((s["start_h"], s["end_h"]))
    return ordered, warnings


def build_sessions(raw_moves, feed_segments):
    """把仿真埋点的原始输油动作转成排期会话。
    raw_moves: list[{src_t,dst_g,crude,tons,trigger_h,oil_avail_h,tph}]（每源罐一条；
      `src_t` 可为 T 罐(T→G)或 G 罐(G→G)，`tph` 为该动作管速，共用同一根管道）
    feed_segments: list[{g_tank,crude,feed_start_h,feed_end_h,tons}]
    返回 (sessions, warnings)；tons≈0 的动作跳过。
    deadline = 该 (dst_g, crude) 供料段中 feed_start_h ≥ trigger_h 的最早一段；
    找不到则回退为 trigger_h 并告警。
    release_floor = max(源油可用时刻, 目标 G 罐在 deadline 前最近一次供料结束时刻)。
    注：deadline 按 (dst_g, crude) 匹配；floor_tank 只按 g_tank 匹配（不限油种），因为任何油种的供料都占着该 G 罐。
    """
    sessions = []
    warns = []
    for m in raw_moves:
        if m["tons"] <= 1e-6:
            continue
        # 触发时刻落入「进行中供料段」→ 中途补料：油在触发当日即被消耗，
        # deadline 取触发时刻（尽早送达），不能指派给下一批次开工时刻，
        # 否则管道把补料排到十几天后、对齐后该 G 罐可用库存转负（吃底油）。
        ongoing = [fs for fs in feed_segments
                   if fs["g_tank"] == m["dst_g"] and fs["crude"] == m["crude"]
                   and fs["feed_start_h"] - 1e-6 <= m["trigger_h"] <= fs["feed_end_h"] + 1e-6]
        if ongoing:
            deadline = m["trigger_h"]
        else:
            cands = [fs["feed_start_h"] for fs in feed_segments
                     if fs["g_tank"] == m["dst_g"] and fs["crude"] == m["crude"]
                     and fs["feed_start_h"] >= m["trigger_h"] - 1e-6]
            if cands:
                deadline = min(cands)
            else:
                deadline = m["trigger_h"]
                warns.append(
                    f"输油动作 {m['src_t']}→{m['dst_g']} {m['crude']} "
                    f"未找到对应供二常供料段，deadline 回退为触发时刻")
        prev_ends = [fs["feed_end_h"] for fs in feed_segments
                     if fs["g_tank"] == m["dst_g"]
                     and fs["feed_end_h"] <= deadline + 1e-6]
        floor_tank = max(prev_ends) if prev_ends else 0.0
        release_floor = max(m["oil_avail_h"], floor_tank)
        sessions.append({
            "src_t": m["src_t"], "dst_g": m["dst_g"], "crude": m["crude"],
            "tons": m["tons"], "dur_h": m["tons"] / m.get("tph", 700),
            "deadline": deadline, "release_floor": release_floor,
            "trigger_h": m["trigger_h"],   # 瞬时整批移动发生日 = transfer_day，供 inv 对齐
        })
    return sessions, warns


def score_solution(P, seq, grid, warn, clog):
    """计算各评分维度：两层达标油种数、批次数、主力罐换油次数、卸油落T罐船次、
    卸油混油(commingling)次数、实际加工总量。
    注：真正参与排序的仅 score_key 中 4 项（达标数 > 实际加工总量 > 批次数少 > 混油次数少）；
    main_switches / unload_to_t_count 只算不排，供展示与诊断。
    （2026-07-05 已移除单批次优选加工时长占比 dur_pct，原 docstring 对应描述已过期。）"""
    proc = P["PROC"]
    main_tanks_in_use = [t for t in P["GTANKS"] if t in P["MAIN_TANKS"]]

    plan = defaultdict(float)
    for b in seq:
        for c in b["u"]["comps"]:
            plan[c["crude"]] += b["dur_h"] * c["load_hr"]

    real = defaultdict(float)
    for t, days in grid.items():
        for d, cell in days.items():
            if cell.get("proc") and cell.get("crude"):
                real[cell["crude"]] += cell["proc"]

    compliant = count_compliant(P, plan, real)

    # 注：优选时长评分(dur_pct)已移除（2026-07-05）——不再按批次时长落区间打分。

    main_switches = 0
    for t in main_tanks_in_use:
        prev = None
        for d in range(1, P["DAYS"] + 1):
            if grid[t][d].get("proc", 0) > 0:
                cur = grid[t][d].get("crude")
                if cur and cur != prev and prev is not None:
                    main_switches += 1
                if cur:
                    prev = cur

    metrics = next((x.get("metrics", {}) for x in clog if isinstance(x, dict)), {})
    unload_to_t_count = metrics.get("unload_to_t_count", 0)
    unload_commingle_count = metrics.get("unload_commingle_count", 0)

    real_total = round(sum(real.values()))

    return {
        "compliant_crudes": compliant,
        "n_batches": len(seq),
        "main_switches": main_switches,
        "unload_to_t_count": unload_to_t_count,
        "unload_commingle_count": unload_commingle_count,
        "real_total": real_total,
    }


def score_key(sc):
    """排序键（越大越好）：达标数 > 实际加工总量 > 批次数少 > 卸油混油次数少"""
    return (
        sc["compliant_crudes"],
        sc.get("real_total", 0),
        -sc["n_batches"],
        -sc.get("unload_commingle_count", 0),   # 第4权重：卸油混油(commingling)次数少优先
    )


def select_solutions(results, n_preferred):
    """从已跑轮次结果中挑选待输出解（已按评分降序）。
    results: list[(sc, seq, grid, warn, clog)]，sc 含 is_preferred。
    - 有优选解：返回全部优选解，按 score_key 降序（轮次循环已按 n_preferred 早停，无需截断）；
    - 无优选解：回退返回评分最高的前 n_preferred 个非优选解（results 不足则全给）；
    - results 空：返回 []。
    """
    if not results:
        return []
    preferred = [r for r in results if r[0].get("is_preferred")]
    if preferred:
        return sorted(preferred, key=lambda r: score_key(r[0]), reverse=True)
    return sorted(results, key=lambda r: score_key(r[0]), reverse=True)[:n_preferred]


def make_seeds(n):
    """生成 n 个确定性、互异的随机种子（多轮调度用）。
    前 8 个为经验种子，超出部分按固定步长扩展，保证可复现且前缀稳定。"""
    base = [42, 137, 256, 512, 1000, 2187, 3141, 9973]
    seeds = base[:n]
    k = 10007
    while len(seeds) < n:
        if k not in seeds:
            seeds.append(k)
        k += 7919
    return seeds


def collect_rounds(seeds, n_preferred, n_judged, round_runner):
    """逐轮跑调度，标记优选解并在收集到 n_preferred 个优选解时早停。
    round_runner(seed, i) 返回 (sc, seq, grid, warn, clog) 或 None(无解跳过)。
    返回所有已跑(非 None)轮次结果列表。"""
    results = []
    preferred = 0
    for i, seed in enumerate(seeds):
        res = round_runner(seed, i)
        if res is None:
            continue
        sc = res[0]
        sc["is_preferred"] = (sc["compliant_crudes"] == n_judged and sc.get("total_gap", 999999) <= 1)
        results.append(res)
        if sc["is_preferred"]:
            preferred += 1
            if preferred >= n_preferred:
                break
    return results


def estimate_batches(P):
    """按【连续供料上限】(供料期不中途补料 ⇒ 一段连续供料 ≤ 一罐可用容量)估算 CDU 加工批次【下限】。
    只累加【可单炼油种】(能做主料的)；掺炼专用组分并入主料运行、不单独成批。
        下限 = Σ_{可单炼C} max( ceil(V(C)/最大可用供料罐容), ceil(V(C)/(负荷×MAX_BATCH_H)) ) + 1
    末尾 +1 为全局缓冲批（吸收掺炼油搭批、腾容时序等带来的额外批次需求）。
    返回 (总下限:int, 分解:dict{crude:(V, C_charge, n, 说明)})。上层从此值起、逐步 +1(最多+2)找最小可行批次。"""
    rate_hr = P.get("RATE_HR", {})
    MAXH = P.get("MAX_BATCH_H", MAX_BATCH_H)
    total = 0; breakdown = {}
    for c, V in P.get("PROC", {}).items():
        if not V or V <= 0:
            continue
        if P.get("CAN_SINGLE", {}).get(c) is False:      # 掺炼专用组分 → 不单独成批
            breakdown[c] = (V, None, 0, "掺炼专用·并入主料")
            continue
        C_charge = charge_cap(P, c)                       # allow 兼容主力罐最大可用容量（共用 helper）
        rate = rate_hr.get(c, DEF_RATE_HR)
        n_cap = math.ceil(V / C_charge) if C_charge > 0 else 0
        n_len = math.ceil(V / (rate * MAXH)) if rate > 0 else 0
        n = max(n_cap, n_len)
        total += n
        breakdown[c] = (V, C_charge, n, "可单炼")
    total += 1                                            # 全局缓冲批（吸收掺炼搭批/腾容时序额外需求）
    return total, breakdown


def multi_round_schedule(P, n_rounds=None, time_limit=None):
    """多轮CP-SAT调度：收集优选解，达 N_PREFERRED 个或跑满 n_rounds 后停止。
    返回按评分降序的待输出解列表 [(sc, seq, grid, warn, clog), ...]。
    无优选解时回退为评分最优单解；全无可行解则抛 RuntimeError。
    """
    if n_rounds is None: n_rounds = P.get("N_ROUNDS", 8)
    if time_limit is None: time_limit = P.get("TIME_LIMIT", 240)
    n_preferred = P.get("N_PREFERRED", 3)

    P["_BATCH_WEIGHTS"], P["_BATCH_ORDER"] = greedy_batch_priority(P)   # order 供 warm-start 复用
    print("批次优先级(剩余批次预估递减序):")
    for c in sorted(P["_BATCH_WEIGHTS"], key=lambda c: -P["_BATCH_WEIGHTS"][c][0]):
        ws = P["_BATCH_WEIGHTS"][c]
        print(f"  {c:8s} 分母={P['DENOM'][c]:.0f} 批权重="
              f"{['%.2f' % w for w in ws]}")
    ship_assignments, cp_constraints, no_process_windows = preprocess_arrivals(P)
    print(f"Phase0: {len(ship_assignments)} 艘船已预分配，{len(cp_constraints)} 条腾容约束，"
          f"{len(no_process_windows)} 条卸油停产窗口")

    extra = P.get("EXTRA_DOMESTIC", set())
    n_judged = len(P["PROC"]) - len(extra)

    # 批次预估下限（连续供料上限驱动）：作 CP-SAT 批次硬上限的起点，逐步 +1（最多 +2）
    batch_est, bd = estimate_batches(P)
    P["_BATCH_ESTIMATE"] = batch_est
    print(f"批次预估下限={batch_est} 批（只累加可单炼油种·罐容驱动）:")
    for c, (V, Cch, n, tag) in bd.items():
        print(f"  {c:8s} 量={V:>8.0f} " +
              (f"罐容={Cch:>7.0f} → {n} 批" if Cch else f"({tag})"))

    def round_runner(seed, i, batch_cap=None, tl=None, stop_on_compliant=False):
        tl = tl if tl is not None else time_limit
        cap_tag = f" cap≤{batch_cap}" if batch_cap is not None else ""
        print(f"\n--- 第{i+1}轮 (seed={seed}{cap_tag} 时限{tl}s) ---", flush=True)
        try:
            t0 = time.perf_counter()
            seq, nb, obj, stt = cp_schedule(P, ship_assignments=ship_assignments,
                                            cp_constraints=cp_constraints,
                                            time_limit=tl, random_seed=seed,
                                            no_process_windows=no_process_windows,
                                            batch_cap=batch_cap,
                                            warm_start=P.get("WARM_START", True),
                                            hint_search=(i < P.get("HINT_SEARCH_ROUNDS", 0)),
                                            stop_on_compliant=stop_on_compliant)
            t_cp = time.perf_counter() - t0
        except RuntimeError as e:
            print(f"  CP-SAT 无解: {e}")
            return None
        cp_tag = "最优" if stt == cp_model.OPTIMAL else "达标即停/时限中止"
        t1 = time.perf_counter()
        grid, warn, clog = tank_fill(P, seq, ship_assignments=ship_assignments)
        t_fill = time.perf_counter() - t1
        sc = score_solution(P, seq, grid, warn, clog)
        sc.update({"seed": seed, "round": i + 1, "nb": nb, "obj": obj})
        print(f"  CP-SAT={t_cp:.1f}s({cp_tag})  罐仿真={t_fill:.2f}s"
              f"  批次数={nb}  两层达标={sc['compliant_crudes']}/{n_judged}"
              f"  实际加工={sc['real_total']:.0f}t"
              f"  主力罐换油={sc['main_switches']}"
              f"  卸油T罐={sc['unload_to_t_count']}")
        return (sc, seq, grid, warn, clog)

    # 增量上限阶梯（2026-07-15）：批次上限从预估下限起、逐档 +1（最多 +2）。
    # 紧档给短时限 + 少种子（下限常本就不可行，快速失败即放宽）；每轮达标即停加速；
    # 某档凑够优选解即【停止放宽】（低档=批次更少=评分更优，无需再跑更松档）。三档不成再回退取消上限。
    LADDER_TIMES = P.get("LADDER_TIME_LIMITS", [40, 90, 180])
    LADDER_SEEDS = P.get("LADDER_SEED_COUNTS", [3, 3, 3])
    START_OFF = P.get("LADDER_START_OFFSET", 0)
    MAX_OFF = P.get("LADDER_MAX_OFFSET", 2)
    all_results = []; hit_n = False
    for lvl, off in enumerate(range(START_OFF, MAX_OFF + 1)):
        cap = batch_est + off
        tl = LADDER_TIMES[min(lvl, len(LADDER_TIMES) - 1)]
        ns = LADDER_SEEDS[min(lvl, len(LADDER_SEEDS) - 1)]
        base = len(all_results)
        print(f"\n[上限阶梯] 档{lvl + 1}: cap≤{cap}（下限{batch_est}+{off}）时限{tl}s × {ns}种子；达标即停")
        lvl_res = collect_rounds(
            make_seeds(ns), n_preferred, n_judged,
            lambda sd, i, cap=cap, tl=tl, base=base:
                round_runner(sd, i + base, batch_cap=cap, tl=tl, stop_on_compliant=True))
        all_results.extend(lvl_res)
        if sum(1 for r in lvl_res if r[0].get("is_preferred")) >= n_preferred:
            print(f"  档{lvl + 1} 已凑够 {n_preferred} 个优选解 → 停止放宽上限")
            hit_n = True
            break
    # 阶梯三档内一个优选解都没有 → 回退取消批次上限再跑
    if not hit_n and not any(r[0].get("is_preferred") for r in all_results):
        print("阶梯内无优选解 → 回退取消批次上限")
        all_results.extend(collect_rounds(
            make_seeds(n_rounds), n_preferred, n_judged,
            lambda sd, i, base=len(all_results):
                round_runner(sd, i + base, batch_cap=None, tl=time_limit, stop_on_compliant=True)))
    results = all_results
    if not results:
        raise RuntimeError("所有轮次均无可行解")

    if extra:
        imp_crudes  = sorted(c for c in P["PROC"] if P.get("ORIGIN", {}).get(c) == "imp")
        dom_crudes  = sorted(c for c in P["PROC"] if P.get("ORIGIN", {}).get(c) != "imp" and c not in extra)
        extra_crudes = sorted(extra)
        print(f"  达标核对：进口油各自必达（{', '.join(imp_crudes)}）"
              f" / 国内油按总量小计判定（{', '.join(dom_crudes)}）")
        print(f"  非计划库存油（产量计入总量、不判达标）：{', '.join(extra_crudes)}")

    results_sorted = sorted(results, key=lambda r: score_key(r[0]), reverse=True)
    print("\n=== 多轮评分对比（★=优选解） ===")
    print(f"  {'轮':>2}  {'达标':>4}  {'实际加工':>10}  {'批次':>4}  {'主罐换油':>6}  "
          f"{'卸油T':>5}  {'优选':>4}")
    for sc, *_ in results_sorted:
        flag = "★" if sc.get("is_preferred") else ""
        print(f"  {sc['round']:>2}  {sc['compliant_crudes']:>4}/{n_judged}  "
              f"{sc['real_total']:>10.0f}  "
              f"{sc['nb']:>4}  "
              f"{sc['main_switches']:>6}  {sc['unload_to_t_count']:>5}  {flag:>4}")

    selected = select_solutions(results, n_preferred)
    n_found = sum(1 for r in selected if r[0].get("is_preferred"))
    if n_found:
        print(f"\n找到 {n_found} 个优选解（目标 {n_preferred} 个），各生成 1 个文件")
    else:
        print(f"\n未找到优选解，回退输出评分最优解 3 个")
    return selected


# ============================================================
# 5. 回填预排产 sheet
# ============================================================
def output_filename(base, idx):
    """优选解输出文件名：base_{idx}.xlsx（idx 从 1 起，按评分序）。"""
    return f"{base}_{idx}.xlsx"


def write_back(path, out, P, grid):
    wb = openpyxl.load_workbook(path)
    ws = wb[sched_sheet(wb)]
    r = 2; rowmap = {}
    while r <= ws.max_row:
        tag = ws.cell(r, 1).value
        if tag:
            t = tank_id(tag); rowmap[t] = r
            r += 6 if t.startswith("G") else 3
        else:
            r += 1

    for t, base in rowmap.items():
        isg = t.startswith("G")
        for day in range(1, P["DAYS"] + 1):
            col = 4 + day - 1
            cell = grid.get(t, {}).get(day, {})
            inv = cell.get("inv"); crude = cell.get("crude")
            if inv is not None: ws.cell(base, col).value = round(inv)
            if crude: ws.cell(base + 1, col).value = crude
            if isg:
                if cell.get("proc"):
                    lc = ws.cell(base + 2, col)
                    lc.value = cell["load"]
                    ws.cell(base + 3, col).value = cell["time"]
                    ws.cell(base + 4, col).value = round(cell["proc"])
                    # 跨批混合格（一天喂了>1个真实负荷段）：负荷格挂批注列出真实配方明细
                    feeds = cell.get("feeds", [])
                    if len({f["load"] for f in feeds}) > 1:
                        detail = " + ".join(f'{f["load"]}×{f["hours"]:g}h' for f in feeds)
                        lc.comment = Comment(detail, "crp")
                # 灰底只标【当天没有加工】的纯不供料日；当天有加工(哪怕只是早段几小时)则
                # 显示加工时间、不刷灰底（那几小时是供料加工、非不供料）。
                elif cell.get("no_feed_h"):
                    tc = ws.cell(base + 3, col)   # 「加工时间」行填不供料小时 + 灰底
                    tc.value = cell["no_feed_h"]
                    tc.fill = NO_FEED_FILL
                if cell.get("recv"):
                    rc = ws.cell(base + 5, col)
                    rc.value = round(cell["recv"])
                    if cell.get("unload_recv"):   # 油轮卸油到库量 → 红底白字粗体
                        rc.fill = UNLOAD_FILL; rc.font = UNLOAD_FONT
                elif cell.get("transfer_out"):    # G→G 源：本 G 罐向其它 G 罐输出（负值）
                    ws.cell(base + 5, col).value = -round(cell["transfer_out"])
            else:
                # T罐：到库量 = 到港卸入量(正)；当天无到港时填 T→G 输油量(负)
                if cell.get("recv"):
                    rc = ws.cell(base + 2, col)
                    rc.value = round(cell["recv"])
                    if cell.get("unload_recv"):   # 油轮卸油到库量 → 红底白字粗体
                        rc.fill = UNLOAD_FILL; rc.font = UNLOAD_FONT
                elif cell.get("transfer_out"):
                    ws.cell(base + 2, col).value = -round(cell["transfer_out"])

    wb.save(out)


def write_solutions(src, base, selected, P, write_fn=write_back):
    """把每个待输出解写成 base_{i}.xlsx（i 从 1，按 selected 顺序），返回路径列表。"""
    paths = []
    for i, (sc, seq, grid, warn, clog) in enumerate(selected, 1):
        out = output_filename(base, i)
        write_fn(src, out, P, grid)
        paths.append(out)
    return paths


# ============================================================
# Benders 分解：子问题 + 割平面生成 + 主循环
# ============================================================
def check_feasibility(P, seq, ship_assignments):
    """子问题：运行 tank_fill 仿真，检测每日缺口。
    返回 (is_feasible, gaps, grid, warn, clog)
    gaps: {crude: {day: gap_tons}}
    """
    grid, warn, clog = tank_fill(P, seq, ship_assignments=ship_assignments)
    day_draw = defaultdict(list)
    for b in seq:
        u = b["u"]; t = b["start_h"]
        remain = {c["crude"]: b["dur_h"] * c.get("load_hr", 0) for c in u["comps"]}
        cr_hr = {c["crude"]: c["load_hr"] for c in u["comps"]}
        while any(v > 1e-6 for v in remain.values()) and t < P["DAYS"] * 24:
            day = int(t // 24) + 1; hleft = 24 - (t % 24)
            for c in u["comps"]:
                k = c["crude"]
                if remain[k] <= 1e-6: continue
                q = min(remain[k], cr_hr[k] * hleft)
                day_draw[day].append((k, q, c["load_hr"]))
                remain[k] -= q
            t += hleft
    gaps = defaultdict(lambda: defaultdict(float))
    for day in range(1, P["DAYS"] + 1):
        agg = defaultdict(float)
        for c, q, lh in day_draw.get(day, []):
            agg[c] += q
        for c, planned in agg.items():
            actual = 0
            for t2, days2 in grid.items():
                cell = days2.get(day, {})
                if cell.get("crude") == c and cell.get("proc"):
                    actual += cell["proc"]
            gap = planned - actual
            if gap > 1:
                gaps[c][day] = gap
    is_feasible = all(len(d) == 0 for d in gaps.values())
    return is_feasible, {c: dict(d) for c, d in gaps.items()}, grid, warn, clog


def generate_benders_cuts(P, gaps):
    """从缺口生成 Benders 割平面。
    返回 [(crude, deadline_hour, max_cumul_tons)]
    """
    cuts = []
    init_by_c = defaultdict(float)
    for t, d in P["INIT"].items():
        c = d.get("crude")
        if c:
            init_by_c[c] += d.get("ton", 0.0) - P.get("HEEL", {}).get(t, 0.0)
    for c, day_gaps in gaps.items():
        if not day_gaps: continue
        first_gap_day = min(day_gaps.keys())
        deadline_h = first_gap_day * 24
        cumul = init_by_c.get(c, 0.0)
        for a in P["ARRIVALS"]:
            if a.get("berth_day") and a["crude"] == c and a["berth_day"] <= first_gap_day:
                cumul += a["ton"]
        cuts.append((c, deadline_h, int(round(cumul))))
        print(f"  Benders cut: {c} deadline=h{deadline_h} max={int(round(cumul))}t")
    return cuts


def fill_month_postprocess(P, seq, ship_assignments):
    """后处理：CP-SAT 出解后，补满月末空闲天数（d30-d31 等）。
    找到最后一批的结束时刻，如果距月末还有 ≥24h 空闲，
    从可用油种中选一个追加一个边批（EDGE_MIN_BATCH_H 小时）填到月末。"""
    if not seq:
        return seq
    H = P["DAYS"] * 24
    EDGE_MIN = P.get("EDGE_MIN_BATCH_H", 24)
    last_end = max(s["start_h"] + s["dur_h"] for s in seq)
    gap = H - last_end
    if gap < EDGE_MIN:
        return seq  # 间隙太小，放不下一个边批

    # 找一个有可用油量的油种来追加
    avail = available_by_crude(P)
    rate_hr = P.get("RATE_HR", {})
    best_crude = None
    best_surplus = 0
    for c, v in avail.items():
        # 已加工量
        processed = sum(s["dur_h"] * comp.get("load_hr", 0)
                        for s in seq for comp in s["u"]["comps"]
                        if comp["crude"] == c)
        surplus = v - processed
        if surplus > best_surplus:
            best_surplus = surplus
            best_crude = c

    if best_crude is None or best_surplus < EDGE_MIN * rate_hr.get(best_crude, 680):
        return seq  # 无足够油量追加

    # 追加一个边批到月末
    load = rate_hr.get(best_crude, 680)
    dur = min(gap, int(best_surplus / load))
    if dur < EDGE_MIN:
        dur = EDGE_MIN  # 至少 EDGE_MIN 小时
    st = H - dur  # 右对齐到月末
    new_batch = {
        "u": {"comps": [{"crude": best_crude, "load_hr": load, "is_main": True}],
              "is_blend": False, "rate": load * 24, "rate_hr": load, "uid": f"FILL_{best_crude}"},
        "start_h": st, "dur_h": dur, "si": 99}
    seq.append(new_batch)
    seq.sort(key=lambda x: x["start_h"])
    print(f"  Post-process: +1 batch {best_crude} {st}h~{H}h ({dur}h, {dur*load}t) to fill month end")
    return seq


def benders_solve(P, ship_assignments=None, cp_constraints=None,
                  no_process_windows=None, batch_cap=None,
                  max_iter=8, time_limit=90, warm_start=True, random_seed=42,
                  stop_on_compliant=True):
    """Benders 分解主循环。"""
    if ship_assignments is None: ship_assignments = {}
    if cp_constraints is None: cp_constraints = []
    if no_process_windows is None: no_process_windows = []
    benders_cuts = []
    best = None
    for iteration in range(max_iter):
        print(f"\n  Benders iter {iteration+1}/{max_iter} (cuts: {len(benders_cuts)})")
        t0 = time.perf_counter()
        try:
            seq, nb, obj, stt = cp_schedule(
                P, ship_assignments=ship_assignments, cp_constraints=cp_constraints,
                time_limit=time_limit, random_seed=random_seed,
                no_process_windows=no_process_windows, batch_cap=batch_cap,
                warm_start=warm_start, stop_on_compliant=stop_on_compliant,
                benders_cuts=benders_cuts)
        except RuntimeError as e:
            print(f"  Master infeasible: {e}")
            if benders_cuts:
                benders_cuts.pop(); continue
            break
        t_master = time.perf_counter() - t0
        print(f"  Master: {t_master:.1f}s nb={nb} obj={obj:.0f}")
        if not seq: break
        t1 = time.perf_counter()
        feasible, gaps, grid, warn, clog = check_feasibility(P, seq, ship_assignments)
        t_sub = time.perf_counter() - t1
        total_gap = sum(sum(d.values()) for d in gaps.values())
        print(f"  Sub: {t_sub:.2f}s feasible={feasible} gap={total_gap:.0f}t")
        sc = score_solution(P, seq, grid, warn, clog)
        if best is None or total_gap < best[0]:
            best = (total_gap, seq, grid, warn, clog, sc)
        if feasible:
            print(f"\n*** Benders converged! iter {iteration+1}, no gap ***")
            seq = fill_month_postprocess(P, seq, ship_assignments)
            # 重新仿真验证补满后的结果
            grid2, warn2, clog2 = tank_fill(P, seq, ship_assignments=ship_assignments)
            sc2 = score_solution(P, seq, grid2, warn2, clog2)
            return seq, len(seq), obj, grid2, warn2, clog2, sc2, iteration + 1
        new_cuts = generate_benders_cuts(P, gaps)
        if not new_cuts: break
        existing = set((c, h) for c, h, _ in benders_cuts)
        added = 0
        for c, h, m in new_cuts:
            if (c, h) not in existing:
                benders_cuts.append((c, h, m)); existing.add((c, h)); added += 1
        if added == 0:
            print("  Duplicate cut, stopping"); break
    if best is None:
        print("\nBenders 全部迭代失败（主问题不可行）")
        return [], 0, 0, {}, [], [], {}, max_iter
    print(f"\nBenders not fully converged, best gap={best[0]:.0f}t")
    seq = fill_month_postprocess(P, best[1], ship_assignments)
    grid2, warn2, clog2 = tank_fill(P, seq, ship_assignments=ship_assignments)
    sc2 = score_solution(P, seq, grid2, warn2, clog2)
    return seq, len(seq), 0, grid2, warn2, clog2, sc2, max_iter


def multi_round_benders(P, n_rounds=None, time_limit=None, max_iter=8):
    """自适应批次阶梯 Benders 调度。

    从预估下限起逐档放宽 batch_cap，每档用不同种子尝试：
      档1: cap=下限+0, 3种子×20s  — 快速试探，大概率不可行
      档2: cap=下限+1, 4种子×60s  — 可能有解
      档3: cap=下限+2, 6种子×90s  — 主力档
      档4: cap=下限+3, 4种子×90s  — 补充档（仅档3不够优选解时触发）
    自适应规则：
      - 某档全不可行 → 跳过剩余种子，直接进入下一档
      - 某档有优选解但不够 → 本档追加2种子再判断
      - 某档凑够 N_PREFERRED → 停止放宽（低档=批次少=评分更优）
      - 连续4档无优选解 → 回退取消 batch_cap
    """
    if n_rounds is None: n_rounds = P.get("N_ROUNDS", 4)
    if time_limit is None: time_limit = P.get("TIME_LIMIT", 90)
    n_preferred = P.get("N_PREFERRED", 3)
    P["_BATCH_WEIGHTS"], P["_BATCH_ORDER"] = greedy_batch_priority(P)
    ship_assign, cp_constraints, no_proc_windows = preprocess_arrivals(P)
    batch_est, bd = estimate_batches(P)
    P["_G_INIT_AVAIL"] = g_init_availability(P)
    P["_ARRIVAL_CUM_SEGS"] = arrival_cum_segments(P)
    extra = P.get("EXTRA_DOMESTIC", set())
    n_judged = len(P["PROC"]) - len(extra)
    print(f"批次预估下限={batch_est} 批")

    MAX_OFF = 3  # 最多放宽到 下限+3
    LADDER_TIMES = P.get("LADDER_TIME_LIMITS", [20, 60, 90, 90])
    LADDER_SEEDS = P.get("LADDER_SEED_COUNTS", [3, 4, 6, 4])

    all_results = []; hit_n = False; round_counter = 0
    no_pref_streak = 0  # 连续无优选解的档数

    for lvl in range(MAX_OFF + 1):
        cap = batch_est + lvl
        tl = LADDER_TIMES[min(lvl, len(LADDER_TIMES) - 1)]
        ns = LADDER_SEEDS[min(lvl, len(LADDER_SEEDS) - 1)]
        lvl_seeds = make_seeds(ns + 2)  # 多取2个备用（自适应追加用）
        print(f"\n[阶梯] 档{lvl+1}: cap≤{cap}（下限{batch_est}+{lvl}）时限{tl}s × {ns}种子")

        lvl_feasible = 0; lvl_preferred = 0; si = 0
        for si in range(ns):
            seed = lvl_seeds[si]
            print(f"\n=== Round {round_counter+1} Benders (cap≤{cap}, seed={seed}) ===")
            try:
                seq, nb, obj, grid, warn, clog, sc, n_iter = benders_solve(
                    P, ship_assignments=ship_assign, cp_constraints=cp_constraints,
                    no_process_windows=no_proc_windows, batch_cap=cap,
                    max_iter=max_iter, time_limit=tl, warm_start=(round_counter==0),
                    random_seed=seed)
            except Exception as e:
                print(f"  Failed: {e}"); round_counter += 1; continue
            round_counter += 1
            if not seq: continue
            lvl_feasible += 1
            sc.update({"seed": seed, "round": round_counter, "nb": nb, "obj": obj})
            sc["is_preferred"] = (sc["compliant_crudes"] == n_judged and sc.get("total_gap", 999999) <= 1)
            total_gap = sum(sum(d.values()) for d in check_feasibility(P, seq, ship_assign)[1].values())
            sc["total_gap"] = total_gap
            all_results.append((sc, seq, grid, warn, clog))
            print(f"  Result: nb={nb} compliant={sc['compliant_crudes']}/{n_judged} "
                  f"real={sc.get('real_total',0):.0f}t gap={total_gap:.0f}t iters={n_iter}")
            if sc["is_preferred"]:
                lvl_preferred += 1
                if lvl_preferred >= n_preferred:
                    break

        # 自适应：有优选解但不够 → 追加2种子
        if 0 < lvl_preferred < n_preferred and si == ns - 1:
            print(f"  档{lvl+1} 有{lvl_preferred}个优选解(目标{n_preferred})，追加2种子...")
            for si2 in range(ns, ns + 2):
                if si2 >= len(lvl_seeds): break
                seed = lvl_seeds[si2]
                print(f"\n=== Round {round_counter+1} Benders (cap≤{cap}, seed={seed}, 追加) ===")
                try:
                    seq, nb, obj, grid, warn, clog, sc, n_iter = benders_solve(
                        P, ship_assignments=ship_assign, cp_constraints=cp_constraints,
                        no_process_windows=no_proc_windows, batch_cap=cap,
                        max_iter=max_iter, time_limit=tl, warm_start=False,
                        random_seed=seed)
                except Exception as e:
                    print(f"  Failed: {e}"); round_counter += 1; continue
                round_counter += 1
                if not seq: continue
                lvl_feasible += 1
                sc.update({"seed": seed, "round": round_counter, "nb": nb, "obj": obj})
                sc["is_preferred"] = (sc["compliant_crudes"] == n_judged and sc.get("total_gap", 999999) <= 1)
                total_gap = sum(sum(d.values()) for d in check_feasibility(P, seq, ship_assign)[1].values())
                sc["total_gap"] = total_gap
                all_results.append((sc, seq, grid, warn, clog))
                print(f"  Result: nb={nb} compliant={sc['compliant_crudes']}/{n_judged} "
                      f"real={sc.get('real_total',0):.0f}t gap={total_gap:.0f}t iters={n_iter}")
                if sc["is_preferred"]:
                    lvl_preferred += 1
                    if lvl_preferred >= n_preferred:
                        break

        # 自适应：全不可行 → 跳过剩余种子（已在 for 中自然跳过）
        if lvl_feasible == 0:
            print(f"  档{lvl+1} 全不可行 → 进入下一档")

        if lvl_preferred >= n_preferred:
            print(f"  档{lvl+1} 已凑够 {n_preferred} 个优选解 → 停止放宽上限")
            hit_n = True
            break

        if lvl_preferred > 0:
            no_pref_streak = 0
        else:
            no_pref_streak += 1

    # 阶梯全部无优选解 → 回退取消批次上限
    if not hit_n and not any(r[0].get("is_preferred") for r in all_results):
        print("\n阶梯内无优选解 → 回退取消批次上限")
        fallback_seeds = make_seeds(max(n_rounds, 4))
        for si, seed in enumerate(fallback_seeds):
            print(f"\n=== Round {round_counter+1} Benders (no cap, seed={seed}) ===")
            try:
                seq, nb, obj, grid, warn, clog, sc, n_iter = benders_solve(
                    P, ship_assignments=ship_assign, cp_constraints=cp_constraints,
                    no_process_windows=no_proc_windows, batch_cap=None,
                    max_iter=max_iter, time_limit=time_limit, warm_start=False,
                    random_seed=seed)
            except Exception as e:
                print(f"  Failed: {e}"); round_counter += 1; continue
            round_counter += 1
            if not seq: continue
            sc.update({"seed": seed, "round": round_counter, "nb": nb, "obj": obj})
            sc["is_preferred"] = (sc["compliant_crudes"] == n_judged and sc.get("total_gap", 999999) <= 1)
            total_gap = sum(sum(d.values()) for d in check_feasibility(P, seq, ship_assign)[1].values())
            sc["total_gap"] = total_gap
            all_results.append((sc, seq, grid, warn, clog))
            print(f"  Result: nb={nb} compliant={sc['compliant_crudes']}/{n_judged} "
                  f"real={sc.get('real_total',0):.0f}t gap={total_gap:.0f}t iters={n_iter}")
            if sc["is_preferred"]:
                if sum(1 for r in all_results if r[0].get("is_preferred")) >= n_preferred:
                    break

    results = all_results
    if not results: raise RuntimeError("No feasible solution")
    selected = select_solutions(results, n_preferred)
    print("\n=== 多轮评分对比 ===")
    print(f"  {'轮':>3}  {'种子':>6}  {'批次':>4}  {'达标':>4}  {'实际加工':>10}  {'缺口':>8}  {'优选':>4}")
    for sc, *_ in sorted(results, key=lambda r: score_key(r[0]), reverse=True):
        flag = "*" if sc.get("is_preferred") else ""
        print(f"  R{sc['round']:>2}  {sc.get('seed','?'):>6}  {sc['nb']:>4}  "
              f"{sc['compliant_crudes']:>4}/{n_judged}  "
              f"{sc.get('real_total',0):>10.0f}  {sc.get('total_gap',0):>8.0f}  {flag:>4}")
    return selected


# ============================================================
# main
# ============================================================
if __name__ == "__main__":
    P = read_all(SRC)
    print("油种加工量(吨):", {k: round(v) for k, v in P["PROC"].items()})
    print("G罐:", P["GTANKS"])
    print("T罐:", P["TTANKS"])
    print("掺炼配方:", list(P["RECIPES"].keys()) or "无")
    for rid, rc in P["RECIPES"].items():
        blend_str = ", ".join(f"{b['crude']}∈{b['cands']}t/h" for b in rc["blends"])
        print(f"  配方{rid}: 主力{rc['main_crude']}限额{rc['total_cap_hr']}t/h | 掺炼: {blend_str}")
    print()


    selected = multi_round_benders(P)

    BASE = "原油加工月计划_预排产结果_benders"
    paths = write_solutions(SRC, BASE, selected, P)

    # 控制台详情打印评分最优（_1）那一个解
    sc0, seq, grid, warn, clog = selected[0]

    print("\n最优加工序列:")
    for x in seq:
        u = x["u"]
        comp_str = "+".join(f"{c['crude']}({c['load_hr']}t/h)" for c in u["comps"])
        tag = "到港片" if x["si"] > 0 else "初始片"
        ton = x["dur_h"] * u["rate_hr"]
        print(f'  {x["start_h"]/24:5.2f}d~{(x["start_h"]+x["dur_h"])/24:5.2f}d'
              f'  {comp_str:30s}  {round(ton):6d}吨  [{tag}]')

    # 达标核对
    plan = defaultdict(float); real = defaultdict(float)
    for b in seq:
        for c in b["u"]["comps"]:
            plan[c["crude"]] += b["dur_h"] * c["load_hr"]
    for t, days in grid.items():
        for d, cell in days.items():
            if cell.get("proc"): real[cell["crude"]] = real[cell["crude"]] + cell["proc"]

    origin = P.get("ORIGIN", {})
    proc_total = P.get("PROC_TOTAL")
    print("\n达标核对（进口油各自必达 / 国内油按总量小计判定）")
    print("  油种 | 计划量(吨) ≥ 要求 | 罐实现量 | 要求 | 口径:")
    for c, req in P["PROC"].items():
        is_imp = origin.get(c) == "imp"
        f1 = "✓" if plan[c] >= req - 1 else "✗"
        f2 = "✓" if real[c] >= req - 1 else "✗"
        kind = "进口·各自" if is_imp else "国内·计总量"
        print(f'  {c:8s}  计划{plan[c]:9.0f}{f1}  实现{real[c]:9.0f}{f2}  要求{req:9.0f}  [{kind}]')
    if proc_total and proc_total > 0:
        tp, tr = sum(plan.values()), sum(real.values())
        tf1 = "✓" if tp >= proc_total - 1 else "✗"
        tf2 = "✓" if tr >= proc_total - 1 else "✗"
        print(f'  {"总加工量":8s}  计划{tp:9.0f}{tf1}  实现{tr:9.0f}{tf2}  小计{proc_total:9.0f}'
              f'  → 国内油{"全达标" if tf1=="✓" and tf2=="✓" else "不达标"}')

    str_clog = [x for x in clog if isinstance(x, str)]
    print("\n接卸/传输日志:", " | ".join(str_clog))
    print("\n告警:")
    [print("  -", w) for w in warn] or print("  (无)")
    print("\n已写出 ->", ", ".join(paths))
